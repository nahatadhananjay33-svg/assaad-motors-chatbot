"""
media_admin.py — Phase 8H owner-facing media + lifecycle management
===================================================================

The inventory/media MANAGEMENT layer. It is **additive** and does NOT touch the
chatbot's retrieval, pricing, Marathi, consultative, follow-up, logging, security,
or inventory-search behaviour. It only writes the parts of Excel the owner would
otherwise have to edit by hand, and the Supabase files the owner would otherwise
have to manage.

What it does
------------
* **Upload photos / videos** for a vehicle → store in Supabase → public URL →
  write into the existing per-slot media columns the chatbot already reads
  (EXTERIOR / VIDEO), and mirror into the consolidated `PHOTO_URLS` / `VIDEO_URLS`
  columns + bump `LAST_UPDATED` + set `MEDIA_FOLDER_ID`. (Steps 4 + 5)
* **Add Instagram / YouTube links** → per-slot INSTAGRAM/YOUTUBE columns + mirror.
* **List vehicles** → CAR NUMB, model, STATUS (for the admin panel). (Step 6)
* **Mark sold** → confirm → move the DNJ row to a `SOLD_CARS` archive sheet →
  delete the vehicle's Supabase files → clear media URLs → refresh inventory →
  the car disappears from the chatbot. (Step 7)

Design guarantees
-----------------
* Excel (DNJ) stays the ONLY source of truth; Supabase is media storage only.
* The chatbot keeps reading the SAME per-slot media columns — no loader change.
* All workbook writes are atomic (temp + os.replace) under a cross-process lock,
  and are skipped/deferred while a human has the file open in Excel.
* Storage is pluggable: `InMemoryMediaStore` (offline tests/validation) or
  `SupabaseMediaStore` (live). Switching is an env-var change only.
"""

from __future__ import annotations

import os
import re
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from media_loader_mapping import detect_media_layout, MediaType

# Reuse the proven concurrency/atomicity helpers from the media_sync package.
_MS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "media_sync")
import sys
if _MS_DIR not in sys.path:
    sys.path.insert(0, _MS_DIR)
from _util import (FileLock, atomic_save_workbook, excel_open_by_user,  # noqa: E402
                   utcnow_iso)

DNJ_SHEET = "DNJ"
SOLD_SHEET = "SOLD_CARS"
DEFAULT_BUCKET = os.getenv("CHAT_MEDIA_BUCKET", "car-photos")
_HEADER_SCAN_ROWS = 6

# allowed media
PHOTO_EXT = {"jpg", "jpeg", "png", "webp"}
VIDEO_EXT = {"mp4", "mov"}
_CONTENT_TYPE = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "webp": "image/webp", "mp4": "video/mp4", "mov": "video/quicktime",
}

# Phase 8H management column headers (final schema, Step 2).
COL_STATUS = "STATUS"
COL_PHOTO_URLS = "PHOTO_URLS"
COL_VIDEO_URLS = "VIDEO_URLS"
COL_YOUTUBE = "YOUTUBE_URL"
COL_INSTAGRAM = "INSTAGRAM_URL"
COL_FOLDER = "MEDIA_FOLDER_ID"
COL_LAST_UPDATED = "LAST_UPDATED"
COL_SOLD_AT = "SOLD_AT"
COL_SOLD_BY = "SOLD_BY"
MGMT_COLS = [COL_STATUS, COL_PHOTO_URLS, COL_VIDEO_URLS, COL_YOUTUBE,
             COL_INSTAGRAM, COL_FOLDER, COL_LAST_UPDATED]

STATUS_AVAILABLE = "AVAILABLE"
STATUS_RESERVED = "RESERVED"
STATUS_SOLD = "SOLD"

_URL_RE = re.compile(r"^https?://", re.IGNORECASE)
_HAS_DIGIT = re.compile(r"\d")
_MODEL_HEADERS = {"MODEL", "CAR"}


# ─────────────────────────────────────────────────────────────────────────────
# Pluggable media store (bytes in / public URL out) — offline-safe
# ─────────────────────────────────────────────────────────────────────────────
class InMemoryMediaStore:
    """Offline store used for tests/validation. Idempotent on object_path."""

    def __init__(self, project: str = "demo"):
        self.project = project
        self.objects: Dict[str, bytes] = {}

    def _url(self, bucket: str, object_path: str) -> str:
        return (f"https://{self.project}.supabase.co/storage/v1/object/public/"
                f"{bucket}/{object_path}")

    def upload(self, bucket: str, object_path: str, data: bytes, ctype: str) -> str:
        if not data:
            raise ValueError("empty file")
        self.objects[f"{bucket}/{object_path}"] = data       # idempotent overwrite
        return self._url(bucket, object_path)

    def list_paths(self, bucket: str, prefix: str = "") -> List[str]:
        pre = f"{bucket}/{prefix}"
        return sorted(k[len(bucket) + 1:] for k in self.objects if k.startswith(pre))

    def delete(self, bucket: str, object_path: str) -> bool:
        return self.objects.pop(f"{bucket}/{object_path}", None) is not None


class SupabaseMediaStore:  # pragma: no cover - exercised only against live Supabase
    """Live Storage adapter (Storage API only — never inventory/cars tables)."""

    def __init__(self, url: str, key: str):
        from supabase import create_client
        self._client = create_client(url, key)

    def upload(self, bucket: str, object_path: str, data: bytes, ctype: str) -> str:
        store = self._client.storage.from_(bucket)
        try:
            store.upload(object_path, data, {"content-type": ctype, "upsert": "true"})
        except Exception:
            pass                                              # already exists → idempotent
        return store.get_public_url(object_path)

    def list_paths(self, bucket: str, prefix: str = "") -> List[str]:
        store = self._client.storage.from_(bucket)
        out: List[str] = []
        folder = prefix.rstrip("/")
        for sub in ("photos", "videos", "exterior", "interior", "video",
                    "instagram", "youtube"):
            listed = store.list(f"{folder}/{sub}") or []
            out.extend(f"{folder}/{sub}/{x.get('name')}" for x in listed if x.get("name"))
        return out

    def delete(self, bucket: str, object_path: str) -> bool:
        self._client.storage.from_(bucket).remove([object_path])
        return True


def make_store():
    """Pick the live store when Supabase creds are present, else in-memory."""
    url, key = os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY")
    if url and key:
        return SupabaseMediaStore(url, key)
    return InMemoryMediaStore()


# ─────────────────────────────────────────────────────────────────────────────
# Workbook layout discovery
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class Layout:
    header_row: int
    car_col: int                                  # 1-based
    model_col: Optional[int]
    slots: Dict[str, List[int]] = field(default_factory=dict)   # media_type -> [cols]
    mgmt: Dict[str, int] = field(default_factory=dict)          # header -> col


def _norm_reg(value: Any) -> str:
    return ("" if value is None else str(value)).strip().upper().replace(" ", "")


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _discover_layout(ws) -> Optional[Layout]:
    """Find the header row (CAR NUMB) and map media + management columns.
    Returns None if no CAR NUMB column can be located."""
    max_col = ws.max_column or 0
    max_scan = min(ws.max_row or 1, _HEADER_SCAN_ROWS)
    for r in range(1, max_scan + 1):
        header = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        ml = detect_media_layout(header)
        if ml.car_col is None:
            continue
        slots: Dict[str, List[int]] = {}
        for mtype, entries in ml.groups.items():
            slots[mtype] = [col for _slot, col in sorted(entries)]
        model_col = None
        mgmt: Dict[str, int] = {}
        for c in range(1, max_col + 1):
            text = str(ws.cell(r, c).value or "").strip()
            up = text.upper()
            if model_col is None and up in _MODEL_HEADERS:
                model_col = c
            if up in {h.upper() for h in MGMT_COLS}:
                mgmt[up] = c
        return Layout(header_row=r, car_col=ml.car_col, model_col=model_col,
                      slots=slots, mgmt=mgmt)
    return None


def _ensure_mgmt_cols(ws, layout: Layout) -> None:
    """Create any missing Phase 8H management columns (append to the right)."""
    for header in MGMT_COLS:
        if header.upper() not in layout.mgmt:
            col = (ws.max_column or 0) + 1
            ws.cell(layout.header_row, col).value = header
            layout.mgmt[header.upper()] = col


def _index_rows(ws, layout: Layout) -> Dict[str, int]:
    """reg -> row index. Skips header/description rows (regs always contain a digit)."""
    rows: Dict[str, int] = {}
    for r in range(layout.header_row + 1, (ws.max_row or 0) + 1):
        reg = _norm_reg(ws.cell(r, layout.car_col).value)
        if reg and _HAS_DIGIT.search(reg):
            rows[reg] = r
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Result type
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class FileOutcome:
    filename: str
    status: str                 # uploaded | invalid_file | no_slot | upload_error
    url: Optional[str] = None
    detail: Optional[str] = None


# ─────────────────────────────────────────────────────────────────────────────
# The management service
# ─────────────────────────────────────────────────────────────────────────────
class MediaAdmin:
    """Owner-facing media + lifecycle operations over the inventory workbook."""

    def __init__(self, xlsx_path: str, *, store=None, bucket: str = DEFAULT_BUCKET,
                 sheet: str = DNJ_SHEET, backup_dir: Optional[str] = None,
                 lock_path: Optional[str] = None, clock=utcnow_iso):
        self.xlsx_path = xlsx_path
        self.store = store or make_store()
        self.bucket = bucket
        self.sheet = sheet
        self.backup_dir = backup_dir or os.path.join(
            os.path.dirname(os.path.abspath(xlsx_path)), "inventory_backups")
        self.lock_path = lock_path or (xlsx_path + ".lock")
        self.clock = clock

    # ── helpers ──
    def _open(self):
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.active
        layout = _discover_layout(ws)
        if layout is None:
            wb.close()
            raise ValueError("No CAR NUMB column found in the inventory sheet.")
        return wb, ws, layout

    @staticmethod
    def _slot_cols_for(layout: Layout, mtype: str) -> List[int]:
        return layout.slots.get(mtype, [])

    def _write_url_to_slot(self, ws, row: int, layout: Layout, mtype: str,
                           url: str) -> Tuple[bool, Optional[str]]:
        """Write url into the next empty per-slot cell of `mtype`.
        Returns (written, reason_if_not)."""
        cols = self._slot_cols_for(layout, mtype)
        if not cols:
            return False, "no_slot"
        existing = [ws.cell(row, c).value for c in cols]
        if url in [v for v in existing if v]:
            return False, "duplicate"
        empty = next((c for c in cols if not ws.cell(row, c).value), None)
        if empty is None:
            return False, "no_slot"
        ws.cell(row, empty).value = url
        return True, None

    def _refresh_mirror(self, ws, row: int, layout: Layout, reg: str, now: str) -> None:
        """Recompute the consolidated *_URLS mirror cells + folder + timestamp."""
        def joined(mtypes) -> str:
            urls: List[str] = []
            for mt in mtypes:
                for c in self._slot_cols_for(layout, mt):
                    v = ws.cell(row, c).value
                    if v and _URL_RE.match(str(v)):
                        urls.append(str(v))
            return "\n".join(urls)

        m = layout.mgmt
        ws.cell(row, m[COL_PHOTO_URLS.upper()]).value = joined(
            [MediaType.EXTERIOR_PHOTO, MediaType.INTERIOR_PHOTO]) or None
        ws.cell(row, m[COL_VIDEO_URLS.upper()]).value = joined([MediaType.VIDEO]) or None
        ws.cell(row, m[COL_YOUTUBE.upper()]).value = joined([MediaType.YOUTUBE]) or None
        ws.cell(row, m[COL_INSTAGRAM.upper()]).value = joined([MediaType.INSTAGRAM]) or None
        ws.cell(row, m[COL_FOLDER.upper()]).value = reg
        ws.cell(row, m[COL_LAST_UPDATED.upper()]).value = now
        # default a blank STATUS to AVAILABLE so the field is always explicit
        scol = m[COL_STATUS.upper()]
        if not ws.cell(row, scol).value:
            ws.cell(row, scol).value = STATUS_AVAILABLE

    def _guard_writable(self) -> Optional[Dict[str, Any]]:
        if excel_open_by_user(self.xlsx_path):
            return {"status": "deferred",
                    "detail": "The Excel file is open in Excel. Close it and retry."}
        return None

    # ── READ: list vehicles (Step 6) ──
    def list_vehicles(self) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.active
            layout = _discover_layout(ws)
            if layout is None:
                return []
            status_col = layout.mgmt.get(COL_STATUS.upper())
            car_i = layout.car_col - 1
            model_i = (layout.model_col - 1) if layout.model_col else None
            status_i = (status_col - 1) if status_col else None
            out: List[Dict[str, Any]] = []
            # Stream rows with iter_rows() instead of random ws.cell() access:
            # in read_only mode random cell access re-scans the sheet and is
            # O(n^2)-slow (~20s for this workbook); iter_rows is ~0.2s for the
            # identical result. Same extraction logic as before.
            for row in ws.iter_rows(min_row=layout.header_row + 1, values_only=True):
                car_val = row[car_i] if car_i < len(row) else None
                reg = _norm_reg(car_val)
                if not reg or not _HAS_DIGIT.search(reg):
                    continue
                model_val = row[model_i] if (model_i is not None and model_i < len(row)) else None
                model = str(model_val).strip() if model_val else ""
                status_val = row[status_i] if (status_i is not None and status_i < len(row)) else None
                status = str(status_val).strip().upper() if status_val else STATUS_AVAILABLE
                out.append({"car_number": reg, "model": model, "status": status})
            return out
        finally:
            wb.close()

    # ── WRITE: upload photos / videos (Steps 4 + 5) ──
    def upload_files(self, reg: str, files: List[Tuple[str, bytes]], *,
                     kind: str) -> Dict[str, Any]:
        """kind = 'photos' | 'videos'. files = list of (filename, bytes)."""
        if kind not in ("photos", "videos"):
            return {"status": "error", "detail": f"unknown kind '{kind}'"}
        reg = _norm_reg(reg)
        allowed = PHOTO_EXT if kind == "photos" else VIDEO_EXT
        slot_type = MediaType.EXTERIOR_PHOTO if kind == "photos" else MediaType.VIDEO
        now = self.clock()

        guard = self._guard_writable()
        if guard:
            return guard

        outcomes: List[FileOutcome] = []
        with FileLock(self.lock_path):
            wb, ws, layout = self._open()
            try:
                _ensure_mgmt_cols(ws, layout)
                rows = _index_rows(ws, layout)
                if reg not in rows:
                    wb.close()
                    return {"status": "rejected", "vehicle": reg,
                            "detail": "No matching CAR NUMB row in inventory.",
                            "uploaded": 0, "files": []}
                row = rows[reg]
                changed = False
                for filename, data in files:
                    ext = _ext(filename)
                    oc = FileOutcome(filename, "uploaded")
                    if ext not in allowed:
                        oc.status, oc.detail = "invalid_file", f"bad extension '.{ext}'"
                        outcomes.append(oc); continue
                    if not data:
                        oc.status, oc.detail = "invalid_file", "empty file"
                        outcomes.append(oc); continue
                    chash = hashlib.sha1(data).hexdigest()
                    object_path = f"{reg}/{kind}/{chash}.{ext}"
                    try:
                        url = self.store.upload(self.bucket, object_path, data,
                                                _CONTENT_TYPE.get(ext, "application/octet-stream"))
                    except Exception as e:
                        oc.status, oc.detail = "upload_error", str(e)
                        outcomes.append(oc); continue
                    oc.url = url
                    written, reason = self._write_url_to_slot(ws, row, layout, slot_type, url)
                    if written:
                        changed = True
                    elif reason == "no_slot":
                        oc.status, oc.detail = "no_slot", f"all {kind} slots full"
                    # duplicate => still report uploaded (url known), no slot change
                    outcomes.append(oc)

                if changed:
                    self._refresh_mirror(ws, row, layout, reg, now)
                    atomic_save_workbook(wb, self.xlsx_path, self.backup_dir, now)
                else:
                    wb.close()
            except Exception:
                wb.close()
                raise

        uploaded = sum(1 for o in outcomes if o.status in ("uploaded",))
        return {"status": "ok", "vehicle": reg, "kind": kind,
                "uploaded": sum(1 for o in outcomes if o.url),
                "written_to_sheet": uploaded,
                "updated_at": now,
                "files": [vars(o) for o in outcomes]}

    # ── WRITE: add Instagram / YouTube link ──
    def add_link(self, reg: str, platform: str, url: str) -> Dict[str, Any]:
        platform = (platform or "").strip().lower()
        if platform not in ("instagram", "youtube"):
            return {"status": "error", "detail": f"unknown platform '{platform}'"}
        if not url or not _URL_RE.match(url.strip()):
            return {"status": "rejected", "detail": "Link must be an http(s) URL."}
        reg = _norm_reg(reg)
        url = url.strip()
        slot_type = MediaType.INSTAGRAM if platform == "instagram" else MediaType.YOUTUBE
        now = self.clock()

        guard = self._guard_writable()
        if guard:
            return guard

        with FileLock(self.lock_path):
            wb, ws, layout = self._open()
            try:
                _ensure_mgmt_cols(ws, layout)
                rows = _index_rows(ws, layout)
                if reg not in rows:
                    wb.close()
                    return {"status": "rejected", "vehicle": reg,
                            "detail": "No matching CAR NUMB row in inventory."}
                row = rows[reg]
                written, reason = self._write_url_to_slot(ws, row, layout, slot_type, url)
                if not written and reason == "no_slot":
                    wb.close()
                    return {"status": "rejected", "vehicle": reg,
                            "detail": f"all {platform} slots are full."}
                if not written and reason == "duplicate":
                    wb.close()
                    return {"status": "ok", "vehicle": reg, "platform": platform,
                            "detail": "link already present", "updated_at": now}
                self._refresh_mirror(ws, row, layout, reg, now)
                atomic_save_workbook(wb, self.xlsx_path, self.backup_dir, now)
            except Exception:
                wb.close()
                raise
        return {"status": "ok", "vehicle": reg, "platform": platform,
                "url": url, "updated_at": now}

    # ── WRITE: mark sold (Step 7) ──
    def mark_sold(self, reg: str, *, confirm: bool, user: str = "owner") -> Dict[str, Any]:
        reg = _norm_reg(reg)
        if not confirm:
            return {"status": "confirm_required", "vehicle": reg,
                    "detail": "Re-send with confirm=true to mark this car sold."}
        now = self.clock()
        guard = self._guard_writable()
        if guard:
            return guard

        files_deleted = 0
        with FileLock(self.lock_path):
            wb, ws, layout = self._open()
            try:
                _ensure_mgmt_cols(ws, layout)
                rows = _index_rows(ws, layout)
                if reg not in rows:
                    wb.close()
                    return {"status": "rejected", "vehicle": reg,
                            "detail": "No matching CAR NUMB row in inventory."}
                row = rows[reg]

                # 1) delete every Supabase object under this vehicle's folder
                try:
                    for path in self.store.list_paths(self.bucket, prefix=f"{reg}/"):
                        if self.store.delete(self.bucket, path):
                            files_deleted += 1
                except Exception:
                    pass                                       # storage best-effort; Excel still moves

                # 2) snapshot the row. Clear ONLY the Supabase-backed photo/video
                #    slots (their files were just deleted, so those URLs would
                #    dangle) + the photo/video mirror columns. PRESERVE the external
                #    Instagram/YouTube URLs so they remain in the SOLD_CARS archive
                #    (they are not Supabase objects and are not deleted).
                max_col = ws.max_column
                values = [ws.cell(row, c).value for c in range(1, max_col + 1)]
                pv_slots = {c for mtype, cols in layout.slots.items() for c in cols
                            if mtype in (MediaType.EXTERIOR_PHOTO,
                                         MediaType.INTERIOR_PHOTO, MediaType.VIDEO)}
                pv_mirror = {layout.mgmt[h.upper()] for h in (COL_PHOTO_URLS, COL_VIDEO_URLS)
                             if h.upper() in layout.mgmt}
                for c in (pv_slots | pv_mirror):
                    if c - 1 < len(values):
                        values[c - 1] = None
                values[layout.mgmt[COL_STATUS.upper()] - 1] = STATUS_SOLD
                values[layout.mgmt[COL_LAST_UPDATED.upper()] - 1] = now

                # 3) ensure SOLD_CARS sheet + header, append the archived row
                self._archive_row(wb, ws, layout, values, now, user)

                # 4) remove the row from DNJ → vehicle leaves the chatbot's view
                ws.delete_rows(row, 1)

                atomic_save_workbook(wb, self.xlsx_path, self.backup_dir, now)
            except Exception:
                wb.close()
                raise

        return {"status": "ok", "vehicle": reg, "moved_to": SOLD_SHEET,
                "files_deleted": files_deleted, "sold_at": now, "sold_by": user}

    def _archive_row(self, wb, ws_src, layout: Layout, values: List[Any],
                     now: str, user: str) -> None:
        if SOLD_SHEET in wb.sheetnames:
            ws_sold = wb[SOLD_SHEET]
        else:
            ws_sold = wb.create_sheet(SOLD_SHEET)
            header = [ws_src.cell(layout.header_row, c).value
                      for c in range(1, ws_src.max_column + 1)]
            header += [COL_SOLD_AT, COL_SOLD_BY]
            ws_sold.append(header)
        ws_sold.append(list(values) + [now, user])

    # ── verification helper (Step 7 / 8): is the reg gone from live stock? ──
    def is_live(self, reg: str) -> bool:
        reg = _norm_reg(reg)
        return reg in {r for r in self.list_vehicles_regs()}

    def list_vehicles_regs(self) -> List[str]:
        return [v["car_number"] for v in self.list_vehicles()]


# ─────────────────────────────────────────────────────────────────────────────
# HTTP glue — used by chat_api.route (Step 6). Additive; mirrors inventory_upload.
# ─────────────────────────────────────────────────────────────────────────────
def parse_multipart(body: bytes, content_type: str) -> Tuple[Dict[str, str], List[Tuple[str, bytes]]]:
    """Return ({text_field: value}, [(filename, bytes), ...]) for a
    multipart/form-data body. File parts (with a filename) go to the list; plain
    fields go to the dict. Binary-safe."""
    m = re.search(r'boundary=(?:"([^"]+)"|([^;]+))', content_type or "", re.I)
    if not m:
        raise ValueError("missing multipart boundary")
    boundary = (m.group(1) or m.group(2)).strip()
    delim = b"--" + boundary.encode()
    fields: Dict[str, str] = {}
    files: List[Tuple[str, bytes]] = []
    for part in body.split(delim):
        if b"Content-Disposition" not in part:
            continue
        idx = part.find(b"\r\n\r\n")
        if idx == -1:
            continue
        headers = part[:idx].decode("utf-8", "replace")
        content = part[idx + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        name_m = re.search(r'name="([^"]*)"', headers)
        file_m = re.search(r'filename="([^"]*)"', headers)
        if file_m and file_m.group(1):
            files.append((file_m.group(1), content))
        elif name_m:
            fields[name_m.group(1)] = content.decode("utf-8", "replace").strip()
    return fields, files


def get_admin(service: Any) -> "MediaAdmin":
    """Return a per-service cached MediaAdmin (keeps an in-memory store alive
    for the life of the process when Supabase is not configured)."""
    admin = getattr(service, "_media_admin", None)
    if admin is None or admin.xlsx_path != service.xlsx_path:
        admin = MediaAdmin(service.xlsx_path)
        try:
            service._media_admin = admin
        except Exception:
            pass
    return admin


def handle_vehicles(service: Any) -> Tuple[int, Dict[str, Any]]:
    admin = get_admin(service)
    return 200, {"vehicles": admin.list_vehicles()}


def handle_upload(service: Any, body: bytes, content_type: str, *, kind: str) -> Tuple[int, Dict[str, Any]]:
    try:
        fields, files = parse_multipart(body, content_type)
    except ValueError as e:
        return 400, {"status": "rejected",
                     "detail": f"Upload must be multipart/form-data ({e})."}
    reg = fields.get("car_number") or fields.get("registration") or ""
    if not reg:
        return 400, {"status": "rejected", "detail": "Missing 'car_number' field."}
    if not files:
        return 400, {"status": "rejected", "detail": "No files in upload."}
    admin = get_admin(service)
    result = admin.upload_files(reg, files, kind=kind)
    if result.get("status") == "ok" and result.get("written_to_sheet"):
        result["refresh"] = service.refresh_inventory()
    code = 200 if result.get("status") in ("ok", "deferred") else 400
    return code, result


def handle_add_link(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    import json
    try:
        payload = json.loads(body.decode("utf-8") or "{}")
    except Exception:
        return 400, {"status": "rejected", "detail": "Body must be JSON."}
    reg = payload.get("car_number") or payload.get("registration") or ""
    platform = payload.get("platform") or ""
    url = payload.get("url") or ""
    admin = get_admin(service)
    result = admin.add_link(reg, platform, url)
    if result.get("status") == "ok":
        result["refresh"] = service.refresh_inventory()
    code = 200 if result.get("status") in ("ok", "deferred") else 400
    return code, result


def handle_mark_sold(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    import json
    try:
        payload = json.loads(body.decode("utf-8") or "{}")
    except Exception:
        return 400, {"status": "rejected", "detail": "Body must be JSON."}
    reg = payload.get("car_number") or payload.get("registration") or ""
    confirm = bool(payload.get("confirm"))
    user = payload.get("user") or "owner"
    if not reg:
        return 400, {"status": "rejected", "detail": "Missing 'car_number' field."}
    admin = get_admin(service)
    result = admin.mark_sold(reg, confirm=confirm, user=user)
    if result.get("status") == "ok":
        result["refresh"] = service.refresh_inventory()
        live = {i.registration_no for i in service.engine.all_facing}
        result["disappeared_from_chatbot"] = _norm_reg(reg) not in live
    code = 200 if result.get("status") in ("ok", "confirm_required", "deferred") else 400
    return code, result
