"""
media_admin_tests.py — Phase 8H unit/integration tests.

Covers the media management service (MediaAdmin) and the HTTP route layer
(chat_api.route) for the new /admin/media/* endpoints, plus the admin auth gate.
All offline: in-memory store + a temp test workbook. Never touches production
data or Supabase.
"""
from __future__ import annotations

import io
import json
import os
import tempfile

import pytest
import openpyxl

import media_admin as MA
from media_admin_validate import build_test_workbook
from chat_service import ChatService
import chat_api
from security import SecurityGate, RateLimiter


REG = "MH02EZ6001"


@pytest.fixture()
def wb_path():
    d = tempfile.mkdtemp(prefix="phase8h_t_")
    p = os.path.join(d, "test_inventory.xlsx")
    build_test_workbook(p)
    return p


@pytest.fixture()
def admin(wb_path):
    return MA.MediaAdmin(wb_path, store=MA.InMemoryMediaStore(), bucket="car-photos",
                         backup_dir=os.path.join(os.path.dirname(wb_path), "bk"))


# ── service-level ────────────────────────────────────────────────────────────
def test_list_vehicles(admin):
    v = admin.list_vehicles()
    assert len(v) == 3
    assert {x["car_number"] for x in v} == {REG, "MH12AB1234", "MH14CD5678"}
    assert all(x["status"] == "AVAILABLE" for x in v)


def test_upload_photos_writes_slots_and_mirror(admin, wb_path):
    r = admin.upload_files(REG, [("a.jpg", b"jpgdata"), ("b.png", b"pngdata")], kind="photos")
    assert r["status"] == "ok" and r["uploaded"] == 2
    wb = openpyxl.load_workbook(wb_path); ws = wb[MA.DNJ_SHEET]
    lay = MA._discover_layout(ws); row = MA._index_rows(ws, lay)[REG]
    photo_mirror = ws.cell(row, lay.mgmt["PHOTO_URLS"]).value
    assert photo_mirror.count("http") == 2
    assert ws.cell(row, lay.mgmt["MEDIA_FOLDER_ID"]).value == REG
    assert ws.cell(row, lay.mgmt["LAST_UPDATED"]).value
    wb.close()


def test_invalid_extension_rejected(admin):
    r = admin.upload_files(REG, [("x.txt", b"nope")], kind="photos")
    assert r["uploaded"] == 0
    assert r["files"][0]["status"] == "invalid_file"


def test_upload_unknown_vehicle(admin):
    r = admin.upload_files("ZZ00ZZ0000", [("a.jpg", b"x")], kind="photos")
    assert r["status"] == "rejected"


def test_add_links(admin):
    assert admin.add_link(REG, "instagram", "https://instagram.com/p/x")["status"] == "ok"
    assert admin.add_link(REG, "youtube", "https://youtu.be/x")["status"] == "ok"
    assert admin.add_link(REG, "youtube", "noturl")["status"] == "rejected"


def test_mark_sold_requires_confirm(admin):
    assert admin.mark_sold(REG, confirm=False)["status"] == "confirm_required"


def test_mark_sold_moves_deletes_and_hides(admin, wb_path):
    admin.upload_files(REG, [("a.jpg", b"x")], kind="photos")
    admin.upload_files(REG, [("v.mp4", b"y")], kind="videos")
    res = admin.mark_sold(REG, confirm=True)
    assert res["status"] == "ok"
    assert res["files_deleted"] == 2
    # gone from Supabase store
    assert admin.store.list_paths("car-photos", f"{REG}/") == []
    # moved to SOLD_CARS, removed from DNJ
    wb = openpyxl.load_workbook(wb_path)
    assert MA.SOLD_SHEET in wb.sheetnames
    wb.close()
    assert REG not in admin.list_vehicles_regs()


# ── HTTP route layer ─────────────────────────────────────────────────────────
@pytest.fixture()
def service(wb_path):
    svc = ChatService(xlsx_path=wb_path)
    svc._media_admin = MA.MediaAdmin(wb_path, store=MA.InMemoryMediaStore(),
                                     bucket="car-photos",
                                     backup_dir=os.path.join(os.path.dirname(wb_path), "bk"))
    yield svc
    svc.close()


def _multipart(car_number, files):
    boundary = "----phase8h"
    parts = []
    parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="car_number"\r\n\r\n{car_number}\r\n'.encode())
    for fn, data in files:
        parts.append(f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{fn}"\r\n'
                     f'Content-Type: application/octet-stream\r\n\r\n'.encode() + data + b"\r\n")
    parts.append(f'--{boundary}--\r\n'.encode())
    return b"".join(parts), f"multipart/form-data; boundary={boundary}"


def test_route_vehicles(service):
    status, payload = chat_api.route("GET", "/admin/media/vehicles", b"", service)
    assert status == 200 and len(payload["vehicles"]) == 3


def test_route_upload_photos_and_refresh(service):
    body, ct = _multipart(REG, [("a.jpg", b"jpg")])
    status, payload = chat_api.route("POST", "/admin/media/upload_photos", body, service, ct)
    assert status == 200 and payload["status"] == "ok"
    assert payload["uploaded"] == 1
    assert "refresh" in payload                      # chatbot was refreshed live


def test_route_add_link(service):
    body = json.dumps({"car_number": REG, "platform": "youtube", "url": "https://youtu.be/z"}).encode()
    status, payload = chat_api.route("POST", "/admin/media/add_link", body, service)
    assert status == 200 and payload["status"] == "ok"


def test_route_mark_sold_disappears(service):
    body = json.dumps({"car_number": REG, "confirm": True}).encode()
    status, payload = chat_api.route("POST", "/admin/media/mark_sold", body, service)
    assert status == 200 and payload["status"] == "ok"
    assert payload["disappeared_from_chatbot"] is True
    assert REG not in {i.registration_no for i in service.engine.all_facing}


# ── admin gate (fail-closed) ─────────────────────────────────────────────────
def test_media_endpoints_are_admin_gated():
    # no admin key configured → admin surface disabled (fail-closed)
    gate = SecurityGate(api_keys=set(), admin_keys=set(),
                        limiter=RateLimiter(0, 60), allowed_origins=["*"])
    denial = gate.authorize("GET", "/admin/media/vehicles", {}, "1.2.3.4")
    assert denial is not None and denial[0] in (401, 403)

    # with a key, wrong key is rejected, right key passes
    gate2 = SecurityGate(api_keys=set(), admin_keys={"secret"},
                         limiter=RateLimiter(0, 60), allowed_origins=["*"])
    assert gate2.authorize("POST", "/admin/media/mark_sold", {"X-API-Key": "wrong"}, "ip") is not None
    assert gate2.authorize("POST", "/admin/media/mark_sold", {"X-API-Key": "secret"}, "ip") is None
