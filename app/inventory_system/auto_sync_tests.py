"""
auto_sync_tests.py — Inventory <-> Chatbot auto-sync regression suite
=====================================================================

Verifies the production rule: **Excel is the single source of truth**, and the
chatbot always reflects the CURRENT inventory after a refresh — ADD / REMOVE /
MODIFY / DUPLICATE / NEW-MODEL / PINNED-safety — with NO code change and NO
restart.

Design:
  * Runs against a TEMP COPY of the live sheet (never touches real inventory).
  * Mutates the Excel then calls the SAME `service.refresh_inventory()` that the
    Owner/Staff save endpoints (inventory_edit / media_admin / upload) call —
    so this exercises the real sync path.
  * DATA-DRIVEN: every expected value is derived from the fixture or from the
    row we just wrote, and counts are asserted as deltas (N -> N+1 -> N). These
    tests never break just because the dealership's fleet changes size/content.
"""
import os
import shutil
import tempfile
import unittest

import openpyxl

import inventory_loader as L
from inventory_loader import COL, DATA_START_ROW, DNJ_SHEET
from chat_service import ChatService
import query_parser as QP

LIVE = os.path.join(os.path.dirname(__file__), "..", "IVR_Sheet.xlsx")

# core columns we write on ADD (a minimal, always-viewable, quotable car)
def _core(make, model, year, fuel, trans, owners, rate, km=None,
          color="White", location="A1"):
    d = {"make": make, "model": model, "year": year, "fuel": fuel,
         "transmission": trans, "ownership": owners, "rate": rate,
         "color": color, "location": location}
    if km is not None:
        d["km"] = km
    return d


@unittest.skipUnless(os.path.exists(LIVE), "live IVR_Sheet.xlsx not found")
class AutoSyncBase(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix="autosync_")
        self.xlsx = os.path.join(self.tmp, "inv.xlsx")
        shutil.copy2(LIVE, self.xlsx)
        self.svc = ChatService(xlsx_path=self.xlsx,
                               leads_db=os.path.join(self.tmp, "l.db"),
                               analytics_db=os.path.join(self.tmp, "a.db"),
                               unknown_db=os.path.join(self.tmp, "u.db"))

    def tearDown(self):
        try:
            self.svc.close()
        except Exception:
            pass
        shutil.rmtree(self.tmp, ignore_errors=True)

    # ── low-level Excel helpers ──────────────────────────────────────────────
    def _car_col(self):
        return COL["car_numb"] + 1

    def _find_row(self, ws, reg):
        cc = self._car_col()
        for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
            v = ws.cell(r, cc).value
            if v and str(v).strip().upper() == reg.upper():
                return r
        return None

    def _header_col(self, ws, header):
        for c in range(1, (ws.max_column or 0) + 1):
            if str(ws.cell(2, c).value or "").strip().lower() == header.lower():
                return c
        return None

    # ── inventory mutators (write Excel, then reload exactly like production) ──
    def _add(self, reg, core=None, headers=None):
        wb = openpyxl.load_workbook(self.xlsx)
        ws = wb[DNJ_SHEET]
        row = (ws.max_row or DATA_START_ROW - 1) + 1
        ws.cell(row, self._car_col(), reg)
        for k, v in (core or {}).items():
            ws.cell(row, COL[k] + 1, v)
        for h, v in (headers or {}).items():
            c = self._header_col(ws, h)
            if c:
                ws.cell(row, c, v)
        wb.save(self.xlsx)
        wb.close()
        self.svc.refresh_inventory()

    def _modify(self, reg, core=None, headers=None):
        wb = openpyxl.load_workbook(self.xlsx)
        ws = wb[DNJ_SHEET]
        row = self._find_row(ws, reg)
        self.assertIsNotNone(row, f"{reg} not found to modify")
        for k, v in (core or {}).items():
            ws.cell(row, COL[k] + 1, v)
        for h, v in (headers or {}).items():
            c = self._header_col(ws, h)
            self.assertIsNotNone(c, f"header {h!r} not present in sheet")
            ws.cell(row, c, v)
        wb.save(self.xlsx)
        wb.close()
        self.svc.refresh_inventory()

    def _remove(self, reg):
        wb = openpyxl.load_workbook(self.xlsx)
        ws = wb[DNJ_SHEET]
        row = self._find_row(ws, reg)
        self.assertIsNotNone(row, f"{reg} not found to remove")
        ws.delete_rows(row, 1)
        wb.save(self.xlsx)
        wb.close()
        self.svc.refresh_inventory()

    def _ask(self, msg, sid="s"):
        return self.svc.handle(msg, session_id=sid)

    def _n(self):
        return self.svc.inventory_count


# ─────────────────────────────────────────────────────────────────────────────
# ADD / REMOVE / MODIFY
# ─────────────────────────────────────────────────────────────────────────────
class TestAddRemoveModify(AutoSyncBase):
    def test_add_new_vehicle_is_discovered(self):
        n = self._n()
        reg = "MH01ZZ9001"
        self._add(reg, _core("TOYO", "Corolla Altis", 2020, "P", "A", 1, 900000, km=30000))
        self.assertEqual(self._n(), n + 1)                       # data-driven delta
        r = self._ask(f"{reg} available hai?")
        self.assertEqual(r.count, 1)
        self.assertIn("available", r.response.lower())

    def test_removed_vehicle_disappears(self):
        n = self._n()
        reg = "MH01ZZ9002"
        self._add(reg, _core("HOND", "City", 2019, "P", "M", 1, 700000))
        self.assertEqual(self._n(), n + 1)
        r = self._ask(f"{reg} available hai?")
        self.assertEqual(r.count, 1)                             # searchable while present
        self._remove(reg)
        self.assertEqual(self._n(), n)                           # back to N
        r = self._ask(f"{reg} available hai?")
        self.assertEqual(r.count, 0)                             # gone, not stale
        self.assertIn("nahi", r.response.lower())

    def test_modified_price_reflected(self):
        reg = "MH01ZZ9003"
        self._add(reg, _core("MARU", "Swift", 2018, "P", "M", 1, 500000))
        self.assertIn("5.00", self._ask(f"{reg} ka price?").response)
        self._modify(reg, {"rate": 650000})
        self.assertIn("6.50", self._ask(f"{reg} ka price?", "p2").response)

    def test_modified_km_reflected(self):
        reg = "MH01ZZ9004"
        self._add(reg, _core("MARU", "Swift", 2018, "P", "M", 1, 500000, km=40000))
        self.assertIn("40,000", self._ask(f"{reg} km kitna hai?", "k").response)
        self._modify(reg, {"km": 55000})
        self.assertIn("55,000", self._ask(f"{reg} km kitna hai?", "k2").response)

    def test_modified_owners_reflected(self):
        reg = "MH01ZZ9006"
        self._add(reg, _core("MARU", "Baleno", 2019, "P", "M", 1, 600000))
        self.assertIn("1", self._ask(f"{reg} kitne owner?", "o").response)
        self._modify(reg, {"ownership": 3})
        self.assertIn("3", self._ask(f"{reg} kitne owner?", "o2").response)

    def test_modified_spec_reflected(self):
        # extended spec (read by header). Owner value in Excel wins over any
        # model_specs default, and a later edit is reflected after refresh.
        reg = "MH01ZZ9005"
        self._add(reg, _core("TOYO", "Fortuner", 2019, "D", "A", 1, 2000000),
                  headers={"Airbags": 2})
        self.assertIn("2", self._ask(f"{reg} airbags kitne?", "a").response)
        self._modify(reg, headers={"Airbags": 7})
        self.assertIn("7", self._ask(f"{reg} airbags kitne?", "a2").response)


# ─────────────────────────────────────────────────────────────────────────────
# NEW MODEL / DUPLICATE MODEL (no code change)
# ─────────────────────────────────────────────────────────────────────────────
class TestNewAndDuplicateModel(AutoSyncBase):
    def test_completely_new_model_searchable_without_code_change(self):
        # a model NOT in the static market vocabulary is searchable purely from
        # Excel, because the recognizer is refreshed from the live inventory.
        reg = "MH01ZZ9010"
        novel = "Zephyrus"
        self._add(reg, _core("FORD", novel, 2017, "P", "M", 1, 400000))
        r = self._ask(f"{novel} available hai?", "nm")
        self.assertGreaterEqual(r.count, 1)
        self.assertEqual(r.vehicles[0]["registration_no"], reg)

    def test_duplicate_model_not_silently_picked(self):
        # two cars, same model, DIFFERING price -> must clarify / list, never
        # silently answer a single car's price.
        novel = "Duplomo"
        self._add("MH01ZZ9021", _core("MARU", novel, 2018, "P", "M", 1, 500000))
        self._add("MH01ZZ9022", _core("MARU", novel, 2020, "P", "M", 2, 700000))
        r = self._ask(f"{novel} price kya hai?", "d")
        self.assertNotEqual(r.count, 1, "must not silently pick one of two")
        self.assertTrue(r.status in ("multi", "clarify") or r.count >= 2)

    def test_duplicate_then_remove_one_resolves_single(self):
        novel = "Duplotwo"
        self._add("MH01ZZ9023", _core("MARU", novel, 2018, "P", "M", 1, 500000))
        self._add("MH01ZZ9024", _core("MARU", novel, 2020, "P", "M", 2, 700000))
        self.assertGreaterEqual(self._ask(f"{novel} hai kya?", "r1").count, 2)
        self._remove("MH01ZZ9024")
        r = self._ask(f"{novel} price kya hai?", "r2")
        self.assertEqual(r.count, 1)                             # now unambiguous
        self.assertIn("5.00", r.response)                        # remaining car's price


# ─────────────────────────────────────────────────────────────────────────────
# PINNED VEHICLE SAFETY (inventory changes mid-conversation)
# ─────────────────────────────────────────────────────────────────────────────
class TestPinnedVehicleSafety(AutoSyncBase):
    def test_pinned_then_removed_no_stale_answer(self):
        reg = "MH01ZZ9030"
        self._add(reg, _core("TOYO", "Innova", 2018, "D", "M", 1, 900000, km=60000))
        sid = "pin1"
        self.assertEqual(self._ask(f"{reg} available hai?", sid).count, 1)   # pinned
        self._remove(reg)
        r = self._ask("price kya hai?", sid)                     # continue same session
        self.assertNotIn("9.00", r.response, "answered the REMOVED car's price")
        self.assertNotIn(reg, str(r.vehicles), "returned the removed car")

    def test_pinned_removed_replaced_no_identity_transfer(self):
        # pin A; remove A; add B (same NOVEL model, different price); ask price.
        # The removed car's identity/price must not silently transfer to B.
        model = "Pinmodelx"
        regA, regB = "MH01ZZ9031", "MH01ZZ9032"
        self._add(regA, _core("TOYO", model, 2018, "D", "M", 1, 800000))
        sid = "pin2"
        self.assertEqual(self._ask(f"{regA} available hai?", sid).count, 1)
        self._remove(regA)
        self._add(regB, _core("TOYO", model, 2021, "D", "A", 1, 1500000))
        r = self._ask("price kya hai?", sid)
        self.assertNotIn(regA, str(r.vehicles), "claimed the removed car A")
        self.assertNotIn("8.00", r.response, "used car A's stale price")

    def test_pinned_modified_uses_current_value(self):
        reg = "MH01ZZ9033"
        self._add(reg, _core("HOND", "City", 2019, "P", "M", 1, 700000))
        sid = "pin3"
        self.assertEqual(self._ask(f"{reg} available hai?", sid).count, 1)
        self._modify(reg, {"rate": 850000})
        r = self._ask("price kya hai?", sid)                     # same pinned session
        self.assertIn("8.50", r.response)                        # CURRENT price
        self.assertNotIn("7.00", r.response)                     # not the old price


# ─────────────────────────────────────────────────────────────────────────────
# MISSING DATA (no fabrication of per-vehicle facts)
# ─────────────────────────────────────────────────────────────────────────────
class TestMissingData(AutoSyncBase):
    def test_blank_km_not_fabricated(self):
        reg = "MH01ZZ9050"
        self._add(reg, _core("MARU", "Swift", 2018, "P", "M", 1, 500000))  # no km
        r = self._ask(f"{reg} km kitna hai?", "m1")
        self.assertEqual(r.count, 1)
        # never invents a number; hedges to a visit
        self.assertNotIn("km chali hai", r.response.replace("exact km", ""))
        self.assertTrue("confirm" in r.response.lower()
                        or "visit" in r.response.lower())


# ─────────────────────────────────────────────────────────────────────────────
# RELOAD / RESTART / COUNT-INDEPENDENCE
# ─────────────────────────────────────────────────────────────────────────────
class TestReloadRestart(AutoSyncBase):
    def test_restart_loads_latest_inventory(self):
        reg = "MH01ZZ9060"
        self._add(reg, _core("MARU", "Dzire", 2019, "P", "M", 1, 600000))
        # simulate a process/Docker restart: a brand-new service on the same file
        svc2 = ChatService(xlsx_path=self.xlsx,
                           leads_db=os.path.join(self.tmp, "l2.db"),
                           analytics_db=os.path.join(self.tmp, "a2.db"),
                           unknown_db=os.path.join(self.tmp, "u2.db"))
        try:
            self.assertEqual(svc2.handle(f"{reg} available hai?", session_id="rs").count, 1)
        finally:
            svc2.close()

    def test_count_is_dynamic_not_hardcoded(self):
        base = self._n()
        self._add("MH01ZZ9061", _core("MARU", "Alto", 2017, "P", "M", 1, 300000))
        self._add("MH01ZZ9062", _core("MARU", "Alto", 2018, "P", "M", 1, 350000))
        self.assertEqual(self._n(), base + 2)
        self._remove("MH01ZZ9061")
        self.assertEqual(self._n(), base + 1)
        # the catalogue reply reflects the live number, whatever it is
        r = self._ask("all cars", "cnt")
        self.assertEqual(r.status, "catalogue")
        self.assertIn(str(self._n()), r.response)


# ─────────────────────────────────────────────────────────────────────────────
# REAL Owner/Staff SAVE endpoint -> sync (production path, not just the helper)
# ─────────────────────────────────────────────────────────────────────────────
class TestOwnerSaveEndpointSync(AutoSyncBase):
    def test_edit_car_endpoint_reflected_in_chatbot(self):
        import json
        import inventory_edit
        reg = "MH01ZZ9070"
        self._add(reg, _core("MARU", "Swift", 2018, "P", "M", 1, 500000))
        # drive the REAL edit endpoint (FileLock + atomic save + refresh_inventory)
        rate_col = COL["rate"] + 1
        body = json.dumps({"car_number": reg,
                           "values": {f"c{rate_col}": 675000}}).encode("utf-8")
        status, payload = inventory_edit.handle_update_car(self.svc, body)
        self.assertEqual(status, 200, payload)
        r = self._ask(f"{reg} ka price?", "es")
        self.assertIn("6.75", r.response)                        # chatbot sees the save


if __name__ == "__main__":
    unittest.main(verbosity=2)
