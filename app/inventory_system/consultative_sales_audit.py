"""
consultative_sales_audit.py  —  Phase 7P.1 STEP 1 (AUDIT ONLY)
===============================================================

Audits how "salesperson-like" the bot is on broad ENTRY intents, replaying the
REAL pilot conversation log (`data/pilot_query_log.db`; 33,740 turns). For each
entry intent we measure the PRE-7P.1 behaviour (CONSULTATIVE_LAYER disabled):

  * total turns of that intent
  * inventory-dump responses   — a multi-car listing ("Haan, N options hain …"
                                 / >= DUMP_THRESHOLD vehicles) with no
                                 recommendation and no follow-up question
  * consultative questions      — replies that recommend then ask ONE short
                                 question (essentially zero before the fix)

Entry intents audited (the brief's list):
  family car · budget only · SUV · sedan · hatchback · first car · city use ·
  office commute · CNG · low maintenance · automatic · finance required

Writes `consultative_sales_audit.xlsx` with sheets:
  Summary · Family Car · Budget Only · SUV · Sedan · Hatchback · Finance ·
  CNG · Recommendations

Read-only on the pilot log; isolated temp DBs (no production DB is touched).
"""
from __future__ import annotations

import os
import re
import sqlite3
import time
from collections import defaultdict
from typing import Dict, List, Tuple

import chat_service as CS
import consultative_sales as CSL
from astor_leak_audit import _speedup, _new_service

LOG_DB = os.path.join(os.path.dirname(__file__), "..", "..", "data",
                      "pilot_query_log.db")
DUMP_THRESHOLD = 4          # a listing of >=4 cars reads as an inventory dump

# ── audit-side entry-intent detector ─────────────────────────────────────────
# Broader than the production CSL.detect_intent: it also tags the intents that
# the consultative layer intentionally does NOT rewrite (finance -> FAQ,
# low-maintenance -> Astor-guard, office-commute -> unknown) so the audit can
# document them. Pure keyword scan over the raw turn.
_AUDIT_KEYWORDS = {
    "family":          (" family", " parivar", " ghar ke liye", " bacho", " kids",
                        " muv", " mpv", " 7 seater", " seven seater"),
    "budget":          (" budget", " sasti", " saste", " cheap", " affordable",
                        " kam paise", " kam daam", " under ", " ke andar", " tak ",
                        " lakh me", " lakh mein"),
    "suv":             (" suv",),
    "sedan":           (" sedan",),
    "hatchback":       (" hatchback", " hatch back"),
    "first_car":       (" first car", " pehli car", " pehli gaadi", " pehli gadi",
                        " naya driver", " learner", " beginner"),
    "city_use":        (" city use", " city ke liye", " shahar", " local use"),
    "office_commute":  (" office", " commute", " daftar", " kaam ke liye",
                        " duty ke liye"),
    "cng":             (" cng",),
    "low_maintenance": (" low maintenance", " kam maintenance", " maintenance kam",
                        " maintenance free"),
    "automatic":       (" automatic", " auto gear", " automatic gear"),
    "finance":         (" finance", " emi", " loan", " installment", " kist",
                        " down payment", " downpayment", " kisht"),
}
# Display name + the sheet it belongs to (only some get a dedicated sheet).
_INTENT_LABEL = {
    "family": "Family Car", "budget": "Budget Only", "suv": "SUV",
    "sedan": "Sedan", "hatchback": "Hatchback", "first_car": "First Car",
    "city_use": "City Use", "office_commute": "Office Commute", "cng": "CNG",
    "low_maintenance": "Low Maintenance", "automatic": "Automatic",
    "finance": "Finance",
}
_SHEET_INTENTS = ["family", "budget", "suv", "sedan", "hatchback", "finance", "cng"]
_SHEET_TITLE = {"family": "Family Car", "budget": "Budget Only", "suv": "SUV",
                "sedan": "Sedan", "hatchback": "Hatchback", "finance": "Finance",
                "cng": "CNG"}


def _audit_intent(message: str) -> str:
    t = " " + re.sub(r"\s+", " ", (message or "").strip().lower()) + " "
    # specific-model queries are not broad entry intents
    q = CS.parse(CS.normalize_typos(message))
    if q.model or q.make or q.registration:
        return ""
    order = ["finance", "first_car", "family", "city_use", "office_commute",
             "suv", "sedan", "hatchback", "cng", "low_maintenance", "automatic",
             "budget"]
    for intent in order:
        if any(k in t for k in _AUDIT_KEYWORDS[intent]):
            return intent
    return ""


def _is_dump(r) -> bool:
    """A multi-car inventory dump: many cards and no follow-up question."""
    resp = r.response or ""
    if "?" in resp and len(r.vehicles) <= 2:
        return False
    return bool(r.vehicles) and (r.count >= DUMP_THRESHOLD or "options hain" in resp)


def _is_consultative(r) -> bool:
    """Recommends (<=2 cars) and asks ONE short follow-up question."""
    resp = r.response or ""
    return bool(resp) and resp.rstrip().endswith("?") and resp.count("?") == 1 \
        and len(resp.splitlines()) <= 3


def _load_conversations() -> List[Tuple[str, List[str]]]:
    c = sqlite3.connect(LOG_DB)
    rows = c.execute(
        "SELECT conversation_id, user_query FROM query_log ORDER BY conversation_id, id"
    ).fetchall()
    c.close()
    convs: Dict[str, List[str]] = defaultdict(list)
    for cid, uq in rows:
        if uq:
            convs[cid].append(uq)
    return list(convs.items())


def audit() -> Dict:
    # PRE-7P.1 behaviour: disable the consultative layer for the audit replay.
    CS.CONSULTATIVE_LAYER = False
    svc = _new_service()
    convs = _load_conversations()
    stats = {k: {"turns": 0, "dump": 0, "consultative": 0, "examples": []}
             for k in _INTENT_LABEL}
    t0 = time.time()
    for n, (cid, turns) in enumerate(convs, 1):
        sid = f"caudit::{cid}"
        for turn in turns:
            intent = _audit_intent(turn)
            r = svc.handle(turn, session_id=sid)
            if not intent:
                continue
            s = stats[intent]
            s["turns"] += 1
            if _is_dump(r):
                s["dump"] += 1
            if _is_consultative(r):
                s["consultative"] += 1
            if len(s["examples"]) < 40:
                s["examples"].append({
                    "conversation_id": cid, "query": turn,
                    "route": r.meta.get("route", ""), "intent": r.intent,
                    "count": r.count, "vehicles": len(r.vehicles),
                    "dump": _is_dump(r), "consultative": _is_consultative(r),
                    "response": (r.response or "").replace("\n", " ")[:160]})
        if n % 1500 == 0:
            print(f"  ...{n}/{len(convs)} convs ({round(time.time()-t0,1)}s)",
                  flush=True)
    svc.close()
    CS.CONSULTATIVE_LAYER = True
    return {"conversations": len(convs), "stats": stats}


def _recommendation_rows() -> List[List]:
    """What the consultative layer WOULD recommend for each intent, drawn from
    CURRENT inventory only (so nothing unavailable is ever listed)."""
    import inventory_loader as L
    import config
    from retrieval_engine import RetrievalEngine
    items = L.load_inventory(config.XLSX_PATH)
    engine = RetrievalEngine(items)

    # Mirror the LIVE path exactly: parse a representative phrase per intent so
    # the recommended models are the same ones a real customer turn produces.
    _PHRASE = {
        "family": "family car batao", "suv": "SUV chahiye",
        "sedan": "sedan dikhao", "hatchback": "hatchback chahiye",
        "cng": "CNG car", "automatic": "automatic car chahiye",
        "budget": "car under 5 lakh", "first_car": "first car under 5 lakh",
        "city_use": "hatchback for city use",
    }

    def matches_for(intent):
        q = CS.parse(CS.normalize_typos(_PHRASE[intent]))
        res = engine.search(q)
        seen, out = set(), []
        for it in res.matches:
            if it.model and it.model not in seen:
                seen.add(it.model); out.append(it.model)
        return out

    rows = []
    for intent in ["family", "budget", "suv", "sedan", "hatchback", "cng",
                   "automatic", "city_use", "first_car"]:
        mm = matches_for(intent)
        recs = CSL.pick_recommendations(intent, mm)
        q_text = CSL._INTENTS[intent]["question"]
        intro = CSL.consultative_intro(intent, mm) or ""
        rows.append([_INTENT_LABEL[intent], ", ".join(recs) or "(none in stock)",
                     q_text, intro.replace("\n", " / ")])
    # finance / low-maintenance / office-commute are intentionally NOT rewritten
    rows.append(["Finance", "(handled by Finance FAQ — no inventory rewrite)",
                 "Finance lena hai ya cash purchase?", "n/a — FAQ path untouched"])
    rows.append(["Low Maintenance", "(hits Astor attr-guard — left untouched)",
                 "Petrol theek hai ya CNG chahiye?", "n/a — Astor fix untouched"])
    rows.append(["Office Commute", "(routes to unknown — no inventory match)",
                 "Roz ka commute kitne km ka hai?", "n/a — no retrieval to enrich"])
    return rows


def _write_xlsx(res: Dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    stats = res["stats"]
    wb = Workbook()

    # ── Summary ──
    s = wb.active
    s.title = "Summary"
    s.append(["Phase 7P.1 — Consultative Selling Audit (PRE-fix replay)"])
    s.append([f"Pilot log conversations replayed: {res['conversations']}"])
    s.append([])
    s.append(["Entry Intent", "Turns", "Inventory Dumps", "Dump %",
              "Consultative Qs", "Consultative %"])
    for k in ["family", "budget", "suv", "sedan", "hatchback", "first_car",
              "city_use", "office_commute", "cng", "low_maintenance",
              "automatic", "finance"]:
        d = stats[k]
        dump_pct = f"{(d['dump']/d['turns']*100):.0f}%" if d["turns"] else "-"
        con_pct = f"{(d['consultative']/d['turns']*100):.0f}%" if d["turns"] else "-"
        s.append([_INTENT_LABEL[k], d["turns"], d["dump"], dump_pct,
                  d["consultative"], con_pct])
    tot_turns = sum(d["turns"] for d in stats.values())
    tot_dump = sum(d["dump"] for d in stats.values())
    tot_con = sum(d["consultative"] for d in stats.values())
    s.append([])
    s.append(["TOTAL", tot_turns, tot_dump,
              f"{(tot_dump/tot_turns*100):.0f}%" if tot_turns else "-",
              tot_con, f"{(tot_con/tot_turns*100):.0f}%" if tot_turns else "-"])
    s.append([])
    s.append(["Finding: broad entry intents are answered as inventory dumps with"])
    s.append(["near-zero consultative questions. Phase 7P.1 adds ONE recommendation"])
    s.append(["+ ONE question on the inventory-routed entry intents."])
    s["A1"].font = Font(bold=True, size=12)
    for c in s[4]:
        c.font = Font(bold=True)
    s.column_dimensions["A"].width = 20
    for col in "BCDEF":
        s.column_dimensions[col].width = 16

    # ── per-intent example sheets ──
    def _sheet(title, intent):
        ws = wb.create_sheet(title)
        ws.append(["conversation_id", "query", "route", "intent", "count",
                   "vehicles", "is_dump", "is_consultative", "response (pre-fix)"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in stats[intent]["examples"]:
            ws.append([r["conversation_id"], r["query"], r["route"], r["intent"],
                       r["count"], r["vehicles"], r["dump"], r["consultative"],
                       r["response"]])
        for col, w in zip("ABCDEFGHI", [16, 26, 11, 13, 7, 9, 9, 14, 90]):
            ws.column_dimensions[col].width = w

    for intent in _SHEET_INTENTS:
        _sheet(_SHEET_TITLE[intent], intent)

    # ── Recommendations design sheet ──
    rec = wb.create_sheet("Recommendations")
    rec.append(["Entry Intent", "Recommended (from current inventory)",
                "Consultative Question", "After-fix intro"])
    for c in rec[1]:
        c.font = Font(bold=True)
    for row in _recommendation_rows():
        rec.append(row)
    for col, w in zip("ABCD", [18, 40, 34, 70]):
        rec.column_dimensions[col].width = w
    for r in rec.iter_rows(min_row=2):
        r[3].alignment = Alignment(wrap_text=True)

    out = os.path.join(os.path.dirname(__file__), "consultative_sales_audit.xlsx")
    wb.save(out)
    print("Wrote", out, flush=True)
    print(f"TOTAL entry turns {tot_turns} | dumps {tot_dump} "
          f"({(tot_dump/tot_turns*100):.0f}%) | consultative {tot_con}", flush=True)


def main() -> None:
    _speedup()
    print("Auditing consultative selling (pre-fix replay)...", flush=True)
    res = audit()
    _write_xlsx(res)


if __name__ == "__main__":
    main()
