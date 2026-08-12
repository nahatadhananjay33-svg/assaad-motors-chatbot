"""
marathi_response_audit.py  —  Phase 7O.5 STEP 1 (AUDIT)
=======================================================

Measures the RESPONSE-language quality for Marathi customers. The referenced
`evaluation_results_marathi_v4.xlsx` and `final_conversations_only.txt` are not in
the repo, so the pilot log (`data/pilot_query_log.db`) is the available dataset.

For every conversation that contains a Marathi customer turn, replay it in order
and — for each Marathi customer turn — classify the BOT reply's language:

  * pure_marathi  — Marathi markers present, NO Hindi/Hinglish markers.
  * mixed         — Marathi markers AND Hindi/Hinglish markers.
  * hindi_hinglish— no Marathi markers, Hindi/Hinglish markers present.
  * neutral       — neither (very short / numeric / English proper-noun only).

"Marathi reply" = pure_marathi + mixed (reply is at least partly Marathi); the
quality target is on pure_marathi. Also tallies the failing replies by route /
intent so the root cause is visible. Read-only on the pilot log; isolated DBs.
"""
from __future__ import annotations

import os
import re
import sqlite3
from collections import defaultdict, Counter
from typing import Dict, List, Tuple

import chat_service as CS
from language_detector import detect_language
from astor_leak_audit import _speedup, _new_service

LOG_DB = os.path.join(os.path.dirname(__file__), "..", "..", "data",
                      "pilot_query_log.db")

# Marathi response markers (Devanagari) — copula / availability / common words.
_MAR = ["आहे", "आहेत", "नाही", "उपलब्ध", "किंमत", "गाडी", "गाड्या", "फोटो",
        "व्हिडिओ", "येऊन", "येऊ", "बघा", "बघू", "नक्की", "धन्यवाद", "मिळ",
        "तुम्ही", "कोणत्या", "कोणती", "हवे", "हवी", "हवा", "सांगा", "पाठव",
        "स्वागत", "शोरूम", "वॉरंटी", "विमा", "सर्व्हिस", "लाख", "हो,", "हो.",
        "नमस्कार", "लोकेशन", "ठरलेल्या", "पर्याय", "गरज", "वेळ"]
# Hinglish (Latin) response markers.
_HING = ["haan", "hai", "hain", "available", "lakh", "aap ", "dekh", "kar deta",
         "kitne", "kaunsi", "chahiye", "milega", "bhej", "gaadi", "photo",
         "video", "abhi", "karo", "batao", "deta hoon",
         "konsi", "dikha", "hoga", "sakte", "bata do", "options hain"]
# Hindi (Devanagari) response markers — distinct from Marathi.
_HIN = ["है", "हैं", "नहीं", "कीमत", "गाड़ी", "चाहिए", "दिखा", "मिलेगा", "हाँ",
        "क्या", "देता हूँ", "सकते हो", "रहे", "लीजिए"]

_LATIN_WORD = re.compile(r"[a-zA-Z]{3,}")


def _classify_reply(resp: str) -> str:
    if not resp:
        return "neutral"
    low = resp.lower()
    m = sum(1 for w in _MAR if w in resp)
    h = sum(1 for w in _HING if w in low) + sum(1 for w in _HIN if w in resp)
    if m > 0 and h == 0:
        return "pure_marathi"
    if m > 0 and h > 0:
        return "mixed"
    if h > 0:
        return "hindi_hinglish"
    return "neutral"


def _load_marathi_conversations() -> List[Tuple[str, List[str]]]:
    c = sqlite3.connect(LOG_DB)
    rows = c.execute(
        "SELECT conversation_id, user_query FROM query_log ORDER BY conversation_id, id"
    ).fetchall()
    c.close()
    convs: Dict[str, List[str]] = defaultdict(list)
    for cid, uq in rows:
        if uq:
            convs[cid].append(uq)
    return [(cid, t) for cid, t in convs.items()
            if any(detect_language(x) == "marathi" for x in t)]


def audit():
    svc = _new_service()
    convs = _load_marathi_conversations()
    cls = Counter()
    fail_by_route = Counter()
    fail_by_intent = Counter()
    samples = defaultdict(list)
    import time
    t0 = time.time()
    for n, (cid, turns) in enumerate(convs, 1):
        sid = f"mar::{cid}"
        for turn in turns:
            is_mar = detect_language(turn) == "marathi"
            r = svc.handle(turn, session_id=sid)
            if not is_mar:
                continue
            c = _classify_reply(r.response)
            cls[c] += 1
            if c in ("hindi_hinglish", "mixed"):
                route = r.meta.get("route", "?")
                fail_by_route[route] += 1
                fail_by_intent[r.intent] += 1
                if len(samples[r.intent]) < 5:
                    samples[r.intent].append((turn, r.intent, route,
                                              r.response.replace("\n", " ")[:90]))
        if n % 300 == 0:
            print(f"  ...{n}/{len(convs)} ({round(time.time()-t0,1)}s)", flush=True)
    svc.close()
    return {"conversations": len(convs), "cls": cls,
            "fail_by_route": fail_by_route, "fail_by_intent": fail_by_intent,
            "samples": samples}


def _write_xlsx(res):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    cls = res["cls"]
    total = sum(cls.values()) or 1
    pure = cls["pure_marathi"]; mixed = cls["mixed"]
    hh = cls["hindi_hinglish"]; neu = cls["neutral"]
    marathi_reply = pure + mixed
    wb = Workbook()
    s = wb.active; s.title = "Summary"
    for row in [
        ["Metric", "Count", "%"],
        ["Marathi customer turns measured", total, "100%"],
        ["Pure Marathi reply", pure, f"{pure/total*100:.1f}%"],
        ["Mixed Hindi/Marathi reply", mixed, f"{mixed/total*100:.1f}%"],
        ["Hindi / Hinglish reply (fail)", hh, f"{hh/total*100:.1f}%"],
        ["Neutral (numeric / proper-noun only)", neu, f"{neu/total*100:.1f}%"],
        ["", "", ""],
        ["Marathi Reply % (pure + mixed)", marathi_reply, f"{marathi_reply/total*100:.1f}%"],
        ["Pure Marathi %", pure, f"{pure/total*100:.1f}%"],
        ["Mixed %", mixed, f"{mixed/total*100:.1f}%"],
    ]:
        s.append(row)
    for c in s[1]:
        c.font = Font(bold=True)
    s.column_dimensions["A"].width = 42

    r1 = wb.create_sheet("Fail by Route")
    r1.append(["route", "fail_count"]);
    for c in r1[1]: c.font = Font(bold=True)
    for k, v in res["fail_by_route"].most_common():
        r1.append([k, v])
    r2 = wb.create_sheet("Fail by Intent")
    r2.append(["intent", "fail_count"])
    for c in r2[1]: c.font = Font(bold=True)
    for k, v in res["fail_by_intent"].most_common():
        r2.append([k, v])
    r3 = wb.create_sheet("Fail Samples")
    r3.append(["query", "intent", "route", "response"])
    for c in r3[1]: c.font = Font(bold=True)
    for intent, lst in res["samples"].items():
        for rec in lst:
            r3.append(list(rec))
    for col, w in zip("ABCD", [24, 16, 14, 90]):
        r3.column_dimensions[col].width = w

    wb.save(os.path.join(os.path.dirname(__file__), "marathi_response_audit.xlsx"))
    print(f"PURE {pure} | MIXED {mixed} | HINDI/HINGLISH {hh} | NEUTRAL {neu} "
          f"| total {total} | Marathi-reply% {marathi_reply/total*100:.1f} "
          f"| pure% {pure/total*100:.1f}", flush=True)
    print("Top fail routes:", res["fail_by_route"].most_common(8), flush=True)
    print("Top fail intents:", res["fail_by_intent"].most_common(10), flush=True)


def main():
    _speedup()
    # STEP-1 baseline = pre-fix state: disable the Phase-7O.5 post-processor so
    # the audit measures the original Hinglish replies regardless of whether the
    # fix is present in the tree.
    CS.to_marathi = lambda x: x
    print("Auditing Marathi response language (pre-fix)...", flush=True)
    res = audit()
    _write_xlsx(res)
    print("Wrote marathi_response_audit.xlsx", flush=True)


if __name__ == "__main__":
    main()
