"""
astor_leak_audit.py  —  Phase 7O.2 validation harness
======================================================

Reproduces the Astor default-leak audit from the REAL pilot conversation log
(`data/pilot_query_log.db`) and writes `astor_leak_audit.xlsx`.

A turn is an **Astor default leak** when, replaying the conversation in order
through the live `ChatService`:

  * the user turn is a vehicle-specific attribute question (insurance / warranty
    / service history / condition / accident / owner count / RC / loan / flood /
    body / engine), AND
  * no vehicle was ever selected earlier in the conversation (so there is no
    active vehicle context to anchor the answer), AND
  * the bot's reply names the default rank-#1 inventory vehicle (the 2022 Black
    Astor) even though the customer never selected it.

`before` = pre-fix behaviour, reproduced by disabling ONLY the Phase 7O.2 guard
(`_attr_clarification` forced to return None) — exactly the code path that
produced the leak. `after` = the shipped guard. Nothing else differs.

Conversation context is preserved turn-by-turn (one ChatService session per
conversation_id), so a prior vehicle selection correctly suppresses a leak —
this is why the leak count is far below the raw bare-attribute turn count.

Isolated temp DBs are used for analytics/leads/unknown so the production
databases and the pilot log are never written to.
"""
from __future__ import annotations

import os
import re
import sqlite3
import sys
import tempfile
import time
from typing import Dict, List, Tuple

import chat_service as CS
from chat_service import ChatService

LEAK_MODEL = "Astor"          # current inventory rank #1 == 2022 Black Astor
LOG_DB = os.path.join(os.path.dirname(__file__), "..", "..", "data",
                      "pilot_query_log.db")

# Cheap pre-filter so we only deep-inspect plausibly-attribute turns.
_ATTR_KW = ("insurance", "service", "warranty", "rc", "loan", "flood",
            "condition", "owner", "malik", "maalik", "malak", "accident",
            "aksident", "axident", "noc", "hypothec", "engine", "body", "damage")


def _load_conversations() -> List[Tuple[str, List[str]]]:
    c = sqlite3.connect(LOG_DB)
    rows = c.execute(
        "SELECT conversation_id, user_query FROM query_log ORDER BY conversation_id, id"
    ).fetchall()
    c.close()
    convs: Dict[str, List[str]] = {}
    for cid, uq in rows:
        if uq:
            convs.setdefault(cid, []).append(uq)
    # Only replay conversations that contain at least one plausible attribute
    # turn — the rest can never produce an Astor attribute leak.
    out = []
    for cid, turns in convs.items():
        if any(any(k in t.lower() for k in _ATTR_KW) for t in turns):
            out.append((cid, turns))
    return out


def _speedup() -> None:
    """Harness-only performance shim. The parser recompiles regexes on every
    call (~4 parses/turn); over thousands of turns that dominates runtime.
    `parse` and `normalize_typos` are PURE deterministic functions of their input
    string, so memoizing them (deepcopy on return, so callers can still mutate
    the result) changes ONLY speed, never the response. Patched into every module
    that imported the name. Also no-ops the lead-capture / analytics tail, which
    never affects the bot reply or the leak verdict."""
    import copy
    import query_parser as QP
    import faq_router as FR

    _pcache: Dict[str, object] = {}
    _real_parse = QP.parse

    def fast_parse(raw):
        key = raw if isinstance(raw, str) else str(raw)
        if key not in _pcache:
            _pcache[key] = _real_parse(raw)
        return copy.deepcopy(_pcache[key])

    _ncache: Dict[str, str] = {}
    _real_norm = QP.normalize_typos

    def fast_norm(s):
        if s not in _ncache:
            _ncache[s] = _real_norm(s)
        return _ncache[s]

    for mod in (QP, FR, CS):
        if hasattr(mod, "parse"):
            mod.parse = fast_parse
        if hasattr(mod, "normalize_typos"):
            mod.normalize_typos = fast_norm


def _new_service() -> ChatService:
    d = tempfile.mkdtemp(prefix="astor_audit_")
    svc = ChatService(
        analytics_db=os.path.join(d, "a.db"),
        leads_db=os.path.join(d, "l.db"),
        unknown_db=os.path.join(d, "u.db"),
    )
    # lead capture / analytics never change the response or the leak verdict —
    # disable for the audit (both are regex-heavy on the hot path).
    class _Lead:
        score_level = None
        visit_ready = False
    svc.lead_engine.capture = lambda *a, **k: _Lead()
    svc.analytics.record = lambda *a, **k: None
    return svc


def run(guard_on: bool) -> Tuple[int, int, List[Dict]]:
    """Replay every candidate conversation. Returns (leaks, convs_affected, rows).

    When guard_on is False we monkeypatch `_attr_clarification` to always return
    None — the exact pre-fix path that leaked the default Astor."""
    orig = CS._attr_clarification
    if not guard_on:
        CS._attr_clarification = lambda q: None   # disable the Phase 7O.2 guard
    try:
        svc = _new_service()
        convs = _load_conversations()
        leaks = 0
        affected = set()
        rows: List[Dict] = []
        t0 = time.time()
        for n, (cid, turns) in enumerate(convs, 1):
            sid = f"audit::{cid}"
            has_context = False          # has a vehicle been selected in a PRIOR turn?
            for turn in turns:
                q = CS.parse(CS.normalize_typos(turn))
                is_attr = orig(q) is not None    # the real guard predicate
                r = svc.handle(turn, session_id=sid)
                # leak test FIRST, against context from prior turns only: an
                # attribute turn, no vehicle ever selected, yet Astor surfaced.
                if is_attr and not has_context and LEAK_MODEL in r.response:
                    leaks += 1
                    affected.add(cid)
                    rows.append({"conversation_id": cid, "query": turn,
                                 "intent": r.intent,
                                 "response": r.response[:160]})
                # THEN update context — mirror Phase 7I.2 `_followup_ctx`
                # establishment: a named vehicle, a single resolved car, or media.
                if (q.model or q.make or q.registration
                        or len(r.vehicles) == 1
                        or (r.media and r.media.get("status") == "ok")):
                    has_context = True
            if n % 200 == 0:
                print(f"  ...{n}/{len(convs)} convs  "
                      f"(leaks so far {leaks}, {round(time.time()-t0,1)}s)",
                      flush=True)
        svc.close()
        return leaks, len(affected), rows
    finally:
        CS._attr_clarification = orig


def main() -> None:
    _speedup()
    print("Replaying pilot conversations (BEFORE — guard disabled)...", flush=True)
    b_leaks, b_convs, b_rows = run(guard_on=False)
    print(f"BEFORE: leaks={b_leaks}  conversations_affected={b_convs}", flush=True)

    print("Replaying pilot conversations (AFTER — guard enabled)...", flush=True)
    a_leaks, a_convs, a_rows = run(guard_on=True)
    print(f"AFTER:  leaks={a_leaks}  conversations_affected={a_convs}", flush=True)

    _write_xlsx(b_leaks, b_convs, b_rows, a_leaks, a_convs, a_rows)
    print("Wrote astor_leak_audit.xlsx", flush=True)


def _write_xlsx(b_leaks, b_convs, b_rows, a_leaks, a_convs, a_rows) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    s = wb.active
    s.title = "Summary"
    s.append(["Metric", "Before", "After"])
    s.append(["Astor default leaks", b_leaks, a_leaks])
    s.append(["Conversations affected", b_convs, a_convs])
    d = wb.create_sheet("Before_Leaks")
    d.append(["conversation_id", "query", "intent", "response"])
    for r in b_rows:
        d.append([r["conversation_id"], r["query"], r["intent"], r["response"]])
    d2 = wb.create_sheet("After_Leaks")
    d2.append(["conversation_id", "query", "intent", "response"])
    for r in a_rows:
        d2.append([r["conversation_id"], r["query"], r["intent"], r["response"]])
    wb.save(os.path.join(os.path.dirname(__file__), "astor_leak_audit.xlsx"))


if __name__ == "__main__":
    main()
