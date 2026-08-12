"""
media_loader_mapping.py
=======================

Phase 4A.5 — Final Wiring.

The one missing link between the **media_sync** writer and the **MediaService**
reader. media_sync writes Supabase public URLs into the DNJ media columns
(EXTERIOR 1..10, INTERIOR 1..10, VIDEO 1..5, INSTAGRAM 1..5, YOUTUBE 1..5).
`inventory_loader.py` reads columns A..Q only and leaves `InventoryItem.media`
empty, so `InventoryMediaProvider.fetch()` always returns nothing. This module
closes that gap: it reads the media URL columns straight from the DNJ sheet and
maps them onto `InventoryItem.media` as `InventoryMedia` records.

    Excel media columns ──► InventoryItem.media ──► InventoryMediaProvider ──► MediaService

Design rules honoured (Phase 4A.5 constraints)
----------------------------------------------
* Excel (DNJ) stays the ONLY source of truth. We read media URLs FROM Excel.
* Supabase is media storage only — we never query it here; we only read the URLs
  the sync layer already wrote into the sheet.
* No new tables, no inventory in Supabase, no architecture redesign. This is a
  pure read-mapping step that plugs into the existing loader and the existing
  `InventoryMedia` / `MediaType` models.
* Resilient to column drift: media columns are located by their HEADER TEXT
  (row 2), not by hard-coded letters, so adding/removing a slot does not break
  the mapping. A fixed reference layout is documented below for grounding.

Reference layout in `IVR_Sheet.xlsx` → sheet DNJ (header row 2, data from row 4):

    cols 18..27  EXTERIOR 1..10   -> MediaType.EXTERIOR_PHOTO
    cols 28..37  INTERIOR 1..10   -> MediaType.INTERIOR_PHOTO
    cols 38..42  VIDEO 1..5       -> MediaType.VIDEO
    cols 43..47  INSTAGRAM 1..5   -> MediaType.INSTAGRAM
    cols 48..52  YOUTUBE 1..5     -> MediaType.YOUTUBE
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from inventory_models import InventoryItem, InventoryMedia, MediaType

# ── Sheet geometry (kept in lockstep with inventory_loader.py) ──
HEADER_ROW = 2          # the row carrying "EXTERIOR 1", "INTERIOR 1", ...
DATA_START_ROW = 4      # first real data row
DNJ_SHEET = "DNJ"

# Header keyword (uppercased prefix) -> canonical MediaType.
# The first column of each group carries the keyword ("EXTERIOR 1 (10 photos)"),
# the rest carry only the slot number ("2", "3", ...). We attach those trailing
# numeric columns to the most recent keyword group.
MEDIA_KEYWORDS: Dict[str, str] = {
    "EXTERIOR": MediaType.EXTERIOR_PHOTO,
    "INTERIOR": MediaType.INTERIOR_PHOTO,
    "VIDEO": MediaType.VIDEO,
    "INSTAGRAM": MediaType.INSTAGRAM,
    "YOUTUBE": MediaType.YOUTUBE,
}

# CAR NUMB header aliases (matches the sync layer's vocabulary).
_CAR_NUMB_HEADERS = {"CAR NUMB", "CAR NUMBER", "CAR_NUMB", "REGISTRATION", "REG NO"}

# A value only counts as media if it is an http(s) URL. Anything else in a media
# cell (stray notes, codes) is ignored — Supabase public URLs are always http(s).
_URL_RE = re.compile(r"^https?://", re.IGNORECASE)

_NUMERIC_RE = re.compile(r"^\d+$")


def _media_keyword(upper: str) -> Optional[str]:
    """Return the media keyword a header begins with, but ONLY when it is a real
    slot header — i.e. the keyword is followed by a slot number, a space, or the
    end of the text (e.g. "EXTERIOR 1", "VIDEO 1", "INSTAGRAM"). This deliberately
    rejects management columns that merely *start* with a keyword, such as
    "VIDEO_URLS", "YOUTUBE_URL", "INSTAGRAM_URL" (Phase 8H), so they are never
    mistaken for media URL slots."""
    for kw in MEDIA_KEYWORDS:
        if upper.startswith(kw):
            rest = upper[len(kw):]
            # accept bare keyword ("INSTAGRAM"), a spaced slot ("VIDEO 1"), or any
            # slot-numbered form ("VIDEO_1", "EXTERIOR_PHOTO_3"). Reject digit-less
            # suffixes like "_URLS"/"_URL" so management columns never match.
            if rest == "" or rest[0].isspace() or any(ch.isdigit() for ch in rest):
                return kw
    return None


@dataclass
class MediaLayout:
    """Where the media lives in a given sheet, discovered from the header row."""
    car_col: Optional[int] = None                         # 1-based
    # media_type -> ordered list of (slot_number, column_index_1based)
    groups: Dict[str, List[Tuple[int, int]]] = field(default_factory=dict)

    def has_media_columns(self) -> bool:
        return any(self.groups.values())


def _norm(value: Any) -> str:
    return ("" if value is None else str(value)).strip()


def _norm_reg(value: Any) -> str:
    """Same normalization the loader/sync use as the registration key."""
    return _norm(value).upper().replace(" ", "")


# ─────────────────────────────────────────────────────────────────────────────
# 1. Discover media columns from the header row (column-drift resilient)
# ─────────────────────────────────────────────────────────────────────────────
def detect_media_layout(header_row: List[Any]) -> MediaLayout:
    """
    Map the header row (a list of cell values, index 0 == column A) to a
    MediaLayout. Groups are the contiguous runs: one keyword cell followed by
    its numeric continuation cells.
    """
    layout = MediaLayout()
    current_type: Optional[str] = None
    slot_in_group = 0

    for i, cell in enumerate(header_row):
        col = i + 1                                        # 1-based
        text = _norm(cell)
        upper = text.upper()

        if upper in _CAR_NUMB_HEADERS:
            layout.car_col = col
            current_type = None
            continue

        keyword = _media_keyword(upper)
        if keyword:
            current_type = MEDIA_KEYWORDS[keyword]
            slot_in_group = 1
            layout.groups.setdefault(current_type, []).append((slot_in_group, col))
        elif current_type and _NUMERIC_RE.match(upper):
            slot_in_group += 1
            layout.groups[current_type].append((slot_in_group, col))
        else:
            current_type = None                            # run broken

    return layout


# ─────────────────────────────────────────────────────────────────────────────
# 2. Build InventoryMedia records for one row
# ─────────────────────────────────────────────────────────────────────────────
def build_media_for_row(registration_no: str, row: Tuple[Any, ...],
                        layout: MediaLayout) -> List[InventoryMedia]:
    """
    Read every populated media slot for one data row and return the matching
    InventoryMedia records (URLs only). Empty / non-URL cells are skipped.
    """
    out: List[InventoryMedia] = []
    for media_type, slots in layout.groups.items():
        for slot_number, col in slots:
            idx = col - 1                                  # row tuple is 0-based
            value = row[idx] if idx < len(row) else None
            url = _norm(value)
            if not url or not _URL_RE.match(url):
                continue
            out.append(InventoryMedia(
                registration_no=registration_no,
                media_type=media_type,
                slot=slot_number,
                url=url,
                # the first exterior photo is the natural hero image
                is_primary=(media_type == MediaType.EXTERIOR_PHOTO and slot_number == 1),
                sort_order=slot_number,
            ))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 3. Read the whole media index from the workbook  {reg -> [InventoryMedia]}
# ─────────────────────────────────────────────────────────────────────────────
def read_media_index(path: str, sheet: str = DNJ_SHEET) -> Dict[str, List[InventoryMedia]]:
    """
    Read media URL columns for every registration in `sheet`. Returns a map
    keyed by normalized registration_no. Sheets with no media columns yield {}.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {path}. Sheets: {wb.sheetnames}")
        ws = wb[sheet]

        header = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
        layout = detect_media_layout(header)
        if not layout.has_media_columns() or layout.car_col is None:
            return {}

        car_idx = layout.car_col - 1
        index: Dict[str, List[InventoryMedia]] = {}
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            reg = _norm_reg(row[car_idx] if car_idx < len(row) else None)
            if not reg:
                continue
            media = build_media_for_row(reg, row, layout)
            if media:
                index.setdefault(reg, []).extend(media)
        return index
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# 4. Attach media to already-loaded InventoryItem objects (the wiring step)
# ─────────────────────────────────────────────────────────────────────────────
def attach_media(items: List[InventoryItem], path: str,
                 sheet: str = DNJ_SHEET) -> int:
    """
    Mutate `items` in place, filling `InventoryItem.media` from the sheet's media
    columns (matched by registration_no). Returns the number of media records
    attached. Idempotent: replaces any existing media list for matched items.
    """
    index = read_media_index(path, sheet)
    if not index:
        return 0
    attached = 0
    for item in items:
        media = index.get(item.registration_no)
        if media:
            # deterministic order: type, then slot
            item.media = sorted(media, key=lambda m: (m.media_type, m.slot))
            attached += len(item.media)
    return attached


if __name__ == "__main__":
    import sys
    from inventory_loader import load_inventory

    p = sys.argv[1] if len(sys.argv) > 1 else "../IVR_Sheet.xlsx"
    items = load_inventory(p)          # load_inventory now attaches media itself
    with_media = [i for i in items if i.media]
    total = sum(len(i.media) for i in items)
    print(f"items={len(items)}  with_media={len(with_media)}  media_records={total}")
    for it in with_media[:5]:
        print(f"  {it.registration_no}: {len(it.media)} media "
              f"({sorted({m.media_type for m in it.media})})")
