"""
media_admin_validate.py — Phase 8H Step 8 validation harness.

Exercises the full owner workflow against a TEST workbook with the OFFLINE
in-memory store (no Supabase, no production data touched):

  upload photo -> upload video -> add IG/YT -> auto-update Excel ->
  verify chatbot read path sees the media -> mark sold -> media deleted ->
  row moved to SOLD_CARS -> vehicle disappears from chatbot ->
  regression: other cars still load unchanged.

Run:  python media_admin_validate.py
Exit code 0 = all assertions passed; prints a JSON result.
"""
from __future__ import annotations

import os
import json
import tempfile

import openpyxl

import media_admin as MA
from inventory_loader import load_inventory, customer_facing


# ── build a DNJ-faithful test workbook (header row 2, data from row 4) ──
def build_test_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = MA.DNJ_SHEET

    # header row 2 (row 1 + row 3 left like the real sheet: title / descriptions)
    hdr = {4: "MODEL", 6: "INS", 7: "VARI", 8: "F", 9: "T", 10: "O", 11: "KM",
           12: "COL", 13: "RATE", 14: "CAR NUMB", 16: "LOC", 17: "RTO"}
    # media slots starting col 18 (R)
    media = (["EXTERIOR 1", 2, 3, 4, 5]      # 18-22
             + ["INTERIOR 1", 2, 3]          # 23-25
             + ["VIDEO 1", 2, 3]             # 26-28
             + ["INSTAGRAM 1", 2]            # 29-30
             + ["YOUTUBE 1", 2])             # 31-32
    for c, v in hdr.items():
        ws.cell(2, c).value = v
    for i, v in enumerate(media):
        ws.cell(2, 18 + i).value = v
    ws.cell(3, 4).value = "Car"               # description row (must be skipped)

    # data rows 4..6  (A..Q meaningful to the loader)
    cars = [
        # stock, sr, make, model, year, ins, vari, fuel, trans, own, km, col, rate, carnumb, last4, loc, rto
        [1, 1, "MARU", "Swift", 2019, None, "VXI", "P", "M", 2, 42000, "WHITE", 550000, "MH02EZ6001", 6001, "B2", "MH02"],
        [2, 2, "HYUN", "Creta", 2020, None, "SX", "D", "A", 1, 31000, "SILVER", 1150000, "MH12AB1234", 1234, "Y5", "MH12"],
        [3, 3, "HOND", "City",  2018, None, "VX", "P", "M", 2, 58000, "GREY",   720000, "MH14CD5678", 5678, "A1", "MH14"],
    ]
    for r, row in enumerate(cars, start=4):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val

    # the loader reconciles against the legacy sold sheet — keep an empty one
    wb.create_sheet("DONT TOUCH SOLD")
    wb.save(path)


def _media_urls_for(path: str, reg: str):
    items = load_inventory(path)
    by_reg = {i.registration_no: i for i in items}
    it = by_reg.get(reg)
    return [m.url for m in (it.media if it else [])]


def main() -> int:
    results = {}
    failures = []

    def check(name, cond, extra=None):
        results[name] = bool(cond)
        if not cond:
            failures.append((name, extra))

    d = tempfile.mkdtemp(prefix="phase8h_")
    wb_path = os.path.join(d, "test_inventory.xlsx")
    build_test_workbook(wb_path)

    store = MA.InMemoryMediaStore()
    admin = MA.MediaAdmin(wb_path, store=store, bucket="car-photos",
                          backup_dir=os.path.join(d, "backups"))

    REG = "MH02EZ6001"

    # Step 6: list vehicles
    vlist = admin.list_vehicles()
    check("list_3_vehicles", len(vlist) == 3, vlist)
    check("list_status_default_available",
          all(v["status"] == "AVAILABLE" for v in vlist), vlist)
    check("list_has_model", any(v["model"] == "Swift" for v in vlist), vlist)

    # Step 4+5: upload photos
    photos = [("front.jpg", b"\xff\xd8\xff jpeg-bytes-1"),
              ("side.png", b"\x89PNG png-bytes-2"),
              ("bad.txt", b"nope")]
    rp = admin.upload_files(REG, photos, kind="photos")
    check("photo_upload_ok", rp["status"] == "ok", rp)
    check("photo_uploaded_2", rp["uploaded"] == 2, rp)
    check("photo_rejected_bad_ext",
          any(f["status"] == "invalid_file" for f in rp["files"]), rp)
    check("store_has_2_photos",
          len(store.list_paths("car-photos", f"{REG}/photos/")) == 2,
          store.list_paths("car-photos", f"{REG}/photos/"))

    # videos
    rv = admin.upload_files(REG, [("walkaround.mp4", b"\x00\x00\x00 mp4-bytes")], kind="videos")
    check("video_upload_ok", rv["status"] == "ok" and rv["uploaded"] == 1, rv)
    check("store_has_1_video",
          len(store.list_paths("car-photos", f"{REG}/videos/")) == 1, None)

    # IG + YT links
    ri = admin.add_link(REG, "instagram", "https://instagram.com/reel/abc123")
    ry = admin.add_link(REG, "youtube", "https://youtu.be/XYZ789")
    check("instagram_ok", ri["status"] == "ok", ri)
    check("youtube_ok", ry["status"] == "ok", ry)
    check("bad_link_rejected",
          admin.add_link(REG, "youtube", "not-a-url")["status"] == "rejected")

    # Step 5: auto-updated Excel — mirror + management columns
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[MA.DNJ_SHEET]
    layout = MA._discover_layout(ws)
    rows = MA._index_rows(ws, layout)
    row = rows[REG]
    g = layout.mgmt
    photo_mirror = ws.cell(row, g["PHOTO_URLS"]).value or ""
    video_mirror = ws.cell(row, g["VIDEO_URLS"]).value or ""
    check("mirror_photo_has_2", photo_mirror.count("http") == 2, photo_mirror)
    check("mirror_video_has_1", video_mirror.count("http") == 1, video_mirror)
    check("mirror_youtube_set", "youtu.be" in (ws.cell(row, g["YOUTUBE_URL"]).value or ""))
    check("mirror_instagram_set", "instagram" in (ws.cell(row, g["INSTAGRAM_URL"]).value or ""))
    check("folder_id_set", ws.cell(row, g["MEDIA_FOLDER_ID"]).value == REG)
    check("last_updated_set", bool(ws.cell(row, g["LAST_UPDATED"]).value))
    check("status_available", (ws.cell(row, g["STATUS"]).value or "").upper() == "AVAILABLE")
    wb.close()

    # Chatbot read path: loader must surface the uploaded media (no loader change)
    chatbot_urls = _media_urls_for(wb_path, REG)
    check("chatbot_sees_photos", sum(1 for u in chatbot_urls if "/photos/" in u) == 2, chatbot_urls)
    check("chatbot_sees_video", any("/videos/" in u for u in chatbot_urls), chatbot_urls)
    check("chatbot_sees_youtube", any("youtu.be" in u for u in chatbot_urls), chatbot_urls)

    # baseline live count before sold
    before = customer_facing(load_inventory(wb_path))
    before_regs = {i.registration_no for i in before}
    check("baseline_has_3_live", len(before_regs) == 3, before_regs)

    # Step 7: mark sold (confirm gate first)
    needs = admin.mark_sold(REG, confirm=False)
    check("mark_sold_requires_confirm", needs["status"] == "confirm_required", needs)

    sold = admin.mark_sold(REG, confirm=True, user="owner")
    check("mark_sold_ok", sold["status"] == "ok", sold)
    check("sold_deleted_files", sold["files_deleted"] == 3, sold)   # 2 photos + 1 video
    check("store_emptied_for_reg",
          len(store.list_paths("car-photos", f"{REG}/")) == 0,
          store.list_paths("car-photos", f"{REG}/"))

    # row moved to SOLD_CARS, removed from DNJ
    wb = openpyxl.load_workbook(wb_path)
    check("sold_cars_sheet_exists", MA.SOLD_SHEET in wb.sheetnames, wb.sheetnames)
    if MA.SOLD_SHEET in wb.sheetnames:
        ws_sold = wb[MA.SOLD_SHEET]
        sold_regs = [MA._norm_reg(ws_sold.cell(r, layout.car_col).value)
                     for r in range(2, ws_sold.max_row + 1)]
        check("reg_in_sold_cars", REG in sold_regs, sold_regs)
    wb.close()

    # Step 7/8: vehicle disappears from chatbot
    after = customer_facing(load_inventory(wb_path))
    after_regs = {i.registration_no for i in after}
    check("vehicle_gone_from_chatbot", REG not in after_regs, after_regs)

    # Regression: the OTHER two cars still load unchanged
    check("other_cars_intact",
          after_regs == {"MH12AB1234", "MH14CD5678"}, after_regs)
    creta = next((i for i in after if i.registration_no == "MH12AB1234"), None)
    check("regression_fields_intact",
          creta is not None and creta.model == "Creta" and creta.price_inr == 1150000,
          None if creta is None else (creta.model, creta.price_inr))

    summary = {"passed": sum(1 for v in results.values() if v),
               "total": len(results),
               "all_passed": not failures,
               "failures": [{"check": n, "detail": e} for n, e in failures],
               "checks": results,
               "workdir": d}
    print(json.dumps(summary, indent=2, default=str))
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
