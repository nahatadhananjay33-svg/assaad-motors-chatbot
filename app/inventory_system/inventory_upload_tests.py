"""
inventory_upload_tests.py — Phase 8G tests for the inventory upload workflow.

Covers: valid upload, invalid sheet, missing columns, duplicate registrations,
empty inventory, rollback, successful refresh, and status. Uses TEMP copies of
IVR_Sheet.xlsx only — the real inventory file is never touched. New file; no
existing test is modified.
"""
from __future__ import annotations

import os
import shutil
import tempfile
import unittest

import openpyxl

import inventory_upload as IU
from inventory_loader import DNJ_SHEET, load_inventory, customer_facing

HERE = os.path.dirname(os.path.abspath(__file__))
REAL_XLSX = os.path.join(HERE, "..", "IVR_Sheet.xlsx")


def _multipart(file_bytes: bytes, filename: str = "IVR_Sheet.xlsx"):
    boundary = "----iuTestBoundary42"
    pre = (f"--{boundary}\r\n"
           f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
           f"Content-Type: application/octet-stream\r\n\r\n").encode()
    post = f"\r\n--{boundary}--\r\n".encode()
    return pre + file_bytes + post, f"multipart/form-data; boundary={boundary}"


class _FakeService:
    """Minimal stand-in: xlsx_path + inventory_count + a refresh that re-loads
    from the current file (using the REAL loader) — and can fail once."""
    def __init__(self, xlsx, fail_once=False):
        self.xlsx_path = xlsx
        self.inventory_count = len(customer_facing(load_inventory(xlsx)))
        self._fail_once = fail_once
        self.refresh_calls = 0

    def refresh_inventory(self):
        self.refresh_calls += 1
        if self._fail_once and self.refresh_calls == 1:
            raise RuntimeError("simulated refresh failure")
        self.inventory_count = len(customer_facing(load_inventory(self.xlsx_path)))
        return {"status": "ok", "inventory_count": self.inventory_count, "sync": {}}


class UploadTestBase(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.xlsx = os.path.join(self.dir, "IVR_Sheet.xlsx")
        shutil.copy2(REAL_XLSX, self.xlsx)            # live file = temp copy

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def _make_variant(self, mutate) -> bytes:
        """Load the real xlsx, apply `mutate(wb)`, return the bytes."""
        wb = openpyxl.load_workbook(REAL_XLSX)
        mutate(wb)
        out = os.path.join(self.dir, "variant.xlsx")
        wb.save(out)
        with open(out, "rb") as f:
            return f.read()


class TestValidUpload(UploadTestBase):
    def test_valid_upload_refreshes(self):
        svc = _FakeService(self.xlsx)
        before = svc.inventory_count
        with open(REAL_XLSX, "rb") as f:
            body, ct = _multipart(f.read())
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 200, payload)
        self.assertEqual(payload["status"], "ok")
        self.assertGreater(payload["vehicles_loaded"], 0)
        self.assertEqual(payload["previous_count"], before)
        self.assertIsNotNone(payload["backup"])               # backup created
        self.assertTrue(os.path.isdir(os.path.join(self.dir, "inventory_backups")))
        self.assertEqual(svc.inventory_count, payload["vehicles_loaded"])  # refresh ran

    def test_status(self):
        svc = _FakeService(self.xlsx)
        status, payload = IU.handle_status(svc)
        self.assertEqual(status, 200)
        self.assertEqual(payload["current_file"], "IVR_Sheet.xlsx")
        self.assertGreater(payload["inventory_count"], 0)
        self.assertIsNotNone(payload["last_updated"])


class TestRejections(UploadTestBase):
    def test_invalid_sheet(self):
        body = self._make_variant(lambda wb: [wb.remove(wb[DNJ_SHEET])])
        body, ct = _multipart(body)
        svc = _FakeService(self.xlsx)
        before = svc.inventory_count
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 400)
        self.assertFalse(payload["live_inventory_changed"])
        self.assertTrue(any("DNJ" in e for e in payload["errors"]))
        self.assertEqual(svc.inventory_count, before)         # untouched
        self.assertEqual(svc.refresh_calls, 0)                # never refreshed

    def test_missing_columns(self):
        def mutate(wb):
            ws = wb[DNJ_SHEET]
            slim = wb.create_sheet("TMP")
            for r in ws.iter_rows(min_row=1, max_row=3, max_col=5, values_only=True):
                slim.append(list(r))
            wb.remove(ws)
            slim.title = DNJ_SHEET
        body, ct = _multipart(self._make_variant(mutate))
        svc = _FakeService(self.xlsx)
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 400)
        self.assertTrue(any("CAR NUMB" in e or "column" in e.lower() for e in payload["errors"]))
        self.assertEqual(svc.refresh_calls, 0)

    def test_duplicate_registrations_warns_not_blocks(self):
        # The loader dedups by registration deterministically, so a duplicate is a
        # WARNING (surfaced) and the upload still succeeds — blocking would reject
        # the real production sheet, which itself contains a benign duplicate.
        def mutate(wb):
            ws = wb[DNJ_SHEET]
            col = IU._CAR_NUMB_IDX + 1                          # 1-based column N
            vals = [(i, ws.cell(row=i, column=col).value) for i in range(1, ws.max_row + 1)]
            regs = [(i, v) for i, v in vals if v and str(v).strip()
                    and str(v).strip().upper() not in IU._HEADER_TOKENS]
            self.assertGreaterEqual(len(regs), 2)
            ws.cell(row=regs[1][0], column=col).value = regs[0][1]   # force a duplicate
        body, ct = _multipart(self._make_variant(mutate))
        svc = _FakeService(self.xlsx)
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 200, payload)
        self.assertTrue(any("Duplicate" in w for w in payload["warnings"]))
        self.assertGreater(payload["vehicles_loaded"], 0)

    def test_empty_inventory(self):
        def mutate(wb):
            ws = wb[DNJ_SHEET]
            col = IU._CAR_NUMB_IDX + 1
            for i in range(1, ws.max_row + 1):
                ws.cell(row=i, column=col).value = None         # wipe all registrations
        body, ct = _multipart(self._make_variant(mutate))
        svc = _FakeService(self.xlsx)
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 400)
        self.assertTrue(any("0 live cars" in e or "CAR NUMB" in e for e in payload["errors"]))
        self.assertEqual(svc.refresh_calls, 0)

    def test_non_xlsx_rejected(self):
        body, ct = _multipart(b"not really an excel", filename="cars.csv")
        svc = _FakeService(self.xlsx)
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 400)
        self.assertTrue(any(".xlsx" in e for e in payload["errors"]))


class TestRollback(UploadTestBase):
    def test_rollback_on_refresh_failure(self):
        svc = _FakeService(self.xlsx, fail_once=True)
        before = svc.inventory_count
        with open(REAL_XLSX, "rb") as f:
            body, ct = _multipart(f.read())
        status, payload = IU.handle_upload(svc, body, ct)
        self.assertEqual(status, 500)
        self.assertEqual(payload["status"], "error")
        self.assertFalse(payload["live_inventory_changed"])   # rolled back
        self.assertEqual(svc.refresh_calls, 2)                # failed, then restore-refresh
        self.assertEqual(svc.inventory_count, before)         # inventory restored
        self.assertTrue(os.path.exists(self.xlsx))            # live file intact
        self.assertGreater(len(customer_facing(load_inventory(self.xlsx))), 0)


if __name__ == "__main__":
    unittest.main(verbosity=2)
