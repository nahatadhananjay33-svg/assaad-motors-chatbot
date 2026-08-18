"""
inventory_edit.py
=================

Phase 9A → 9C — owner-facing inventory editing over the SAME server-side Excel
(single source of truth). The Excel is never uploaded wholesale; these handlers
update / append ONE row and then call the existing ``service.refresh_inventory()``.

Phase 9C makes the Add/Edit form **fully dynamic**: the field list is discovered
from the actual Excel columns at request time, so any column added to the Excel
automatically appears in Add / Edit / View — no code change. Nothing is hardcoded
beyond friendly labels for the loader's known core columns and the keyword rules
used to group fields into sections.

Scope: reuses inventory_loader (column map), inventory_upload.handle_status,
refresh_inventory(), and the media_sync atomic-save/lock helpers. It does NOT
touch chatbot / retrieval / media upload / Supabase / logging / security.

Endpoints (all admin-key gated by the existing /admin rule):
  GET  /admin/inventory/dashboard        summary + table (Phase 9B)
  GET  /admin/inventory/schema           dynamic grouped field schema
  POST /admin/inventory/get_car          {car_number} -> {values{colkey:val}}
  POST /admin/inventory/update_car       {car_number, values{colkey:val}}
  POST /admin/inventory/add_car          {values{colkey:val}}
  POST /admin/inventory/duplicate_audit  scan workbook -> duplicate_inventory_report.xlsx
  GET  /admin/inventory/vehicles         (kept) simple list
"""

from __future__ import annotations

import json
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

import inventory_upload
from inventory_loader import COL, DATA_START_ROW, DNJ_SHEET

SOLD_SHEET = "SOLD_CARS"

# media_sync atomic-save / lock / excel-open helpers (same as media_admin.py)
_MS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "media_sync")
if _MS_DIR not in sys.path:
    sys.path.insert(0, _MS_DIR)
from _util import FileLock, atomic_save_workbook, excel_open_by_user  # noqa: E402

HEADER_ROW = 2          # loader header row (short labels)
DESC_ROW = 3            # loader description row

# Friendly labels + groups for the loader's known core columns (cols 1..17).
# Derived from inventory_loader.COL — reuse, not a second source of truth for the
# field LIST (the field list still comes from the live Excel).
_CORE = {  # loader key -> (label, group)
    "stock_no":     ("Stock No",        "Vehicle Information"),
    "sr_no":        ("Sr No",           "Vehicle Information"),
    "make":         ("Company / Make",  "Vehicle Information"),
    "model":        ("Model",           "Vehicle Information"),
    "year":         ("Year",            "Vehicle Information"),
    "insurance":    ("Insurance (date)","Documents"),
    "variant":      ("Variant",         "Vehicle Information"),
    "fuel":         ("Fuel Type",       "Vehicle Information"),
    "transmission": ("Transmission",    "Vehicle Information"),
    "ownership":    ("Owners",          "Vehicle Information"),
    "km":           ("KM Driven",       "Vehicle Details"),
    "color":        ("Colour",          "Vehicle Information"),
    "rate":         ("Rate (Rs)",       "Pricing"),
    "car_numb":     ("CAR NUMB",        "Vehicle Information"),
    "reg_last4":    ("Reg Last 4",      "Vehicle Information"),
    "location":     ("Location",        "Vehicle Information"),
    "rto":          ("RTO",             "Documents"),
}
CORE_LABEL = {COL[k] + 1: v[0] for k, v in _CORE.items() if k in COL}
CORE_GROUP = {COL[k] + 1: v[1] for k, v in _CORE.items() if k in COL}
CAR_COL = COL["car_numb"] + 1           # 14

_MEDIA_TYPES = ("EXTERIOR", "INTERIOR", "VIDEO", "INSTAGRAM", "YOUTUBE")
_READONLY_KW = ("PHOTO_URLS", "VIDEO_URLS", "YOUTUBE_URL", "INSTAGRAM_URL",
                "MEDIA_FOLDER", "PHOTO COUNT", "VIDEO COUNT", "STATUS", "LAST_UPDATED")

GROUP_ORDER = ["Vehicle Information", "Pricing", "Vehicle Details",
               "Documents", "Media", "Status", "Other Information"]

_NUM_RE = re.compile(r"^-?\d+(?:,\d{3})*(?:\.\d+)?$")

# Owner-editable link slots (Instagram/YouTube) accept only http(s) URLs — the same
# rule the media loader uses, so a saved value is guaranteed to reach the chatbot.
_URL_OK = re.compile(r"^https?://", re.IGNORECASE)
_LINK_LABEL = {"INSTAGRAM": "Instagram", "YOUTUBE": "YouTube"}


# ─────────────────────────────────────────────────────────────────────────────
# small helpers
# ─────────────────────────────────────────────────────────────────────────────
def _norm_reg(value: Any) -> str:
    return ("" if value is None else str(value)).strip().upper().replace(" ", "")


def _coerce(raw: Any) -> Any:
    """Blank -> None. Clean numeric string -> int/float. Else the raw string."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "":
        return None
    if _NUM_RE.match(s):
        s2 = s.replace(",", "")
        try:
            f = float(s2)
            return int(f) if f.is_integer() else f
        except ValueError:
            return s
    return s


def _sheet(wb):
    return wb[DNJ_SHEET] if DNJ_SHEET in wb.sheetnames else wb.active


def _find_row(ws, car_number: str) -> Optional[int]:
    target = _norm_reg(car_number)
    if not target:
        return None
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        if _norm_reg(ws.cell(r, CAR_COL).value) == target:
            return r
    return None


def _last_data_row(ws) -> int:
    last = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
        if _norm_reg(ws.cell(r, CAR_COL).value):
            last = r
    return last


def _header_map(ws, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            m[str(v).strip().upper()] = c
    return m


def _backup_dir(xlsx_path: str) -> str:
    return os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), "inventory_backups")


def _backup_keep() -> int:
    """How many recent timestamped backups to retain (configurable). 0/negative
    disables pruning. Default 30 — enough to roll back several edits without
    letting the backup folder grow without bound."""
    try:
        return int(os.getenv("INVENTORY_BACKUP_KEEP", "") or 30)
    except ValueError:
        return 30


def _prune_backups(xlsx_path: str) -> None:
    """Keep only the most recent N backups of this workbook in inventory_backups/.
    The backup suffix is a zero-padded, sortable UTC stamp, so a lexical sort is
    chronological. Fully guarded — backup hygiene must never break an edit."""
    try:
        keep = _backup_keep()
        bdir = _backup_dir(xlsx_path)
        if keep <= 0 or not os.path.isdir(bdir):
            return
        import glob
        base = os.path.basename(xlsx_path)
        files = sorted(glob.glob(os.path.join(bdir, base + ".*.bak")))
        for old in files[:-keep]:
            try:
                os.remove(old)
            except OSError:
                pass
    except Exception:
        pass


def _lock_path(xlsx_path: str) -> str:
    return xlsx_path + ".lock"


_AUDIT_LOG = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "data", "admin_audit.log")


def _audit(action: str, car_number: str, detail: Any = None) -> None:
    """Append one JSON line per staff inventory action (edit/add/restore).
    Mirrors media_audit for media actions — answers 'who changed what, when'."""
    try:
        os.makedirs(os.path.dirname(_AUDIT_LOG), exist_ok=True)
        from datetime import datetime, timezone
        entry = {"ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                 "action": action, "car_number": car_number, "detail": detail}
        with open(_AUDIT_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass                                    # auditing must never break the action


def _excel_open_response(xlsx_path: str) -> Optional[Tuple[int, Dict[str, Any]]]:
    if excel_open_by_user(xlsx_path):
        return 409, {"status": "deferred",
                     "detail": "The Excel file is open in Excel. Close it and retry."}
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Dynamic schema discovery — the heart of Phase 9C
# ─────────────────────────────────────────────────────────────────────────────
def _label_for(col: int, r2: str, r3: str, media_base: Optional[str]) -> str:
    if col in CORE_LABEL:
        return CORE_LABEL[col]
    r2c = (r2 or "").strip()
    r3c = (r3 or "").strip()
    if r2c and r2c.replace(".", "").isdigit():          # numbered media slot
        return (media_base + " " + r2c) if media_base else ("Slot " + r2c)
    if r2c and len(r2c) > 4 and not r2c.replace(".", "").isdigit():
        return r2c.split("(")[0].strip()
    if r3c:
        return r3c.split("(")[0].strip()
    if r2c:
        return r2c
    return "Column " + openpyxl.utils.get_column_letter(col)


def _group_and_ro(col: int, r2u: str, r3u: str, in_media: bool,
                  media_group: Optional[str] = None) -> Tuple[str, bool]:
    """Return (group, read_only) for a column, from keyword rules on the header."""
    text = (r2u + " " + r3u).strip()
    if col == CAR_COL:
        return CORE_GROUP.get(col, "Vehicle Information"), True   # identity key
    is_mirror = any(k in text for k in ("PHOTO_URLS", "VIDEO_URLS", "YOUTUBE_URL",
                                        "INSTAGRAM_URL", "MEDIA_FOLDER",
                                        "PHOTO COUNT", "VIDEO COUNT"))
    if in_media or is_mirror:
        # Instagram / YouTube LINK slots are owner-editable (plain URLs the chatbot
        # reads straight off the row). Photo/video slots and the consolidated mirror
        # columns stay read-only — those are managed via the Media panel / Supabase
        # upload, which we do NOT change.
        link_slot = (media_group in ("INSTAGRAM", "YOUTUBE")) and not is_mirror
        return "Media", (not link_slot)
    if r2u == "STATUS" or "LAST_UPDATED" in r2u:        # exact: not "RC STATUS"
        return "Status", True
    if col in CORE_GROUP:
        return CORE_GROUP[col], False
    if any(k in text for k in ("RATE", "PRICE", "NEGOTIABLE", "EMI", "DOWN PAYMENT")):
        return "Pricing", False
    if any(k in text for k in ("INSURANCE", "RTO", "RC STATUS", "HYPOTHECATION",
                               "LOAN", "NOC", "FINANCE", "WARRANTY", "ZERO DEP")):
        return "Documents", False
    if any(k in text for k in ("KM", "KILOMET", "MILEAGE", "ACCIDENT", "FLOOD",
                               "REPAINT", "CONDITION", "SERVICE", "REASON FOR SALE",
                               "BEST FEATURES", "KNOWN ISSUES", "OWNERSHIP")):
        return "Vehicle Details", False
    if any(k in text for k in ("COMPANY", "MAKE", "MODEL", "MANUFACTUR", "VARIANT",
                               "VARI", "FUEL", "TRANSMISSION", "COLOUR", "COLOR",
                               "CAR NUMB", "LAST 4", "LOCATION", "STOCK", "SR NO",
                               "YEAR")):
        return "Vehicle Information", False
    return "Other Information", False


def _discover_fields(ws) -> List[Dict[str, Any]]:
    """Build the ordered field list from the live Excel columns."""
    fields: List[Dict[str, Any]] = []
    media_group: Optional[str] = None
    slot_num = 0
    for c in range(1, (ws.max_column or 0) + 1):
        r2 = str(ws.cell(HEADER_ROW, c).value or "").strip()
        r3 = str(ws.cell(DESC_ROW, c).value or "").strip()
        r2u, r3u = r2.upper(), r3.upper()
        is_num_slot = bool(r2) and r2.replace(".", "").isdigit()

        started = next((k for k in _MEDIA_TYPES if r2u.startswith(k)), None)
        if started:
            media_group = started
            slot_num = 1
        elif media_group is not None and is_num_slot:
            slot_num += 1                               # continuation cell of a media run
        elif r2 and not is_num_slot:
            media_group = None                          # a named header ends media block
            slot_num = 0
        in_media = media_group is not None and (started is not None or is_num_slot)

        # include a column if it has any header OR is a known core column
        if not r2 and not r3 and c not in CORE_LABEL:
            continue

        media_base = media_group.title() if media_group else None
        label = _label_for(c, r2, r3, media_base)
        group, read_only = _group_and_ro(c, r2u, r3u, in_media, media_group)
        # Instagram / YouTube LINK slots become owner-editable here; flag them so the
        # editor UI groups them and the save path validates their URL format.
        is_link = (not read_only) and in_media and media_group in ("INSTAGRAM", "YOUTUBE")
        if is_link:
            label = "%s %d" % (_LINK_LABEL.get(media_group, media_group.title()),
                               slot_num or 1)
        # numeric type hint from a sample data cell
        sample = ws.cell(DATA_START_ROW, c).value
        ftype = "number" if isinstance(sample, (int, float)) else "text"
        fields.append({
            "key": "c%d" % c, "col": c, "label": label, "group": group,
            "editable": not read_only, "type": ftype,
            "header": r2,          # Phase 12C: raw row-2 header (UI control mapping)
            "is_link": is_link,    # Instagram/YouTube URL slot (owner-editable)
        })
    return fields


def _grouped(fields: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for f in fields:
        groups.setdefault(f["group"], []).append(f)
    ordered = []
    for g in GROUP_ORDER:
        if groups.get(g):
            ordered.append({"name": g, "fields": groups[g]})
    for g, fs in groups.items():                        # any group not in GROUP_ORDER
        if g not in GROUP_ORDER:
            ordered.append({"name": g, "fields": fs})
    return ordered


def handle_schema(service: Any) -> Tuple[int, Dict[str, Any]]:
    wb = openpyxl.load_workbook(service.xlsx_path, data_only=True)
    try:
        fields = _discover_fields(_sheet(wb))
    finally:
        wb.close()
    return 200, {"groups": _grouped(fields),
                 "car_numb_key": "c%d" % CAR_COL,
                 "field_count": len(fields)}


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard (Phase 9B) — summary + table. Reads a few fixed core columns.
# ─────────────────────────────────────────────────────────────────────────────
def _row_view(ws, r: int, status: str, lu_col: Optional[int]) -> Dict[str, Any]:
    def cell(key):
        v = ws.cell(r, COL[key] + 1).value
        return "" if v is None else v
    return {
        "registration_no": _norm_reg(ws.cell(r, CAR_COL).value),
        "brand": cell("make"), "model": cell("model"), "year": cell("year"),
        "km": cell("km"), "price": cell("rate"), "status": status,
        "last_updated": (ws.cell(r, lu_col).value if lu_col and ws.cell(r, lu_col).value else ""),
    }


def handle_dashboard(service: Any) -> Tuple[int, Dict[str, Any]]:
    _code, status = inventory_upload.handle_status(service)
    wb = openpyxl.load_workbook(service.xlsx_path, data_only=True)
    try:
        vehicles: List[Dict[str, Any]] = []
        available = reserved = sold = 0
        seen: set = set()
        if DNJ_SHEET in wb.sheetnames:
            ws = wb[DNJ_SHEET]
            hmap = _header_map(ws, HEADER_ROW)
            status_c, lu_c = hmap.get("STATUS"), hmap.get("LAST_UPDATED")
            for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
                reg = _norm_reg(ws.cell(r, CAR_COL).value)
                if not reg or reg in seen:
                    continue
                seen.add(reg)
                st = (str(ws.cell(r, status_c).value).strip().upper()
                      if status_c and ws.cell(r, status_c).value else "AVAILABLE")
                if st == "SOLD":
                    continue
                if st == "RESERVED":
                    reserved += 1
                else:
                    st = "AVAILABLE"
                    available += 1
                vehicles.append(_row_view(ws, r, st, lu_c))
        if SOLD_SHEET in wb.sheetnames:
            ws = wb[SOLD_SHEET]
            lu_c = _header_map(ws, 1).get("LAST_UPDATED")
            for r in range(2, (ws.max_row or 0) + 1):
                reg = _norm_reg(ws.cell(r, CAR_COL).value)
                if not reg:
                    continue
                sold += 1
                vehicles.append(_row_view(ws, r, "SOLD", lu_c))
        summary = {"live": status.get("inventory_count"), "available": available,
                   "reserved": reserved, "sold": sold,
                   "last_updated": status.get("last_updated"),
                   "current_file": status.get("current_file")}
        return 200, {"summary": summary, "vehicles": vehicles}
    finally:
        wb.close()


def handle_vehicles(service: Any) -> Tuple[int, Dict[str, Any]]:
    """Kept for compatibility — simple {car_number, model, year, rate} list."""
    wb = openpyxl.load_workbook(service.xlsx_path, data_only=True)
    try:
        ws = _sheet(wb)
        out = []
        for r in range(DATA_START_ROW, (ws.max_row or 0) + 1):
            reg = _norm_reg(ws.cell(r, CAR_COL).value)
            if not reg:
                continue
            out.append({"car_number": reg,
                        "model": (str(ws.cell(r, COL["model"] + 1).value).strip()
                                  if ws.cell(r, COL["model"] + 1).value else ""),
                        "year": ws.cell(r, COL["year"] + 1).value or "",
                        "rate": ws.cell(r, COL["rate"] + 1).value or ""})
        return 200, {"vehicles": out}
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# get / update / add — dynamic, keyed by column ("c<col>")
# ─────────────────────────────────────────────────────────────────────────────
def _parse_json(body: bytes) -> Optional[Dict[str, Any]]:
    try:
        p = json.loads(body.decode("utf-8") or "{}")
        return p if isinstance(p, dict) else None
    except Exception:
        return None


def _read_values(ws, row: int, fields: List[Dict[str, Any]]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for f in fields:
        v = ws.cell(row, f["col"]).value
        out[f["key"]] = "" if v is None else v
    return out


def handle_get_car(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    payload = _parse_json(body)
    if payload is None:
        return 400, {"status": "error", "detail": "Body must be a JSON object."}
    car_number = payload.get("car_number") or ""
    if not car_number:
        return 400, {"status": "error", "detail": "Missing 'car_number'."}
    wb = openpyxl.load_workbook(service.xlsx_path, data_only=True)
    try:
        ws = _sheet(wb)
        row = _find_row(ws, car_number)
        if row is None:
            return 404, {"status": "error", "detail": "Car not found."}
        fields = _discover_fields(ws)
        return 200, {"status": "ok", "car_number": _norm_reg(car_number),
                     "values": _read_values(ws, row, fields)}
    finally:
        wb.close()


def _editable_cols(ws) -> Dict[int, bool]:
    """col -> editable? (from the same discovery used by the schema)."""
    return {f["col"]: f["editable"] for f in _discover_fields(ws)}


def _bad_link_urls(values: Dict[str, Any], link_cols: set) -> List[str]:
    """Return the c-keys whose value targets an Instagram/YouTube slot but is not a
    valid http(s) URL. Blank is allowed (it clears the link). Used to reject BEFORE
    writing, so the workbook is never partially saved with a bad link."""
    bad: List[str] = []
    for key, val in values.items():
        if not key.startswith("c"):
            continue
        try:
            col = int(key[1:])
        except ValueError:
            continue
        if col in link_cols:
            s = "" if val is None else str(val).strip()
            if s and not _URL_OK.match(s):
                bad.append(key)
    return bad


def handle_update_car(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    payload = _parse_json(body)
    if payload is None:
        return 400, {"status": "error", "detail": "Body must be a JSON object."}
    car_number = payload.get("car_number") or ""
    values = payload.get("values") or payload.get("fields") or {}
    if not car_number:
        return 400, {"status": "error", "detail": "Missing 'car_number'."}
    if not isinstance(values, dict) or not values:
        return 400, {"status": "error", "detail": "No fields to update."}

    xlsx = service.xlsx_path
    busy = _excel_open_response(xlsx)
    if busy:
        return busy
    with FileLock(_lock_path(xlsx)):
        wb = openpyxl.load_workbook(xlsx)
        try:
            ws = _sheet(wb)
            row = _find_row(ws, car_number)
            if row is None:
                return 404, {"status": "error", "detail": "Car not found."}
            fields = _discover_fields(ws)
            editable = {f["col"]: f["editable"] for f in fields}
            link_cols = {f["col"] for f in fields if f.get("is_link")}
            bad = _bad_link_urls(values, link_cols)     # validate BEFORE any write
            if bad:
                return 400, {"status": "error",
                             "detail": "Invalid link URL — use a full https:// link "
                                       "(or leave it blank to remove).",
                             "invalid": bad}
            for key, val in values.items():
                if not key.startswith("c"):
                    continue
                try:
                    col = int(key[1:])
                except ValueError:
                    continue
                if col == CAR_COL:                      # CAR NUMB is read-only
                    continue
                if not editable.get(col, False):        # skip media/system cols
                    continue
                ws.cell(row, col).value = _coerce(val)
            atomic_save_workbook(wb, xlsx, backup_dir=_backup_dir(xlsx))
            _prune_backups(xlsx)
        finally:
            wb.close()
    refresh = service.refresh_inventory()
    _audit("update_car", _norm_reg(car_number),
           {"fields_changed": sorted(k for k in values if k.startswith("c"))})
    return 200, {"status": "ok", "car_number": _norm_reg(car_number),
                 "message": "Car updated.", "refresh": refresh}


def handle_add_car(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    payload = _parse_json(body)
    if payload is None:
        return 400, {"status": "error", "detail": "Body must be a JSON object."}
    values = payload.get("values") or payload.get("fields") or {}
    if not isinstance(values, dict):
        return 400, {"status": "error", "detail": "'values' must be an object."}
    car_key = "c%d" % CAR_COL
    car_number = _norm_reg(values.get(car_key) or values.get("car_numb"))
    if not car_number:
        return 400, {"status": "error", "detail": "CAR NUMB is required."}

    xlsx = service.xlsx_path
    busy = _excel_open_response(xlsx)
    if busy:
        return busy
    with FileLock(_lock_path(xlsx)):
        wb = openpyxl.load_workbook(xlsx)
        try:
            ws = _sheet(wb)
            if _find_row(ws, car_number) is not None:
                return 409, {"status": "error", "detail": "This vehicle already exists.",
                             "car_number": car_number, "exists": True}
            new_row = _last_data_row(ws) + 1
            fields = _discover_fields(ws)
            editable = {f["col"]: f["editable"] for f in fields}
            link_cols = {f["col"] for f in fields if f.get("is_link")}
            bad = _bad_link_urls(values, link_cols)     # validate BEFORE any write
            if bad:
                return 400, {"status": "error",
                             "detail": "Invalid link URL — use a full https:// link "
                                       "(or leave it blank).",
                             "invalid": bad}
            for key, val in values.items():
                if not key.startswith("c"):
                    continue
                try:
                    col = int(key[1:])
                except ValueError:
                    continue
                if col != CAR_COL and not editable.get(col, False):
                    continue                            # skip media/system on add too
                ws.cell(new_row, col).value = _coerce(val)
            ws.cell(new_row, CAR_COL).value = car_number    # ensure normalised reg
            atomic_save_workbook(wb, xlsx, backup_dir=_backup_dir(xlsx))
            _prune_backups(xlsx)
        finally:
            wb.close()
    refresh = service.refresh_inventory()
    _audit("add_car", car_number,
           {"fields_set": sorted(k for k in values if k.startswith("c"))})
    return 200, {"status": "ok", "car_number": car_number,
                 "message": "Car added.", "refresh": refresh}


# ─────────────────────────────────────────────────────────────────────────────
# Restore (un-sold) — move a row from SOLD_CARS back to the live DNJ sheet.
# Additive inverse of media_admin's Mark Sold; media stays deleted (re-upload
# later via the Media panel). Does NOT touch the media/Supabase/sold code.
# ─────────────────────────────────────────────────────────────────────────────
def handle_restore_car(service: Any, body: bytes) -> Tuple[int, Dict[str, Any]]:
    payload = _parse_json(body)
    if payload is None:
        return 400, {"status": "error", "detail": "Body must be a JSON object."}
    car_number = _norm_reg(payload.get("car_number"))
    if not car_number:
        return 400, {"status": "error", "detail": "Missing 'car_number'."}

    xlsx = service.xlsx_path
    busy = _excel_open_response(xlsx)
    if busy:
        return busy
    with FileLock(_lock_path(xlsx)):
        wb = openpyxl.load_workbook(xlsx)
        try:
            if SOLD_SHEET not in wb.sheetnames:
                return 404, {"status": "error", "detail": "No SOLD_CARS sheet."}
            sold_ws = wb[SOLD_SHEET]
            srow = None
            for r in range(2, (sold_ws.max_row or 0) + 1):
                if _norm_reg(sold_ws.cell(r, CAR_COL).value) == car_number:
                    srow = r
                    break
            if srow is None:
                return 404, {"status": "error", "detail": "Car not found in SOLD_CARS."}
            dnj = wb[DNJ_SHEET] if DNJ_SHEET in wb.sheetnames else wb.active
            if _find_row(dnj, car_number) is not None:
                return 409, {"status": "error", "detail": "Car is already active."}
            new_row = _last_data_row(dnj) + 1
            maxc = max(sold_ws.max_column or 0, dnj.max_column or 0)
            for c in range(1, maxc + 1):
                dnj.cell(new_row, c).value = sold_ws.cell(srow, c).value
            status_c = _header_map(dnj, HEADER_ROW).get("STATUS")
            if status_c:
                dnj.cell(new_row, status_c).value = "AVAILABLE"
            sold_ws.delete_rows(srow, 1)
            atomic_save_workbook(wb, xlsx, backup_dir=_backup_dir(xlsx))
            _prune_backups(xlsx)
        finally:
            wb.close()
    refresh = service.refresh_inventory()
    _audit("restore_car", car_number, "moved from SOLD_CARS back to DNJ")
    return 200, {"status": "ok", "car_number": car_number,
                 "message": "Car restored to Available. Re-upload media if needed.",
                 "refresh": refresh}


# ─────────────────────────────────────────────────────────────────────────────
# Duplicate audit — scan the workbook, write duplicate_inventory_report.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def handle_duplicate_audit(service: Any) -> Tuple[int, Dict[str, Any]]:
    xlsx = service.xlsx_path
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    dupes: Dict[str, List[Dict[str, Any]]] = {}
    known_sheets = {DNJ_SHEET, SOLD_SHEET, "DONT TOUCH SOLD"}
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # locate CAR NUMB on this sheet; only scan sheets that plausibly hold
            # registrations (a real CAR NUMB header, or a known inventory sheet).
            car_c = None
            for hr in (1, 2, 3):
                hm = _header_map(ws, hr)
                if "CAR NUMB" in hm:
                    car_c = hm["CAR NUMB"]
                    break
            if car_c is None:
                if sheet not in known_sheets:
                    continue
                car_c = CAR_COL
            start = DATA_START_ROW if sheet == DNJ_SHEET else 2
            seen: Dict[str, List[Tuple[str, int, str]]] = {}
            for r in range(start, (ws.max_row or 0) + 1):
                reg = _norm_reg(ws.cell(r, car_c).value)
                # registration-like only: has a digit and a letter, length >= 6
                if not reg or len(reg) < 6 or not any(ch.isdigit() for ch in reg) \
                        or not any(ch.isalpha() for ch in reg):
                    continue
                model = ws.cell(r, COL["model"] + 1).value
                seen.setdefault(reg, []).append((sheet, r, str(model or "")))
            for reg, rows in seen.items():
                if len(rows) > 1:
                    dupes.setdefault(reg, []).extend(rows)
    finally:
        wb.close()

    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Duplicates"
    sh.append(["CAR NUMB", "Occurrences", "Sheet", "Row", "Model", "Suggested Action"])
    total = 0
    for reg, rows in sorted(dupes.items()):
        for i, (sheet, r, model) in enumerate(rows):
            total += 1
            action = ("KEEP (first occurrence) — review"
                      if i == 0 else "REVIEW / MERGE / REMOVE this duplicate row")
            sh.append([reg, len(rows), sheet, r, model, action])
    report_path = os.path.join(os.path.dirname(os.path.abspath(xlsx)),
                               "duplicate_inventory_report.xlsx")
    out.save(report_path)
    out.close()
    return 200, {"status": "ok",
                 "duplicate_registrations": len(dupes),
                 "duplicate_rows": total,
                 "report": os.path.basename(report_path),
                 "report_path": report_path,
                 "duplicates": {reg: rows for reg, rows in dupes.items()}}
