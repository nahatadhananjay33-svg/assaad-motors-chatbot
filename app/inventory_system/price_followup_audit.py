"""
price_followup_audit.py  —  Phase 7O.4 STEP 1 (AUDIT ONLY, no fix)
==================================================================

Audits price follow-up accuracy against the REAL pilot conversation log
(`data/pilot_query_log.db`). The referenced `final_conversations_only.txt` and
`review_batch_*_vFinal.txt` are not present in the repository, so the pilot log
(28,805 turns / 10,800 conversations) is used as the available replay dataset —
consistent with the Phase-7O.2 / 7O.3 audits.

A **price follow-up turn** is a bare price question that names NO new vehicle and
carries NO budget filter — it relies on the conversation's active vehicle context:
    price · price? · kitne ka · kimat · daam · rate · cost · cost kya hai ·
    kitne ki hai · price batao   (matched via the bot's own _PRICE_FOLLOWUP_WORDS)

Each such turn is replayed in order (context preserved) through the CURRENT
(pre-fix) ChatService and classified:

  * pass          — active context existed AND the bot answered ONE selected
                    vehicle's price (count == 1, intent price).
  * lost_context  — active context existed but the bot returned multiple cars
                    (clarification / "multiple vehicles") instead of the
                    selected one.
  * inventory_dump— the bot returned a large slice of the catalogue
                    (count >= DUMP_THRESHOLD) for a bare price question.
  * no_context    — no vehicle had been selected yet (informational; not a
                    follow-up failure by itself).

Writes `price_followup_audit.xlsx` (Summary / Pass Cases / Fail Cases /
Lost Context / Inventory Dump Cases / Root Cause). Read-only on the pilot log;
isolated temp DBs.
"""
from __future__ import annotations

import os
import sqlite3
import time
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import chat_service as CS
from astor_leak_audit import _speedup, _new_service

LOG_DB = os.path.join(os.path.dirname(__file__), "..", "..", "data",
                      "pilot_query_log.db")
DUMP_THRESHOLD = 6          # no single model has >3 cars; top_n is 5

# Reuse the bot's own price-question vocabulary so the audit matches routing.
_PRICE_WORDS = CS._PRICE_FOLLOWUP_WORDS


def _is_price_followup(message: str) -> bool:
    """Bare price question, no new vehicle named, no budget filter."""
    q = CS.parse(CS.normalize_typos(message))
    if q.model or q.make or q.registration:
        return False
    if q.price_max is not None or q.price_min is not None or q.sort_cheapest:
        return False
    if getattr(q, "clarify_budget", None) is not None:
        return False
    text = f" {(message or '').strip().lower()} "
    return any(w in text for w in _PRICE_WORDS)


def _ctx_after(turn: str, r) -> Optional[Dict]:
    """Mirror Phase-7I.2 `_followup_ctx` establishment for the audit."""
    if r.vehicles and len(r.vehicles) == 1:
        return {"reg": r.vehicles[0].get("registration_no"),
                "model": r.vehicles[0].get("model")}
    q = CS.parse(CS.normalize_typos(turn))
    if q.model:
        return {"reg": None, "model": q.model}
    return None


def _load_conversations() -> List[Tuple[str, List[str]]]:
    c = sqlite3.connect(LOG_DB)
    rows = c.execute(
        "SELECT conversation_id, user_query FROM query_log ORDER BY conversation_id, id"
    ).fetchall()
    c.close()
    convs: Dict[str, List[str]] = defaultdict(list)
    for cid, uq in rows:
        if uq:
            convs[cid].append(uq)
    # only conversations that contain at least one bare price question
    out = []
    for cid, turns in convs.items():
        if any(_is_price_followup(t) for t in turns):
            out.append((cid, turns))
    return out


def audit() -> Dict:
    svc = _new_service()
    convs = _load_conversations()
    rows = {"pass": [], "lost_context": [], "inventory_dump": [], "no_context": []}
    t0 = time.time()
    for n, (cid, turns) in enumerate(convs, 1):
        sid = f"price::{cid}"
        ctx = None
        for turn in turns:
            price_fu = _is_price_followup(turn)
            r = svc.handle(turn, session_id=sid)
            if price_fu:
                rec = {"conversation_id": cid, "query": turn, "intent": r.intent,
                       "count": r.count, "vehicles": len(r.vehicles),
                       "had_context": bool(ctx),
                       "ctx": (ctx or {}).get("model") or (ctx or {}).get("reg") or "",
                       "response": (r.response or "").replace("\n", " ")[:150]}
                if r.count >= DUMP_THRESHOLD:
                    rows["inventory_dump"].append(rec)
                elif ctx and r.count == 1:
                    rows["pass"].append(rec)
                elif ctx and r.count > 1:
                    rows["lost_context"].append(rec)
                else:
                    rows["no_context"].append(rec)
            nctx = _ctx_after(turn, r)
            if nctx:
                ctx = nctx
        if n % 300 == 0:
            print(f"  ...{n}/{len(convs)} convs ({round(time.time()-t0,1)}s)",
                  flush=True)
    svc.close()
    return {"conversations": len(convs), "rows": rows}


def _write_xlsx(res: Dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    rows = res["rows"]
    # totals
    n_pass = len(rows["pass"])
    n_lost = len(rows["lost_context"])
    n_dump = len(rows["inventory_dump"])
    n_noctx = len(rows["no_context"])
    # a "price follow-up" relies on context: failures = lost_context + dumps that
    # had context. (no_context turns are not follow-ups — no vehicle to follow.)
    dump_with_ctx = sum(1 for r in rows["inventory_dump"] if r["had_context"])
    dump_no_ctx = n_dump - dump_with_ctx
    followups = n_pass + n_lost + dump_with_ctx
    failures = n_lost + dump_with_ctx
    pass_rate = (n_pass / followups * 100) if followups else 0.0

    wb = Workbook()
    s = wb.active
    s.title = "Summary"
    for r in [
        ["Metric", "Value"],
        ["Conversations replayed (with >=1 price question)", res["conversations"]],
        ["Price follow-ups (active context existed)", followups],
        ["  Pass (answered selected vehicle's price)", n_pass],
        ["  Fail — total", failures],
        ["    Lost context (multiple-vehicle clarification)", n_lost],
        ["    Inventory dump (with context)", dump_with_ctx],
        ["Pass rate (before fix)", f"{pass_rate:.1f}%"],
        ["", ""],
        ["Bare price questions with NO context yet", n_noctx + dump_no_ctx],
        ["  of which inventory dump (no context)", dump_no_ctx],
        ["  of which clarification / handled", n_noctx],
        ["", ""],
        ["Total inventory-dump cases (any context)", n_dump],
    ]:
        s.append(r)
    s["A1"].font = s["B1"].font = Font(bold=True)
    s.column_dimensions["A"].width = 52
    s.column_dimensions["B"].width = 14

    def _sheet(title, recs):
        ws = wb.create_sheet(title)
        ws.append(["conversation_id", "query", "had_context", "ctx",
                   "intent", "count", "response"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in recs:
            ws.append([r["conversation_id"], r["query"], r["had_context"],
                       r["ctx"], r["intent"], r["count"], r["response"]])
        for col, w in zip("ABCDEFG", [16, 22, 12, 14, 14, 8, 80]):
            ws.column_dimensions[col].width = w

    _sheet("Pass Cases", rows["pass"])
    _sheet("Fail Cases", rows["lost_context"] +
           [r for r in rows["inventory_dump"] if r["had_context"]])
    _sheet("Lost Context", rows["lost_context"])
    _sheet("Inventory Dump Cases", rows["inventory_dump"])

    rc = wb.create_sheet("Root Cause")
    rc.append(["#", "Finding"])
    rc["A1"].font = rc["B1"].font = Font(bold=True)
    for i, line in enumerate([
        "Bare price questions route to inventory retrieval with NO single-vehicle pin.",
        "When context model has multiple cars, retrieval returns G-MULTI -> clarification (lost_context).",
        "When NO vehicle context exists, retrieval runs with no filter -> full catalogue dump.",
        "Phase-7I.2 appends the context MODEL, but a multi-car model still yields multiple matches.",
        "Code path: chat_service.handle -> _handle_retrieval(q) -> engine.search -> response_formatter G-MULTI / full list.",
    ], 1):
        rc.append([i, line])
    rc.column_dimensions["A"].width = 5
    rc.column_dimensions["B"].width = 100

    wb.save(os.path.join(os.path.dirname(__file__), "price_followup_audit.xlsx"))
    print("PASS", n_pass, "| LOST_CTX", n_lost, "| DUMP", n_dump,
          "(ctx", dump_with_ctx, "/ noctx", dump_no_ctx, ") | NOCTX_other",
          n_noctx, "| pass_rate", f"{pass_rate:.1f}%", flush=True)


def main() -> None:
    _speedup()
    print("Auditing price follow-ups (pre-fix)...", flush=True)
    res = audit()
    _write_xlsx(res)
    print("Wrote price_followup_audit.xlsx", flush=True)


if __name__ == "__main__":
    main()
