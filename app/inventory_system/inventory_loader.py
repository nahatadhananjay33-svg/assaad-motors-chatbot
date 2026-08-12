"""
inventory_loader.py
===================

Reads `IVR_Sheet.xlsx` → sheet **DNJ** (the inventory source of truth) and the
**DONT TOUCH SOLD** sheet, normalizes every dirty operator value, derives the
fields the retrieval layer needs, reconciles sold cars, and returns a clean list
of `InventoryItem` objects.

All mapping tables are grounded in:
  * Information extraction/02_inventory/inventory_field_dictionary.md
  * Information extraction/02_inventory/inventory_query_mapping.csv
  * Information extraction/02_inventory/inventory_retrieval_risks.md  (R01–R20)
  * Phase 1B discovery report (actual DNJ values)

This module is pure data work — no Supabase, no network. Sync lives in
`inventory_sync.py`.
"""

from __future__ import annotations

import re
import difflib
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl

from inventory_models import (
    InventoryItem,
    FuelType,
    Transmission,
    BodyType,
    ListingStatus,
    LocationType,
    ColorConfidence,
    utcnow_iso,
)
import model_specs                       # Phase 12B: standard-spec auto-fill

# ─────────────────────────────────────────────────────────────────────────────
# Sheet layout (0-based column indices in the row tuple). Header on row 2,
# description on row 3, DATA STARTS ON ROW 4. Columns A..Q + media R.. (empty).
# ─────────────────────────────────────────────────────────────────────────────
DATA_START_ROW = 4
COL = {
    "stock_no": 0,        # A
    "sr_no": 1,           # B
    "make": 2,            # C  (header blank; "Company Name")
    "model": 3,           # D  MODEL
    "year": 4,            # E  (header blank; manufacturing/passing)
    "insurance": 5,       # F  INS
    "variant": 6,         # G  VARI
    "fuel": 7,            # H  F
    "transmission": 8,    # I  T
    "ownership": 9,       # J  O
    "km": 10,             # K  KM
    "color": 11,          # L  COL
    "rate": 12,           # M  RATE
    "car_numb": 13,       # N  CAR NUMB
    "reg_last4": 14,      # O  (header blank; last 4 digits)
    "location": 15,       # P  LOC
    "rto": 16,            # Q  RTO
}

DNJ_SHEET = "DNJ"
SOLD_SHEET = "DONT TOUCH SOLD"


# ─────────────────────────────────────────────────────────────────────────────
# CODE TABLES  (field dictionary §3)
# ─────────────────────────────────────────────────────────────────────────────
MAKE_MAP: Dict[str, str] = {
    "MARU": "Maruti Suzuki",
    "TOYO": "Toyota",
    "HYUN": "Hyundai", "HUYN": "Hyundai",
    "HOND": "Honda", "HONDA": "Honda",
    "MAHI": "Mahindra",
    "TATA": "Tata",
    "MERC": "Mercedes-Benz",
    "VOX": "Volkswagen",
    "SKODA": "Škoda",
    "FORD": "Ford",
    "BMW": "BMW",
    "AUDI": "Audi",
    "JEEP": "Jeep",
    "KIA": "Kia",
    "RENA": "Renault",
    "NISS": "Nissan",
    "MG": "MG Motor",
    "SHEV": "Chevrolet", "SHEVR": "Chevrolet",
    "ISUZU": "Isuzu",
    "JAGUAR": "Jaguar",
    "RANGE": "Range Rover",
    "VOLV": "Volvo",
    "HMFC": "Mitsubishi",      # Hindustan Motors badge (e.g. Pajero)
    "CUST": "Custom",          # non-standard/placeholder listing
}

FUEL_MAP: Dict[str, str] = {
    "P": FuelType.PETROL,
    "SP": FuelType.PETROL,         # dirty single petrol entry
    "D": FuelType.DIESEL,
    "C": FuelType.CNG,
    "CR": FuelType.CNG,            # retrofitted CNG kit
    "PC": FuelType.PETROL_CNG,
    "CP": FuelType.PETROL_CNG,     # same as PC, order swapped
    "PH": FuelType.HYBRID,
    "E": FuelType.ELECTRIC,
}

TRANSMISSION_MAP: Dict[str, str] = {
    "A": Transmission.AUTOMATIC,   # covers AT/AMT/CVT/DCT/iMT
    "M": Transmission.MANUAL,
}

# canonical colour -> set of dirty codes seen in the sheet
_COLOR_VARIANTS: Dict[str, Set[str]] = {
    "White": {"WHI", "WHIT", "WHITE"},
    "Silver": {"SIL", "SILVE", "SILVER"},
    "Grey": {"GREY", "GRY", "GRE"},          # GRE -> Grey (low confidence)
    "Black": {"BLACK", "BLA"},
    "Blue": {"BLUE", "BLU"},
    "Red": {"RED"},
    "Brown": {"BRO", "BROW", "BRWN"},
    "Gold": {"GOLD", "GOL"},
    "Beige": {"BEIGE"},
    "Green": {"GREE"},
    "Champagne": {"CHAM"},
    "Orange": {"ORA"},
    "Platinum": {"PLAT"},
    "Red+Black": {"R+B"},
}
COLOR_MAP: Dict[str, str] = {
    code: canon for canon, codes in _COLOR_VARIANTS.items() for code in codes
}
AMBIGUOUS_COLOR_CODES = {"GRE", "GREE"}      # confidence = low
NON_COLOR_CODES = {"REP"}                    # placeholder, not a colour

# model alias folding (dirty spellings -> canonical). Keep sub-variants distinct.
MODEL_ALIASES: Dict[str, str] = {
    "BREZZA": "Brezza", "BREEZA": "Brezza",
    "SALTOS": "Seltos",
    "RARID": "Rapid", "RAPID": "Rapid",
    "MARAZOO": "Marazzo", "MARAZZO": "Marazzo",
    "MOBILO": "Mobilio", "MOBILIO": "Mobilio",
    "FIESTS": "Fiesta", "FIESTA": "Fiesta",
    "ECCO": "Eeco", "EECO": "Eeco",
    "ESTIOL": "Estilo", "ESTILO": "Estilo", "ZEN ESTIOL": "Zen Estilo",
    "DESIRE": "Dzire", "DZIRE": "Dzire",
    "CRYSTA": "Innova Crysta",
    "ECO SPORT": "EcoSport", "ECOSPORT": "EcoSport",
    "WAGON R": "WagonR", "WAGONR": "WagonR",
    "I 20": "i20", "I20": "i20", "I-20": "i20",
    "I 20 ELITE": "i20 Elite",
    "I 10": "i10", "I10": "i10",
    "I 10 GRAND": "Grand i10", "GRAND I 10": "Grand i10",
    "XCENT": "Xcent",
    "ALTIS": "Corolla Altis",
    "COROLLA": "Corolla",
    "C CLASS": "C Class", "E CLASS": "E Class",
}

# known-good model vocabulary for fuzzy fallback (canonical display names)
KNOWN_MODELS: List[str] = sorted(set(MODEL_ALIASES.values()) | {
    "Alto", "Swift", "WagonR", "Celerio", "Spark", "Tiago", "Kwid", "Santro",
    "Polo", "Brio", "Figo", "Punch", "Amaze", "City", "Verna", "Ciaz", "SX4",
    "Vento", "Manza", "Etios", "Sonata", "Elantra", "Yaris", "Superb", "Ameo",
    "A4", "A6", "Civic", "Creta", "Nexon", "Sonet", "Venue", "WRV", "KUV100",
    "TUV300", "Safari", "XUV500", "Fortuner", "Endeavour", "Compass", "Duster",
    "Terrano", "Q5", "X1", "Innova", "Ertiga", "Mobilio", "BRV", "Omni", "Eeco",
    "Jazz", "Astor", "Pajero", "CRV", "Xylo", "Quanto", "Fiesta", "Alto 800",
})

# ── model -> body type / seats  (field dictionary §6.1) ──
_HATCH = {"Alto", "Alto 800", "Swift", "WagonR", "Celerio", "Spark", "Tiago",
          "Kwid", "i10", "Grand i10", "i20", "i20 Elite", "Santro", "Polo",
          "Brio", "Figo", "Zen Estilo", "Estilo", "Jazz"}
_SEDAN = {"City", "Verna", "Ciaz", "SX4", "Dzire", "Amaze", "Vento", "Rapid",
          "Manza", "Etios", "Sonata", "Elantra", "Corolla Altis", "Yaris",
          "Superb", "Xcent", "Ameo", "A4", "A6", "Civic", "Fiesta",
          "C Class", "E Class", "Corolla", "B180", "320d"}
_COMPACT_SUV = {"Creta", "Nexon", "Brezza", "Sonet", "Seltos", "Astor", "Venue",
                "EcoSport", "WRV", "KUV100", "TUV300", "Punch"}
_SUV = {"Safari", "XUV500", "Fortuner", "Endeavour", "Compass", "Duster",
        "Terrano", "Q5", "X1", "CRV", "Pajero", "Quanto", "Xylo"}
_MUV = {"Innova", "Innova Crysta", "Ertiga", "Mobilio", "BRV", "Marazzo",
        "Omni", "Eeco"}
_LUXURY = {"A4", "A6", "Q5", "X1", "Pajero", "C Class", "E Class", "B180", "320d"}
SEVEN_SEATERS = _MUV | {"Safari", "XUV500", "Fortuner", "Endeavour", "Xylo"}


def _body_type_for(model: Optional[str]) -> str:
    if not model:
        return BodyType.UNKNOWN
    if model in _MUV:
        return BodyType.MUV
    if model in _COMPACT_SUV:
        return BodyType.COMPACT_SUV
    if model in _SUV:
        return BodyType.SUV
    if model in _SEDAN:
        return BodyType.SEDAN
    if model in _HATCH:
        return BodyType.HATCHBACK
    return BodyType.UNKNOWN


def _seats_for(model: Optional[str]) -> Optional[int]:
    if not model:
        return None
    if model in SEVEN_SEATERS:
        return 7
    if _body_type_for(model) != BodyType.UNKNOWN:
        return 5
    return None


# ── RATE status codes (field dictionary §4 / risk R01) ──
RATE_STATUS_CODES = {"4", "33", "44", "DDD", "DDDD"}
PRICE_FLOOR = 10000          # anything below is a code, not rupees

# ── LOCATION word-custody codes (field dictionary §5 / risk R17) ──
CUSTODY_CODES = {"ABR", "SAM", "SONI", "RAS", "RAJ", "DJ", "IMM", "WEST",
                 "NGP", "POLI", "ILL", "REP"}
SLOT_RE = re.compile(r"^[A-Z]\d+$")
IVR_PRICE_THRESHOLD = 200000  # > ₹2 lakh => IVR-eligible


# ─────────────────────────────────────────────────────────────────────────────
# CELL-LEVEL NORMALIZERS
# ─────────────────────────────────────────────────────────────────────────────
def _clean(value: Any) -> str:
    """Trim + collapse internal whitespace; '' for None/blank cells."""
    if value is None:
        return ""
    s = str(value).strip()
    return re.sub(r"\s+", " ", s)


def normalize_make(raw: Any) -> Tuple[Optional[str], Optional[str]]:
    """Return (make_code, make_full). Unknown codes pass through, full=None."""
    code = _clean(raw).upper()
    if not code:
        return None, None
    return code, MAKE_MAP.get(code)


def normalize_model(raw: Any) -> Optional[str]:
    """Alias-fold then fuzzy-fold dirty model spellings to a canonical name."""
    s = _clean(raw).upper()
    if not s:
        return None
    if s in MODEL_ALIASES:
        return MODEL_ALIASES[s]
    # title-case fallback; try fuzzy match against known vocabulary
    pretty = s.title()
    upper_known = {m.upper(): m for m in KNOWN_MODELS}
    if s in upper_known:
        return upper_known[s]
    match = difflib.get_close_matches(s, list(upper_known.keys()), n=1, cutoff=0.82)
    if match:
        return upper_known[match[0]]
    return pretty


def parse_year(raw: Any) -> Optional[int]:
    """First 4 digits = manufacture year. 0/blank/<1990 -> None (risk R10/R13)."""
    s = _clean(raw)
    if not s:
        return None
    m = re.match(r"^(\d{4})", s)
    if not m:
        return None
    year = int(m.group(1))
    if year < 1990 or year > 2100:
        return None
    return year


def normalize_fuel(raw: Any) -> str:
    code = _clean(raw).upper()
    return FUEL_MAP.get(code, FuelType.UNKNOWN)


def normalize_transmission(raw: Any) -> str:
    code = _clean(raw).upper()
    return TRANSMISSION_MAP.get(code, Transmission.UNKNOWN)


def normalize_color(raw: Any) -> Tuple[Optional[str], str]:
    """Return (canonical_colour | None, confidence). REP -> None (placeholder)."""
    code = _clean(raw).upper()
    if not code or code in NON_COLOR_CODES:
        return None, ColorConfidence.HIGH
    canon = COLOR_MAP.get(code)
    conf = ColorConfidence.LOW if code in AMBIGUOUS_COLOR_CODES else ColorConfidence.HIGH
    return canon, conf


def parse_km(raw: Any) -> Optional[int]:
    """Plain km if a plausible integer; else None (blank != 0; risk R12)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        v = int(raw)
        return v if v >= 100 else None
    s = _clean(raw)
    if not s or not s.isdigit():
        return None              # notes like SCRAP / SN 1.75 / T -> unknown
    v = int(s)
    return v if v >= 100 else None


def parse_rate(raw: Any) -> Tuple[Optional[int], Optional[float], bool]:
    """
    Return (price_inr, price_lakh, price_quotable).
    Status codes {4,33,44,DDD,DDDD} and values < 10000 are NOT money (risk R01).
    """
    if raw is None:
        return None, None, False
    s = _clean(raw).upper()
    if not s or s in RATE_STATUS_CODES:
        return None, None, False
    # strip any stray non-digits (e.g. 'D 365000')
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None, None, False
    value = int(digits)
    if value < PRICE_FLOOR:
        return None, None, False
    return value, round(value / 100000, 2), True


def parse_ownership(raw: Any) -> Optional[int]:
    s = _clean(raw)
    if s.isdigit():
        v = int(s)
        return v if 1 <= v <= 10 else None
    return None


def classify_location(raw: Any) -> Tuple[Optional[str], str, bool]:
    """Return (location_code, location_type, customer_viewable)."""
    code = _clean(raw).upper()
    if not code:
        return None, LocationType.UNKNOWN, True
    if SLOT_RE.match(code):
        return code, LocationType.SLOT, True
    if code in CUSTODY_CODES:
        # off main lot (POLI/ILL/NGP/REP especially) -> not immediately viewable
        viewable = code not in {"POLI", "ILL", "NGP", "REP"}
        return code, LocationType.CUSTODY, viewable
    return code, LocationType.UNKNOWN, True


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7D — extended column map (located by header text, not fixed index)
# ─────────────────────────────────────────────────────────────────────────────
_EXT_HEADER_MAP: Dict[str, str] = {
    "Price Range Low":           "price_range_low",
    "Price Range High":          "price_range_high",
    "Negotiable":                "negotiable",
    "Claimed Mileage":           "claimed_mileage_kmpl",
    "Insurance Type":            "insurance_type",
    "Insurance Expiry":          "insurance_expiry",
    "Zero Dep":                  "zero_dep",
    "Insurance Claim History":   "insurance_claim_history",
    "Accident Free":             "accident_free",
    "Flood Damage":              "flood_damage",
    "Repainted":                 "repainted",
    "Repaint Panels":            "repaint_panels",
    "Body Condition":            "body_condition",
    "Engine Condition":          "engine_condition",
    "Interior Condition":        "interior_condition",
    "Tyre Condition":            "tyre_condition",
    "Brake Condition":           "brake_condition",
    "Clutch Condition":          "clutch_condition",
    "Battery Condition":         "battery_condition",
    "Service History Available": "service_history_available",
    "Last Service Date":         "last_service_date",
    "Service Center Type":       "service_center_type",
    "RC Status":                 "rc_status",
    "Hypothecation Bank":        "hypothecation_bank",
    "Loan Closed":               "loan_closed",
    "NOC Available":             "noc_available",
    "Finance Eligible":          "finance_eligible",
    "Warranty Available":        "warranty_available",
    "Warranty Expiry":           "warranty_expiry",
    "Warranty Provider":         "warranty_provider",
    "Reason for Sale":           "reason_for_sale",
    "Best Features":             "best_features",
    "Known Issues":              "known_issues",
    "Photo Count":               "photo_count",
    "Video Count":               "video_count",
}

# ── Phase 12B: new grouped fields, read from Excel by HEADER (optional columns,
#    so this stays backward compatible — absent column => field left None).
#    kind: int | float | bool | text.  Excel value (if present) OVERRIDES the
#    model_specs auto-fill; if the column/value is absent, model_specs fills the
#    STANDARD specs and dealership-only fields stay None ("Data not available").
_NEW_EXT_FIELDS = [
    # standard specs (overridable)
    ("engine_cc", "Engine CC", "int"), ("power_bhp", "Power BHP", "float"),
    ("torque_nm", "Torque NM", "float"), ("aspiration", "Aspiration", "text"),
    ("transmission_subtype", "Transmission Subtype", "text"),
    ("gears", "Gears", "int"), ("drivetrain", "Drivetrain", "text"),
    ("mileage_arai_kmpl", "Mileage ARAI", "float"),
    ("fuel_tank_l", "Fuel Tank", "int"),
    ("length_mm", "Length MM", "int"), ("width_mm", "Width MM", "int"),
    ("height_mm", "Height MM", "int"), ("wheelbase_mm", "Wheelbase MM", "int"),
    ("boot_litres", "Boot Litres", "int"),
    ("ground_clearance_mm", "Ground Clearance MM", "int"),
    ("headlamp_type", "Headlamp Type", "text"), ("drl", "DRL", "bool"),
    ("fog_lamps", "Fog Lamps", "bool"), ("wheel_type", "Wheel Type", "text"),
    ("wheel_size_inch", "Wheel Size Inch", "int"),
    ("sunroof_type", "Sunroof Type", "text"), ("roof_rails", "Roof Rails", "bool"),
    ("spoiler", "Spoiler", "bool"), ("upholstery", "Upholstery", "text"),
    ("ac_type", "AC Type", "text"), ("rear_ac_vents", "Rear AC Vents", "bool"),
    ("power_windows", "Power Windows", "text"),
    ("adjustable_seat", "Adjustable Seat", "text"),
    ("push_button_start", "Push Button Start", "bool"),
    ("keyless_entry", "Keyless Entry", "bool"),
    ("cruise_control", "Cruise Control", "bool"),
    ("auto_folding_orvm", "Auto Folding ORVM", "bool"),
    ("rear_defogger", "Rear Defogger", "bool"), ("rear_wiper", "Rear Wiper", "bool"),
    ("wireless_charging", "Wireless Charging", "bool"),
    ("ventilated_seats", "Ventilated Seats", "bool"),
    ("connected_car", "Connected Car", "bool"),
    ("touchscreen_inches", "Touchscreen Inches", "float"),
    ("android_auto_carplay", "Android Auto CarPlay", "bool"),
    ("speakers", "Speakers", "int"), ("camera_type", "Camera Type", "text"),
    ("steering_controls", "Steering Controls", "bool"),
    ("airbags", "Airbags", "int"), ("abs_ebd", "ABS EBD", "bool"),
    ("esp", "ESP", "bool"), ("hill_hold", "Hill Hold", "bool"),
    ("parking_sensors", "Parking Sensors", "text"), ("isofix", "ISOFIX", "bool"),
    ("ncap_rating", "NCAP Rating", "int"),
    # dealership-entered extras
    ("puc_valid_till", "PUC Valid Till", "text"),
    ("road_tax_status", "Road Tax Status", "text"),
    ("fitness_valid_till", "Fitness Valid Till", "text"),
    ("duplicate_rc", "Duplicate RC", "bool"), ("keys_count", "Keys Count", "int"),
    ("usage_type", "Usage Type", "text"),
    ("tyre_life_pct", "Tyre Life Pct", "int"),
    ("battery_replaced_on", "Battery Replaced On", "text"),
    ("spare_key", "Spare Key", "bool"), ("spare_tyre", "Spare Tyre", "bool"),
    ("toolkit", "Toolkit", "bool"), ("floor_mats", "Floor Mats", "bool"),
    ("accessories_added", "Accessories Added", "text"),
    ("battery_health_pct", "Battery Health Pct", "int"),
    ("real_range_km", "Real Range KM", "int"),
    ("battery_warranty_till", "Battery Warranty Till", "text"),
    ("charger_type", "Charger Type", "text"),
    ("charging_time", "Charging Time", "text"),
    ("battery_owned", "Battery Owned", "bool"),
]
# register the new headers so _load_ext_col_map() locates them by header text
for _f, _h, _k in _NEW_EXT_FIELDS:
    _EXT_HEADER_MAP[_h] = _f

# Populated once per load_inventory() call.
_EXT_COL: Dict[str, int] = {}


def _load_ext_col_map(path: str, sheet: str) -> Dict[str, int]:
    """Read header row 2 and return {field_key: 0-based-col-index} for Phase 7D columns."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet]
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    wb.close()
    result: Dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val and str(val).strip() in _EXT_HEADER_MAP:
            result[_EXT_HEADER_MAP[str(val).strip()]] = idx
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7D — cell-level parsers for new column types
# ─────────────────────────────────────────────────────────────────────────────
_UNKNOWN_SENTINELS = {"unknown", "data not available", "n/a", "na", ""}


def _is_unknown(s: str) -> bool:
    return s.lower().strip() in _UNKNOWN_SENTINELS


def parse_yn(raw: Any) -> Optional[bool]:
    """Y -> True, N -> False, Unknown/blank/sentinel -> None."""
    s = _clean(raw).upper()
    if s == "Y":
        return True
    if s == "N":
        return False
    return None


def parse_enum_field(raw: Any, valid: set) -> Optional[str]:
    """Return canonical value if raw matches any member of valid (case-insensitive), else None."""
    s = _clean(raw)
    if _is_unknown(s):
        return None
    sl = s.lower()
    for v in valid:
        if sl == v.lower():
            return v
    return None


def parse_decimal(raw: Any) -> Optional[float]:
    """Parse positive decimal (e.g. claimed_mileage_kmpl). Unknown/zero -> None."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw) if float(raw) > 0 else None
    s = _clean(raw)
    if _is_unknown(s):
        return None
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def parse_text_field(raw: Any) -> Optional[str]:
    """Free-text field — None for unknown/sentinel values."""
    s = _clean(raw)
    if _is_unknown(s):
        return None
    return s or None


def parse_price_field(raw: Any) -> Optional[int]:
    """Integer price field (range low/high). Sentinel/zero -> None."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        v = int(raw)
        return v if v >= PRICE_FLOOR else None
    s = _clean(raw)
    if _is_unknown(s):
        return None
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    v = int(digits)
    return v if v >= PRICE_FLOOR else None


def parse_count_field(raw: Any) -> Optional[int]:
    """Non-negative integer count (photo/video). Unknown/sentinel -> None."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        v = int(raw)
        return v if v >= 0 else None
    s = _clean(raw)
    if _is_unknown(s):
        return None
    if s.isdigit():
        return int(s)
    return None


def normalize_insurance(raw: Any) -> Optional[str]:
    """Advisory hint / date field. Unknown/blank sentinels -> None."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.date().isoformat() if isinstance(raw, datetime) else raw.isoformat()
    s = _clean(raw)
    if _is_unknown(s):
        return None
    return s or None


def normalize_registration(raw: Any) -> Optional[str]:
    s = _clean(raw).upper().replace(" ", "")
    return s or None


# ─────────────────────────────────────────────────────────────────────────────
# ROW -> InventoryItem
# ─────────────────────────────────────────────────────────────────────────────
def _cell(row: Tuple, key: str) -> Any:
    idx = COL[key]
    return row[idx] if idx < len(row) else None


def _is_blank_row(row: Tuple) -> bool:
    return not any(_clean(c) for c in row)


def build_item(row: Tuple, *, source_sheet: str, as_of: str,
               ext_col: Optional[Dict[str, int]] = None) -> Optional[InventoryItem]:
    """Normalize one sheet row into an InventoryItem (None for blank rows)."""
    if _is_blank_row(row):
        return None

    ext = ext_col or {}

    def _xcell(key: str) -> Any:
        """Read from extended (Phase 7D) column by key; None if column absent."""
        idx = ext.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    reg = normalize_registration(_cell(row, "car_numb"))
    make_code, make_full = normalize_make(_cell(row, "make"))
    model = normalize_model(_cell(row, "model"))
    color_norm, color_conf = normalize_color(_cell(row, "color"))
    price_inr, price_lakh, quotable = parse_rate(_cell(row, "rate"))
    loc_code, loc_type, viewable = classify_location(_cell(row, "location"))

    # placeholder / non-car detection (risk R03)
    is_placeholder = (
        not model
        or make_code == "CUST"
        or _clean(_cell(row, "color")).upper() == "REP"
    )

    stock_raw = _cell(row, "stock_no")
    stock_no = int(stock_raw) if isinstance(stock_raw, (int, float)) and not isinstance(stock_raw, bool) else None

    # a registration is required to be a real, keyable listing
    if not reg:
        # keep as placeholder so sync can quarantine; synthesize a stable key
        reg = f"NOREG-{source_sheet}-{stock_no or _clean(_cell(row,'reg_last4')) or 'X'}"
        is_placeholder = True

    # ── Phase 7D: enum valid-value sets ──
    _COND4 = {"Excellent", "Good", "Fair", "Poor"}
    _COND3 = {"Good", "Fair", "Replace"}
    _CLUTCH = {"Good", "Fair", "Replace", "NA"}
    _SVC    = {"Authorised", "Multi-brand", "Local"}
    _RC     = {"Clear", "Hypothecated", "Pending"}
    _INS    = {"Comprehensive", "Third-Party", "Expired"}

    item = InventoryItem(
        registration_no=reg,
        stock_no=stock_no,
        reg_last4=_clean(_cell(row, "reg_last4")) or None,
        make=make_code,
        make_full=make_full,
        model=model,
        variant=_clean(_cell(row, "variant")) or None,
        year_int=parse_year(_cell(row, "year")),
        fuel_norm=normalize_fuel(_cell(row, "fuel")),
        transmission_norm=normalize_transmission(_cell(row, "transmission")),
        ownership_count=parse_ownership(_cell(row, "ownership")),
        km_driven=parse_km(_cell(row, "km")),
        color_norm=color_norm,
        color_confidence=color_conf,
        price_inr=price_inr,
        price_lakh=price_lakh,
        price_quotable=quotable,
        insurance_hint=normalize_insurance(_cell(row, "insurance")),
        body_type=_body_type_for(model),
        seats=_seats_for(model),
        location_code=loc_code,
        location_type=loc_type,
        customer_viewable=viewable,
        rto=_clean(_cell(row, "rto")) or None,
        listing_status=ListingStatus.AVAILABLE,
        is_ivr_eligible=bool(price_inr and price_inr > IVR_PRICE_THRESHOLD),
        is_placeholder=is_placeholder,
        source_sheet=source_sheet,
        raw={k: (_clean(_cell(row, k)) or None) for k in COL},
        as_of=as_of,
        # ── Phase 7D extended fields ──
        price_range_low=parse_price_field(_xcell("price_range_low")),
        price_range_high=parse_price_field(_xcell("price_range_high")),
        negotiable=parse_yn(_xcell("negotiable")),
        claimed_mileage_kmpl=parse_decimal(_xcell("claimed_mileage_kmpl")),
        insurance_type=parse_enum_field(_xcell("insurance_type"), _INS),
        insurance_expiry=normalize_insurance(_xcell("insurance_expiry")),
        zero_dep=parse_yn(_xcell("zero_dep")),
        insurance_claim_history=parse_yn(_xcell("insurance_claim_history")),
        accident_free=parse_yn(_xcell("accident_free")),
        flood_damage=parse_yn(_xcell("flood_damage")),
        repainted=parse_yn(_xcell("repainted")),
        repaint_panels=parse_text_field(_xcell("repaint_panels")),
        body_condition=parse_enum_field(_xcell("body_condition"), _COND4),
        engine_condition=parse_enum_field(_xcell("engine_condition"), _COND4),
        interior_condition=parse_enum_field(_xcell("interior_condition"), _COND4),
        tyre_condition=parse_enum_field(_xcell("tyre_condition"), _COND3),
        brake_condition=parse_enum_field(_xcell("brake_condition"), _COND3),
        clutch_condition=parse_enum_field(_xcell("clutch_condition"), _CLUTCH),
        battery_condition=parse_enum_field(_xcell("battery_condition"), _CLUTCH),
        service_history_available=parse_yn(_xcell("service_history_available")),
        last_service_date=normalize_insurance(_xcell("last_service_date")),
        service_center_type=parse_enum_field(_xcell("service_center_type"), _SVC),
        rc_status=parse_enum_field(_xcell("rc_status"), _RC),
        hypothecation_bank=parse_text_field(_xcell("hypothecation_bank")),
        loan_closed=parse_yn(_xcell("loan_closed")),
        noc_available=parse_yn(_xcell("noc_available")),
        finance_eligible=parse_yn(_xcell("finance_eligible")),
        warranty_available=parse_yn(_xcell("warranty_available")),
        warranty_expiry=normalize_insurance(_xcell("warranty_expiry")),
        warranty_provider=parse_text_field(_xcell("warranty_provider")),
        reason_for_sale=parse_text_field(_xcell("reason_for_sale")),
        best_features=parse_text_field(_xcell("best_features")),
        known_issues=parse_text_field(_xcell("known_issues")),
        photo_count=parse_count_field(_xcell("photo_count")),
        video_count=parse_count_field(_xcell("video_count")),
    )
    # ── Phase 12B: read new grouped fields from Excel (header-located, optional).
    #    A value present in the sheet is the owner's override and is set now, so it
    #    takes precedence over the model_specs auto-fill that runs next. ──
    _new_parsers = {"int": parse_count_field, "float": parse_decimal,
                    "bool": parse_yn, "text": parse_text_field}
    for _f, _h, _k in _NEW_EXT_FIELDS:
        _v = _new_parsers[_k](_xcell(_f))
        if _v is not None:
            setattr(item, _f, _v)
    # ── auto-fill STANDARD factory specs from the model_specs library where the
    #    dealership has not entered a value. Owner data always wins; unknown
    #    models are left untouched (no fabrication). Fully guarded. ──
    model_specs.apply_specs(item)
    return item


# ─────────────────────────────────────────────────────────────────────────────
# SHEET READERS
# ─────────────────────────────────────────────────────────────────────────────
def _read_rows(path: str, sheet: str) -> List[Tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {path}. Sheets: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
    wb.close()
    return rows


def load_sold_registrations(path: str, sheet: str = SOLD_SHEET) -> Set[str]:
    """
    Build the sold set from the DONT TOUCH SOLD sheet (risk R02).
    A registration is sold if it appears here OR carries RATE in {DDD, DDDD}.
    Rows with no registration are skipped (cannot be matched safely).
    """
    sold: Set[str] = set()
    for row in _read_rows(path, sheet):
        if _is_blank_row(row):
            continue
        reg = normalize_registration(_cell(row, "car_numb"))
        if reg:  # presence in the DONT TOUCH SOLD sheet == sold
            sold.add(reg)
    return sold


def reconcile_sold(items: List[InventoryItem], sold_regs: Set[str]) -> int:
    """
    Mark items as SOLD when their registration is in the sold set OR their own
    RATE was a DDD code. SOLD always wins. Returns count newly marked sold.
    """
    count = 0
    for it in items:
        rate_raw = (it.raw or {}).get("rate", "") or ""
        is_ddd = str(rate_raw).upper() in {"DDD", "DDDD"}
        if it.registration_no in sold_regs or is_ddd:
            if it.listing_status != ListingStatus.SOLD:
                it.listing_status = ListingStatus.SOLD
                it.customer_viewable = False
                count += 1
    return count


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def load_inventory(
    path: str,
    *,
    dnj_sheet: str = DNJ_SHEET,
    sold_sheet: str = SOLD_SHEET,
    as_of: Optional[str] = None,
    with_media: bool = True,
) -> List[InventoryItem]:
    """
    Full pipeline: read DNJ -> normalize/derive -> reconcile sold -> attach media
    -> return items.

    `with_media` (Phase 4A.5) reads the DNJ media URL columns (EXTERIOR/INTERIOR/
    VIDEO/INSTAGRAM/YOUTUBE) the media_sync layer wrote into the sheet and fills
    `InventoryItem.media`, so MediaService serves real Supabase URLs. It is a
    no-op on sheets without media columns (back-compatible with older fixtures).
    """
    as_of = as_of or utcnow_iso()
    ext_col = _load_ext_col_map(path, dnj_sheet)
    items: List[InventoryItem] = []
    for row in _read_rows(path, dnj_sheet):
        item = build_item(row, source_sheet=dnj_sheet, as_of=as_of, ext_col=ext_col)
        if item is not None:
            items.append(item)

    # de-duplicate by registration_no (last write wins, but keep a real over a
    # placeholder) — protects the natural key (risk R09/R11).
    by_reg: Dict[str, InventoryItem] = {}
    for it in items:
        existing = by_reg.get(it.registration_no)
        if existing is None or (existing.is_placeholder and not it.is_placeholder):
            by_reg[it.registration_no] = it
    deduped = list(by_reg.values())

    sold = load_sold_registrations(path, sold_sheet)
    reconcile_sold(deduped, sold)

    if with_media:
        # Wire Excel media columns -> InventoryItem.media. Imported lazily so the
        # loader keeps zero hard dependency on the media mapping module.
        from media_loader_mapping import attach_media
        attach_media(deduped, path, dnj_sheet)

    return deduped


def customer_facing(items: List[InventoryItem]) -> List[InventoryItem]:
    """The subset a customer may ever be offered (excludes sold/placeholder)."""
    return [it for it in items if it.is_customer_facing]


if __name__ == "__main__":
    import sys
    p = sys.argv[1] if len(sys.argv) > 1 else "../IVR_Sheet.xlsx"
    data = load_inventory(p)
    cf = customer_facing(data)
    print(f"Loaded {len(data)} items | customer-facing {len(cf)} | "
          f"sold {sum(1 for i in data if i.listing_status=='sold')} | "
          f"placeholder {sum(1 for i in data if i.is_placeholder)}")
