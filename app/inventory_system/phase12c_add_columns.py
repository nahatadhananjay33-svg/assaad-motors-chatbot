"""
phase12c_add_columns.py
=======================

Phase 12C migration: add the 68 Phase-12B grouped-field COLUMNS to the live DNJ
sheet so the Vehicle Details UI can SAVE them (persistence). Header-located, so
this is purely additive:

  * row 2 = short header (exact _NEW_EXT_FIELDS text — what the loader & edit
    backend locate by),
  * row 3 = friendly description,
  * data rows untouched (new cells start empty → "Data not available").

Idempotent (skips headers already present). Backs up the workbook before saving.
Existing columns A..LAST_UPDATED and all data are left exactly as-is.

Run:  python phase12c_add_columns.py
"""

from __future__ import annotations

import os, shutil, sys, io
from datetime import datetime, timezone

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import inventory_loader as L

XLSX = os.path.join(os.path.dirname(__file__), "..", "IVR_Sheet.xlsx")
HEADER_ROW, DESC_ROW = 2, 3


def _backup(path: str) -> str:
    bdir = os.path.join(os.path.dirname(path), "inventory_backups")
    os.makedirs(bdir, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(bdir, f"IVR_Sheet.phase12c.{stamp}.xlsx")
    shutil.copy2(path, dest)
    return dest


def migrate() -> None:
    backup = _backup(XLSX)
    print(f"Backup written: {backup}")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[L.DNJ_SHEET]

    existing = {str(ws.cell(HEADER_ROW, c).value).strip()
                for c in range(1, ws.max_column + 1) if ws.cell(HEADER_ROW, c).value}
    # first empty header column (append point)
    last = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(HEADER_ROW, c).value:
            last = c
    col = last + 1

    added = 0
    for field, header, kind in L._NEW_EXT_FIELDS:
        if header in existing:
            continue
        ws.cell(HEADER_ROW, col).value = header       # exact header for loader/edit
        ws.cell(DESC_ROW, col).value = header         # friendly label for the panel
        col += 1
        added += 1

    if added == 0:
        print("Nothing to add — all 68 columns already present. No change.")
        wb.close()
        return

    wb.save(XLSX)
    wb.close()
    print(f"Added {added} new columns (from col {last + 1}). Data rows untouched.")
    print("Backward compatible: loader/edit are header-located; existing columns "
          "and data unchanged.")


if __name__ == "__main__":
    migrate()
