"""
One-off Phase 5H.6 script: rebuild EXTERIOR/INTERIOR/VIDEO URL cells in
IVR_Sheet.xlsx/DNJ for the 5 audited vehicles from the actual contents of the
now-public `car-photos` Supabase bucket. Removes stale demo.supabase.co
duplicates and broken hxjxqdufquowvpmucqmf URLs (the bucket previously
returned 404 because it was private; now public -> 200).
"""
import openpyxl
from media_loader_mapping import detect_media_layout, HEADER_ROW, DATA_START_ROW, DNJ_SHEET, MEDIA_KEYWORDS

PATH = "../IVR_Sheet.xlsx"
BASE = "https://hxjxqdufquowvpmucqmf.supabase.co/storage/v1/object/public/car-photos"

# Actual bucket contents (from Storage API list, after making bucket public)
BUCKET = {
    "MH01BK9444": {
        "exterior_photo": [
            "exterior/89dd9e1ae28c673fc9f3f37af7ba4d0ed9d4cf86.jpeg",
            "exterior/4453a291cafc2f9e5c114b17bc7d2627ec39d429.jpeg",
            "exterior/77c92522d227ee6e7f03efff8fd055297066d127.jpeg",
        ],
        "interior_photo": [
            "interior/efa85323f3d676e0fa5156f5d53ccb84ac351c15.jpeg",
        ],
        "video": ["video/27aed25061df513c74c7a0cd18fcb137f7fa991c.mp4"],
    },
    "MH02EZ6001": {
        "exterior_photo": [
            "exterior/8408ad5abbe9d81ba457da92d0e4167f93101c9f.jpeg",
            "exterior/54745b3fd54159ae139689e46d9c947669c3b075.jpeg",
            "exterior/ec0fc53a016c02801bc2130bef931a295992b0d6.jpeg",
        ],
        "interior_photo": [
            "interior/198393a23a93e0cc984a7d27c1d7f433c22be492.jpeg",
        ],
        "video": ["video/249e5974551bfeaff7bca481eb66931ebbb34a0a.mp4"],
    },
    "MH05AF8000": {
        "exterior_photo": [
            "exterior/7b22d2d3782e13c9252dfeb557d6d094975bbf53.jpeg",
            "exterior/7328a19be6da40ed710a08d5d04f82c94b7a3383.jpeg",
        ],
        "interior_photo": [
            "interior/0e679e3dd222901acaa5e6f8f877452fbadfe84c.jpeg",
        ],
        "video": ["video/a5648ff62314031bd89b30f0ad98fa76db51da59.mp4"],
    },
    "MH05EQ8661": {
        "exterior_photo": [
            "exterior/fa1c299f8d9321b5df1b88ce8a1b63942df7ef03.jpeg",
            "exterior/8c5e608cf316deff0ff9d23d588e6b88eebfb84f.jpeg",
            "exterior/c732ed5e8f14d915e06b276044e376f5b1062d81.jpeg",
        ],
        "interior_photo": [
            "interior/589caacf43d27c889f253a3c95e9c77220bfef19.jpeg",
            "interior/4e14840d6500dae6437b21ff549708bd5654d9f6.jpeg",
            "interior/2a1e36a467f52bdd247ce35b0d710cf6bfe360ed.jpeg",
        ],
        "video": ["video/b64fbc83a8068a2049e479d3e0b19b7032483001.mp4"],
    },
    "MH50V0269": {
        "exterior_photo": [
            "exterior/60dba48b93ed69a9948db26ab6995909bb4b50b3.jpeg",
            "exterior/73ad3d1b93e3316c710f57fafd55342c343c61f8.jpeg",
        ],
        "interior_photo": [
            "interior/56f058ca5362085bd3da71346bdc4dce08a4d80d.jpeg",
            "interior/fc30ad9cde7a7131893d9137d159d6831919aeb6.jpeg",
        ],
        "video": ["video/414242b6af1bb291812538ee2885dd2f54c3b756.mp4"],
    },
}

REG_HEADERS = {"CAR NUMB", "CAR NUMBER", "CAR_NUMB", "REGISTRATION", "REG NO"}


def norm_reg(v):
    return ("" if v is None else str(v)).strip().upper().replace(" ", "")


wb = openpyxl.load_workbook(PATH)
ws = wb[DNJ_SHEET]

header = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
layout = detect_media_layout(header)

updated_cells = []

for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    reg = norm_reg(ws.cell(row_idx, layout.car_col).value)
    if reg not in BUCKET:
        continue
    for media_type, slots in layout.groups.items():
        urls = [f"{BASE}/{reg}/{p}" for p in BUCKET[reg].get(media_type, [])]
        for slot_number, col in slots:
            old = ws.cell(row_idx, col).value
            new = urls[slot_number - 1] if slot_number - 1 < len(urls) else None
            if old != new:
                ws.cell(row_idx, col).value = new
                updated_cells.append((reg, media_type, row_idx, col, old, new))

wb.save(PATH)

print(f"Updated {len(updated_cells)} cells")
for reg, mt, row, col, old, new in updated_cells:
    print(f"  {reg} {mt} R{row}C{col}: {old!r} -> {new!r}")
