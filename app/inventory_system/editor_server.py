"""
editor_server.py
================

A tiny, self-contained LOCAL Excel-like editor for the real inventory workbook
(`app/IVR_Sheet.xlsx`).  It is intentionally standalone: it uses only Python's
built-in ``http.server`` plus ``openpyxl`` (already a project dependency) and does
NOT touch the chatbot, auth, Supabase, media, or any other subsystem.

What it does
------------
* Serves a browser spreadsheet grid at   http://localhost:8010/editor
* Reads the ACTUAL workbook sheets (DNJ = editable, "DONT TOUCH SOLD" = view only)
* Lets the owner click / edit / clear / copy-paste / arrow-key around cells
* Name Box + value bar (jump to N95, AS95, FK200, ...)
* Freeze panes (rows 0-3, cols 0-4, any combination)
* SAVE writes the edited cells straight back into the same .xlsx (with a backup)

Run:
    python editor_server.py            # serves on 127.0.0.1:8010
    python editor_server.py 8020        # custom port

Nothing here is wired into the production server; it is a local tool only.
"""

from __future__ import annotations

import json
import os
import base64
import shutil
import sys
import time
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

import openpyxl

# ── paths ────────────────────────────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
# Workbook + backup locations are env-overridable so the SAME script works both
# locally (defaults below → the repo's app/IVR_Sheet.xlsx) and inside the KVM2
# container, where EDITOR_XLSX=/data/IVR_Sheet.xlsx points at the exact same
# writable workbook the chatbot loads (CHAT_XLSX). No behaviour change locally.
WORKBOOK_PATH = os.environ.get("EDITOR_XLSX") or os.path.abspath(os.path.join(HERE, "..", "IVR_Sheet.xlsx"))
BACKUP_DIR = os.environ.get("EDITOR_BACKUP_DIR") or os.path.abspath(os.path.join(HERE, "..", "inventory_backups"))
# Audit log of every saved cell change (one JSON object per line). Kept in the
# app's data dir so the Developer Dashboard (which reads config.DATA_DIR) can show
# it. Default matches config.DATA_DIR locally (../inventory_system/data); the KVM2
# container sets EDITOR_AUDIT_LOG=/data/excel_edit_audit.jsonl (the shared volume).
AUDIT_LOG = os.environ.get("EDITOR_AUDIT_LOG") or os.path.join(HERE, "data", "excel_edit_audit.jsonl")

# Which sheets the editor exposes, and whether each is editable.
# The real workbook uses "DONT TOUCH SOLD" for the sold list; we surface it under
# the friendly name SOLD_CARS but keep it read-only.
SHEET_CONFIG = [
    {"key": "DNJ", "sheet": "DNJ", "label": "DNJ", "editable": True},
    {"key": "SOLD_CARS", "sheet": "DONT TOUCH SOLD", "label": "SOLD_CARS", "editable": False},
]
SHEET_BY_KEY = {c["key"]: c for c in SHEET_CONFIG}

# UI grid target — number of rows the browser shows so the owner can keep adding
# vehicles into empty rows. We do NOT physically pad the .xlsx with blank rows.
UI_MIN_ROWS = 1000
UI_MIN_COLS = 168   # DNJ full width; ensures every real column is visible


# ── helpers ──────────────────────────────────────────────────────────────────
def _cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _coerce(text):
    """Best-effort: keep numbers numeric so PRICE/KM/YEAR stay real numbers."""
    if text is None:
        return None
    s = str(text).strip()
    if s == "":
        return None
    # plain integer
    try:
        if s.lstrip("-").isdigit():
            return int(s)
    except Exception:
        pass
    # float (but not things like phone/version with multiple dots)
    try:
        f = float(s)
        return f
    except Exception:
        return str(text)


def _col_letter(n):
    """1-based column number -> Excel letters (1->A, 27->AA)."""
    s = ""
    while n > 0:
        n, m = divmod(n - 1, 26)
        s = chr(65 + m) + s
    return s


def _utcnow_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _append_audit(entries):
    """Append audit rows (one JSON object per line). Best-effort: a logging
    failure must NEVER break an already-successful save."""
    if not entries:
        return
    try:
        os.makedirs(os.path.dirname(AUDIT_LOG), exist_ok=True)
        with open(AUDIT_LOG, "a", encoding="utf-8") as f:
            for e in entries:
                f.write(json.dumps(e, ensure_ascii=False) + "\n")
    except Exception:
        pass


def load_sheet_data(sheet_name):
    # NOTE: read_only random .cell() access is pathologically slow in openpyxl;
    # a normal load + iter_rows over this small workbook is fast and correct.
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(sheet_name)
    ws = wb[sheet_name]
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    grid_rows = max(max_row, UI_MIN_ROWS)
    grid_cols = max(max_col, UI_MIN_COLS)
    # Build a dense grid of strings for the used range only; the UI extends the
    # rest as blank rows/cols client-side (cheaper payload).
    rows = []
    for row_cells in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        rows.append([_cell_to_str(cell.value) for cell in row_cells])
    wb.close()
    return {
        "usedRows": max_row,
        "usedCols": max_col,
        "gridRows": grid_rows,
        "gridCols": grid_cols,
        "rows": rows,
    }


def save_edits(sheet_name, edits, user="editor"):
    """edits: list of {r, c, v} with 1-based r/c. Backs up then writes in place.
    `user` is recorded in the audit log for each changed cell."""
    if not os.path.exists(WORKBOOK_PATH):
        raise FileNotFoundError(WORKBOOK_PATH)
    # backup first
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    backup = os.path.join(BACKUP_DIR, f"IVR_Sheet.editor_{stamp}.xlsx")
    shutil.copy2(WORKBOOK_PATH, backup)

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(sheet_name)
    ws = wb[sheet_name]
    applied = 0
    audit_rows = []
    ts = _utcnow_iso()
    for e in edits:
        r = int(e["r"])
        c = int(e["c"])
        if r < 1 or c < 1:
            continue
        cell = ws.cell(row=r, column=c)
        old_val = _cell_to_str(cell.value)          # capture BEFORE overwrite
        cell.value = _coerce(e.get("v"))
        new_val = _cell_to_str(cell.value)
        applied += 1
        audit_rows.append({
            "ts": ts, "user": user, "sheet": sheet_name,
            "cell": f"{_col_letter(c)}{r}", "old": old_val, "new": new_val,
        })

    # atomic-ish write: temp then replace
    tmp = WORKBOOK_PATH + ".editortmp"
    wb.save(tmp)
    wb.close()
    os.replace(tmp, WORKBOOK_PATH)
    # Log ONLY after the save actually succeeded (best-effort; never fatal).
    _append_audit(audit_rows)
    return {"applied": applied, "backup": os.path.basename(backup)}


# ── static files ─────────────────────────────────────────────────────────────
def _read(name):
    with open(os.path.join(HERE, name), "rb") as f:
        return f.read()


STATIC = {
    "/editor": ("editor.html", "text/html; charset=utf-8"),
    "/editor.html": ("editor.html", "text/html; charset=utf-8"),
    "/editor.js": ("editor.js", "application/javascript; charset=utf-8"),
    "/editor.css": ("editor.css", "text/css; charset=utf-8"),
}


class Handler(BaseHTTPRequestHandler):
    server_version = "InventoryEditor/1.0"

    def log_message(self, *a):  # quiet
        return

    def _auth_user(self):
        """Username from the HTTP Basic Auth header (nginx gate) for the audit
        log; falls back to 'editor' when there is no auth (local dev)."""
        h = self.headers.get("Authorization", "")
        if h.startswith("Basic "):
            try:
                dec = base64.b64decode(h[6:]).decode("utf-8", "replace")
                return dec.split(":", 1)[0] or "editor"
            except Exception:
                return "editor"
        return "editor"

    def _send(self, status, body, ctype="application/json; charset=utf-8"):
        if isinstance(body, (dict, list)):
            body = json.dumps(body, ensure_ascii=False, default=str).encode("utf-8")
        elif isinstance(body, str):
            body = body.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/" or path == "":
            self.send_response(302)
            self.send_header("Location", "/editor")
            self.end_headers()
            return
        if path in STATIC:
            fname, ctype = STATIC[path]
            try:
                return self._send(200, _read(fname), ctype)
            except FileNotFoundError:
                return self._send(404, {"error": "missing", "file": fname})
        if path == "/api/sheets":
            return self._send(200, {"sheets": [
                {"key": c["key"], "label": c["label"], "editable": c["editable"]}
                for c in SHEET_CONFIG
            ], "workbook": os.path.basename(WORKBOOK_PATH)})
        if path == "/api/data":
            qs = parse_qs(parsed.query)
            key = (qs.get("sheet") or ["DNJ"])[0]
            cfg = SHEET_BY_KEY.get(key)
            if not cfg:
                return self._send(404, {"error": "unknown_sheet", "sheet": key})
            try:
                data = load_sheet_data(cfg["sheet"])
            except KeyError:
                return self._send(404, {"error": "sheet_not_in_workbook", "sheet": cfg["sheet"]})
            data["key"] = key
            data["editable"] = cfg["editable"]
            return self._send(200, data)
        return self._send(404, {"error": "not_found", "path": path})

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path != "/api/save":
            return self._send(404, {"error": "not_found", "path": parsed.path})
        length = int(self.headers.get("Content-Length") or 0)
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw or b"{}")
        except Exception as e:
            return self._send(400, {"error": "bad_json", "detail": str(e)})
        key = payload.get("sheet") or "DNJ"
        cfg = SHEET_BY_KEY.get(key)
        if not cfg:
            return self._send(404, {"error": "unknown_sheet", "sheet": key})
        if not cfg["editable"]:
            return self._send(403, {"error": "read_only", "sheet": key})
        edits = payload.get("edits") or []
        if not isinstance(edits, list):
            return self._send(400, {"error": "edits_must_be_list"})
        try:
            result = save_edits(cfg["sheet"], edits, user=self._auth_user())
        except Exception as e:
            return self._send(500, {"error": "save_failed", "detail": str(e)})
        result["ok"] = True
        return self._send(200, result)


def main():
    # Bind host is env-overridable: default 127.0.0.1 (local-only, safe) but the
    # container sets EDITOR_HOST=0.0.0.0 so the sibling nginx container can reach
    # it on the private docker network (the port is never published to the host).
    host = os.environ.get("EDITOR_HOST", "127.0.0.1")
    port = int(os.environ.get("EDITOR_PORT", "8010"))
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            pass
    httpd = ThreadingHTTPServer((host, port), Handler)
    print(f"Inventory Excel editor running:  http://{host}:{port}/editor")
    print(f"Workbook: {WORKBOOK_PATH}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nstopped")


if __name__ == "__main__":
    main()
