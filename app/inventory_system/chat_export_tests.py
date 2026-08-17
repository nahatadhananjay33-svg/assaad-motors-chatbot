"""
chat_export_tests.py
====================

Tests for the simplified customer-conversation export used by both the
Developer Dashboard and the Owner Dashboard.

Covers the explicit requirements:
  * the export contains the conversation (customer + agent, long format);
  * the export contains ONLY timestamp / conversation_id / speaker / message —
    NO intent, filters, result_count, latency, retrieval or parser metadata;
  * CSV and XLSX both work;
  * date filtering (today / 7d / 30d / all) works;
  * a customer / unauthenticated caller cannot reach the download endpoints,
    Developers can download from /developer, and the Owner endpoint is
    Owner-only (staff are 403).
"""

from __future__ import annotations

import base64
import csv
import io
import os
import shutil
import sqlite3
import tempfile
import unittest
from datetime import datetime, timedelta, timezone

import chat_export
from pilot_query_log import PilotQueryLog, QueryLogEntry
from chat_api import route

# reuse the fully-wired auth/service harness (temp audit/users/pilot stores)
from developer_dashboard_tests import DevDashboardBase, _b


def _iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S+00:00")


class TestExportContent(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix="cexp_")
        self.db = os.path.join(self.tmp, "pilot_query_log.db")
        log = PilotQueryLog(self.db)
        now = datetime.now(timezone.utc)
        # a recent, fully-metadata'd turn — metadata must NOT reach the export
        log.record(QueryLogEntry(
            timestamp=_iso(now), conversation_id="c1", session_id="c1",
            user_query="Swift automatic hai?", bot_response="Haan, Swift automatic available hai.",
            detected_language="hinglish", detected_intent="availability",
            route="inventory", matched_inventory=True, response_time_ms=42.0,
            vehicle_selected="2019 Swift"))
        log.record(QueryLogEntry(
            timestamp=_iso(now + timedelta(seconds=30)), conversation_id="c1",
            session_id="c1", user_query="Sabse sasti wali?",
            bot_response="Sabse sasti Swift ye rahi."))
        # an OLD turn (40 days back) for the date-range test
        log.record(QueryLogEntry(
            timestamp=_iso(now - timedelta(days=40)), conversation_id="c_old",
            session_id="c_old", user_query="purani query", bot_response="purana jawab"))
        log.close()

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _csv_rows(self, range_key="all"):
        st, pl = chat_export.export_payload(self.db, fmt="csv", range_key=range_key)
        self.assertEqual(st, 200, pl)
        data = base64.b64decode(pl["content_b64"]).decode("utf-8-sig")
        return pl, list(csv.DictReader(io.StringIO(data)))

    def test_csv_columns_are_minimal(self):
        pl, rows = self._csv_rows()
        self.assertEqual(list(rows[0].keys()),
                         ["timestamp", "conversation_id", "speaker", "message"])

    def test_csv_has_customer_and_agent_messages(self):
        pl, rows = self._csv_rows()
        pairs = [(r["speaker"], r["message"]) for r in rows]
        self.assertIn(("customer", "Swift automatic hai?"), pairs)
        self.assertIn(("agent", "Haan, Swift automatic available hai."), pairs)
        self.assertIn(("customer", "Sabse sasti wali?"), pairs)

    def test_export_leaks_no_metadata(self):
        _, pl = chat_export.export_payload(self.db, fmt="csv", range_key="all"), None
        st, payload = chat_export.export_payload(self.db, fmt="csv", range_key="all")
        blob = base64.b64decode(payload["content_b64"]).decode("utf-8-sig").lower()
        for needle in ("hinglish", "availability", "inventory", "2019 swift",
                       "42.0", "intent", "filters", "result_count", "latency",
                       "response_time", "vehicle_selected", "route"):
            self.assertNotIn(needle, blob, f"metadata leaked into export: {needle!r}")

    def test_date_range_filters(self):
        # 'all' includes the 40-day-old row; '30d' excludes it
        _, all_rows = self._csv_rows("all")
        _, recent_rows = self._csv_rows("30d")
        cids_all = {r["conversation_id"] for r in all_rows}
        cids_recent = {r["conversation_id"] for r in recent_rows}
        self.assertIn("c_old", cids_all)
        self.assertNotIn("c_old", cids_recent)
        self.assertIn("c1", cids_recent)

    def test_xlsx_export(self):
        from openpyxl import load_workbook
        st, pl = chat_export.export_payload(self.db, fmt="xlsx", range_key="all")
        self.assertEqual(st, 200)
        wb = load_workbook(io.BytesIO(base64.b64decode(pl["content_b64"])))
        ws = wb.active
        header = [c.value for c in ws[1]]
        self.assertEqual(header, ["timestamp", "conversation_id", "speaker", "message"])
        # header + at least the recent conversation's 4 message rows + old pair
        self.assertGreaterEqual(ws.max_row, 5)

    def test_bad_format_rejected(self):
        st, pl = chat_export.export_payload(self.db, fmt="pdf", range_key="all")
        self.assertEqual(st, 400)


class TestExportAuth(DevDashboardBase):
    def _seed(self):
        self.svc.pilot_log.record(QueryLogEntry(
            timestamp=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S+00:00"),
            conversation_id="s1", session_id="s1",
            user_query="Swift hai?", bot_response="Haan available hai."))

    def _login(self, username, password):
        st, pl = route("POST", "/auth/login", _b({"username": username,
                                                  "password": password}), self.svc)
        self.assertEqual(st, 200, pl)
        return pl["token"]

    # ── Developer download endpoint ──
    def test_developer_can_download(self):
        self._seed()
        tok = self._login("devx", "devx-pass-1")
        st, pl = route("GET", "/developer/chats/export?format=csv&range=all", b"",
                       self.svc, session_token=tok)
        self.assertEqual(st, 200)
        self.assertEqual(pl["status"], "ok")
        self.assertIn("content_b64", pl)

    def test_unauthenticated_cannot_download_developer(self):
        self._seed()
        st, pl = route("GET", "/developer/chats/export?format=csv", b"", self.svc)
        self.assertEqual(st, 401)

    def test_staff_cannot_download_developer(self):
        import user_management
        user_management.handle_create(_b({"full_name": "Sam", "username": "staff9",
                                          "password": "pass123", "role": "Inventory Staff"}))
        tok = self._login("staff9", "pass123")
        st, pl = route("GET", "/developer/chats/export?format=csv", b"", self.svc,
                       session_token=tok)
        self.assertEqual(st, 403)

    # ── Owner download endpoint ──
    def test_owner_can_download(self):
        self._seed()
        tok = self._login("owner", "owner123")
        st, pl = route("GET", "/admin/owner/chat_logs/export?format=csv&range=all", b"",
                       self.svc, acting_user="owner", session_token=tok)
        self.assertEqual(st, 200)
        self.assertEqual(pl["status"], "ok")
        st, lst = route("GET", "/admin/owner/chat_logs", b"", self.svc,
                        acting_user="owner", session_token=tok)
        self.assertEqual(st, 200)
        self.assertTrue(lst["available"])

    def test_staff_cannot_download_owner(self):
        import user_management
        user_management.handle_create(_b({"full_name": "Sam", "username": "staff8",
                                          "password": "pass123", "role": "Inventory Staff"}))
        tok = self._login("staff8", "pass123")
        st, pl = route("GET", "/admin/owner/chat_logs/export?format=csv", b"", self.svc,
                       acting_user="staff8", session_token=tok)
        self.assertEqual(st, 403)


if __name__ == "__main__":
    unittest.main(verbosity=2)
