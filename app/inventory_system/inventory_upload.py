"""
inventory_upload.py — Phase 8G owner-friendly inventory upload
==============================================================

Implements the upload workflow designed in Phase 8F. It is ADDITIVE and wraps
the EXISTING, unchanged `ChatService.refresh_inventory()` (atomic, zero-downtime
engine swap). It does NOT touch retrieval, pricing, Marathi, consultative, media,
logging, or security logic — and it does not re-implement inventory parsing
(validation is a read-only dry-run through the existing `inventory_loader`).

Flow per upload:  parse multipart -> validate -> backup current -> atomic
replace -> refresh_inventory() -> (on failure) automatic rollback.

Used by two new admin endpoints (registered in chat_api.route, protected by the
existing fail-closed admin gate):
    POST /admin/upload_inventory   (multipart/form-data, field "file")
    GET  /admin/inventory_status
"""
from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from inventory_loader import load_inventory, customer_facing, DNJ_SHEET
from inventory_models import utcnow_iso

# Max upload size (default 10 MB; override with CHAT_MAX_UPLOAD_MB).
def _max_bytes() -> int:
    try:
        mb = int(os.getenv("CHAT_MAX_UPLOAD_MB", "10"))
    except (TypeError, ValueError):
        mb = 10
    return max(1, mb) * 1024 * 1024

# CAR NUMB is column N → 0-based index 13 (see final_excel_columns.md).
_CAR_NUMB_IDX = 13
_HEADER_TOKENS = {"CAR NUMB", "CAR NUMBER", "CARNUMB", "CAR NO", "REG NO"}


# ── multipart/form-data: extract the single "file" field ──────────────────────
def parse_multipart_file(body: bytes, content_type: str) -> Tuple[str, bytes]:
    """Return (filename, file_bytes) for the 'file' field. Raises ValueError."""
    m = re.search(r'boundary=(?:"([^"]+)"|([^;]+))', content_type or "", re.I)
    if not m:
        raise ValueError("missing multipart boundary")
    boundary = (m.group(1) or m.group(2)).strip()
    delim = b"--" + boundary.encode()
    for part in body.split(delim):
        if b"Content-Disposition" not in part:
            continue
        idx = part.find(b"\r\n\r\n")
        if idx == -1:
            continue
        headers = part[:idx].decode("utf-8", "replace")
        if 'name="file"' not in headers:
            continue
        content = part[idx + 4:]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        fn = re.search(r'filename="([^"]*)"', headers)
        return (fn.group(1) if fn else ""), content
    raise ValueError("no 'file' field in upload")


# ── validation (read-only dry-run; rules from inventory_validation_rules.md) ───
def validate_workbook(path: str) -> Dict[str, Any]:
    """Return {errors, warnings, vehicles_loaded, rows_processed}. Never raises."""
    errors: List[str] = []
    warnings: List[str] = []
    rows = 0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:                       # corrupt / not an xlsx
        return {"errors": [f"File is not a readable .xlsx workbook ({e})."],
                "warnings": [], "vehicles_loaded": 0, "rows_processed": 0}

    try:
        if DNJ_SHEET not in wb.sheetnames:
            return {"errors": ["Sheet 'DNJ' not found — do not rename the sheets."],
                    "warnings": [], "vehicles_loaded": 0, "rows_processed": 0}
        ws = wb[DNJ_SHEET]
        if (ws.max_column or 0) < _CAR_NUMB_IDX + 1:
            errors.append("Required column 'CAR NUMB' (N) is missing — keep the "
                          "column layout unchanged.")
        # scan registrations for presence + duplicates
        seen: Dict[str, int] = {}
        regs = 0
        for r in ws.iter_rows(values_only=True):
            rows += 1
            if not r or len(r) <= _CAR_NUMB_IDX:
                continue
            val = r[_CAR_NUMB_IDX]
            if val is None:
                continue
            s = str(val).strip().upper()
            if not s or s in _HEADER_TOKENS:
                continue
            regs += 1
            seen[s] = seen.get(s, 0) + 1
    finally:
        wb.close()

    # Duplicates are a WARNING, not a blocker: the production loader dedups by
    # registration deterministically (real-over-placeholder, last-write-wins),
    # and the live sheet itself legitimately contains a benign duplicate. Blocking
    # would reject the owner's normal file. We surface them so they can be fixed.
    dups = sorted(k for k, v in seen.items() if v > 1)
    if dups:
        shown = ", ".join(dups[:10]) + (" …" if len(dups) > 10 else "")
        warnings.append(f"Duplicate registration(s) in CAR NUMB: {shown} "
                        "(kept one per registration).")
    if regs == 0 and not errors:
        errors.append("No CAR NUMB values found — column N (registration) is empty.")

    # real facing-vehicle count via the existing loader (read-only)
    count = 0
    if not errors:
        try:
            count = len(customer_facing(load_inventory(path)))
        except Exception as e:
            errors.append(f"Could not load inventory from this file ({e}).")
    if not errors and count == 0:
        errors.append("This file produces 0 live cars — refusing to replace inventory.")

    return {"errors": errors, "warnings": warnings,
            "vehicles_loaded": count, "rows_processed": rows}


# ── upload orchestration ──────────────────────────────────────────────────────
def _backups_dir(xlsx: str) -> str:
    d = os.path.join(os.path.dirname(os.path.abspath(xlsx)), "inventory_backups")
    os.makedirs(d, exist_ok=True)
    return d


def handle_upload(service: Any, body: bytes, content_type: str) -> Tuple[int, Dict[str, Any]]:
    if not body:
        return 400, {"status": "rejected", "errors": ["Empty request body."],
                     "live_inventory_changed": False}
    if len(body) > _max_bytes():
        return 400, {"status": "rejected",
                     "errors": [f"Upload too large (max {_max_bytes() // (1024*1024)} MB)."],
                     "live_inventory_changed": False}
    try:
        filename, content = parse_multipart_file(body, content_type)
    except ValueError as e:
        return 400, {"status": "rejected",
                     "errors": [f"Upload must be multipart/form-data with a 'file' field ({e})."],
                     "live_inventory_changed": False}
    if not filename.lower().endswith(".xlsx"):
        return 400, {"status": "rejected",
                     "errors": ["File must be an Excel .xlsx."],
                     "live_inventory_changed": False}
    if len(content) > _max_bytes():
        return 400, {"status": "rejected",
                     "errors": [f"File too large (max {_max_bytes() // (1024*1024)} MB)."],
                     "live_inventory_changed": False}

    xlsx = service.xlsx_path
    # Stage with a real .xlsx extension — openpyxl validates by extension and
    # rejects unknown suffixes. Same directory as the live file so os.replace is atomic.
    incoming = xlsx + ".incoming.xlsx"
    try:
        with open(incoming, "wb") as f:
            f.write(content)
    except OSError as e:
        return 500, {"status": "error", "detail": f"Could not stage upload ({e}).",
                     "live_inventory_changed": False}

    # 1) validate the candidate (live file still untouched)
    v = validate_workbook(incoming)
    if v["errors"]:
        _silent_remove(incoming)
        return 400, {"status": "rejected", "errors": v["errors"],
                     "warnings": v["warnings"], "live_inventory_changed": False}

    # 2) backup the current live Excel
    backup_path: Optional[str] = None
    if os.path.exists(xlsx):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(_backups_dir(xlsx), f"IVR_Sheet_{ts}.xlsx")
        try:
            shutil.copy2(xlsx, backup_path)
        except OSError as e:
            _silent_remove(incoming)
            return 500, {"status": "error", "detail": f"Could not back up current inventory ({e}).",
                         "live_inventory_changed": False}
        _prune_backups(_backups_dir(xlsx), keep=20)

    prev_count = getattr(service, "inventory_count", None)

    # 3) atomic replace
    try:
        os.replace(incoming, xlsx)
    except OSError as e:
        _silent_remove(incoming)
        return 500, {"status": "error", "detail": f"Could not replace inventory file ({e}).",
                     "live_inventory_changed": False}

    # 4) refresh (existing atomic engine swap) — with automatic rollback
    try:
        report = service.refresh_inventory()
        new_count = report.get("inventory_count")
        if not new_count or new_count <= 0:
            raise RuntimeError("refresh produced 0 vehicles")
    except Exception as e:
        rolled = _rollback(service, xlsx, backup_path)
        return 500, {"status": "error",
                     "detail": f"Refresh failed ({e}); "
                               + ("restored previous inventory." if rolled
                                  else "ROLLBACK ALSO FAILED — inventory may be inconsistent."),
                     "live_inventory_changed": (not rolled),
                     "restored_from": os.path.basename(backup_path) if backup_path else None}

    return 200, {"status": "ok", "message": "Inventory updated",
                 "rows_processed": v["rows_processed"], "vehicles_loaded": new_count,
                 "previous_count": prev_count,
                 "backup": os.path.basename(backup_path) if backup_path else None,
                 "sync": report.get("sync"), "warnings": v["warnings"],
                 "updated_at": utcnow_iso()}


def _rollback(service: Any, xlsx: str, backup_path: Optional[str]) -> bool:
    """Restore the previous Excel and re-refresh. Returns True on success."""
    if not backup_path or not os.path.exists(backup_path):
        return False
    try:
        shutil.copy2(backup_path, xlsx)
        service.refresh_inventory()
        return True
    except Exception:
        return False


def handle_status(service: Any) -> Tuple[int, Dict[str, Any]]:
    xlsx = getattr(service, "xlsx_path", None)
    last_updated = None
    current_file = None
    if xlsx and os.path.exists(xlsx):
        last_updated = datetime.fromtimestamp(os.path.getmtime(xlsx)).isoformat(timespec="seconds")
        current_file = os.path.basename(xlsx)
    return 200, {"inventory_count": getattr(service, "inventory_count", None),
                 "last_updated": last_updated, "current_file": current_file}


# ── helpers ───────────────────────────────────────────────────────────────────
def _silent_remove(path: str) -> None:
    try:
        os.remove(path)
    except OSError:
        pass


def _prune_backups(d: str, keep: int = 20) -> None:
    try:
        files = sorted((f for f in os.listdir(d) if f.endswith(".xlsx")), reverse=True)
        for old in files[keep:]:
            _silent_remove(os.path.join(d, old))
    except OSError:
        pass
