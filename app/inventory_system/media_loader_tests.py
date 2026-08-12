"""
media_loader_tests.py
=====================

Phase 4A.5 — Final Wiring & Live Validation tests.

Proves the full deterministic chain on a workbook shaped exactly like the real
`IVR_Sheet.xlsx` → DNJ sheet (header on row 2, data from row 4, media URL columns
EXTERIOR/INTERIOR/VIDEO/INSTAGRAM/YOUTUBE):

    Upload → URL Generation → Excel Update → Inventory Loader → Media Service → Chat Response

Run:  python media_loader_tests.py      (stdlib unittest; no network, no Supabase)
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# reuse the production media store (Upload → URL generation) without any network
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "media_sync"))

from inventory_models import MediaType
from inventory_loader import load_inventory
from retrieval_engine import RetrievalEngine
from media_service import MediaService, InventoryMediaProvider
import media_loader_mapping as M


# ─────────────────────────────────────────────────────────────────────────────
# A DNJ-shaped fixture (mirrors IVR_Sheet.xlsx column geometry exactly)
# ─────────────────────────────────────────────────────────────────────────────
# 1-based columns the loader cares about: make=3 model=4 year=5 fuel=8 trans=9
# color=12 rate=13 CAR NUMB=14.  Media: exterior 18-27, interior 28-37,
# video 38-42, instagram 43-47, youtube 48-52.
def _make_header_row():
    hdr = [None] * 52
    hdr[3] = "MODEL"          # D
    hdr[5] = "INS"
    hdr[6] = "VARI"
    hdr[7] = "F"
    hdr[8] = "T"
    hdr[10] = "KM"
    hdr[11] = "COL"
    hdr[12] = "RATE"
    hdr[13] = "CAR NUMB"      # N
    hdr[15] = "LOC"
    hdr[16] = "RTO"
    # media groups: keyword cell + numeric continuation cells
    hdr[17] = "EXTERIOR 1 (10 photos)"
    for k, c in enumerate(range(19, 28)):
        hdr[c - 1] = str(k + 2)
    hdr[27] = "INTERIOR 1"
    for k, c in enumerate(range(29, 38)):
        hdr[c - 1] = str(k + 2)
    hdr[37] = "VIDEO 1"
    for k, c in enumerate(range(39, 43)):
        hdr[c - 1] = str(k + 2)
    hdr[42] = "INSTAGRAM 1"
    for k, c in enumerate(range(44, 48)):
        hdr[c - 1] = str(k + 2)
    hdr[47] = "YOUTUBE 1"
    for k, c in enumerate(range(49, 53)):
        hdr[c - 1] = str(k + 2)
    return hdr


def _data_row(stock, make, model, year, fuel, trans, color, rate, car_numb):
    row = [None] * 52
    row[0] = stock
    row[2] = make
    row[3] = model
    row[4] = year
    row[7] = fuel
    row[8] = trans
    row[11] = color
    row[12] = rate
    row[13] = car_numb
    return row


def build_dnj_workbook(path, *, with_media_columns=True):
    """Create a workbook with a DNJ sheet (+ DONT TOUCH SOLD) like the real file."""
    wb = openpyxl.Workbook()
    sold = wb.active
    sold.title = "DONT TOUCH SOLD"
    sold.append(["sold sheet"])           # header noise; no sold regs

    ws = wb.create_sheet("DNJ")
    ws.append(["CTRL F"])                  # row 1 (junk, like real sheet)
    if with_media_columns:
        ws.append(_make_header_row())      # row 2 headers
    else:
        # a minimal A..Q header with no media columns (back-compat case)
        h = [None] * 17
        h[3], h[12], h[13] = "MODEL", "RATE", "CAR NUMB"
        ws.append(h)
    ws.append([None] * 52)                 # row 3 descriptions
    # row 4+ data
    ws.append(_data_row(1, "HYUN", "CRETA", 2019, "P", "A", "WHI", 750000, "MH12AB1234"))
    ws.append(_data_row(2, "MARU", "SWIFT", 2018, "P", "M", "SIL", 550000, "MH14CD5678"))
    ws.append(_data_row(3, "TOYO", "FORTUNER", 2017, "D", "A", "WHI", 2200000, "MH01EF9012"))
    wb.save(path)
    return path


# next-empty-slot writer — mirrors media_sync's Excel-write contract on the DNJ
# layout (located via the SAME header detection the loader uses).
def write_url_to_slot(path, reg, media_type, url):
    wb = openpyxl.load_workbook(path)
    ws = wb["DNJ"]
    header = [ws.cell(M.HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
    layout = M.detect_media_layout(header)
    car_idx = layout.car_col
    target_row = None
    for r in range(M.DATA_START_ROW, ws.max_row + 1):
        if M._norm_reg(ws.cell(r, car_idx).value) == M._norm_reg(reg):
            target_row = r
            break
    assert target_row, f"reg {reg} not found"
    for _slot, col in layout.groups[media_type]:
        if not ws.cell(target_row, col).value:
            ws.cell(target_row, col).value = url
            wb.save(path)
            return col
    raise AssertionError("no free slot")


# ─────────────────────────────────────────────────────────────────────────────
# Tests
# ─────────────────────────────────────────────────────────────────────────────
class TestLayoutDetection(unittest.TestCase):
    def test_detects_all_media_groups_with_correct_columns(self):
        layout = M.detect_media_layout(_make_header_row())
        self.assertEqual(layout.car_col, 14)
        cols = {t: [c for _s, c in v] for t, v in layout.groups.items()}
        self.assertEqual(cols[MediaType.EXTERIOR_PHOTO], list(range(18, 28)))
        self.assertEqual(cols[MediaType.INTERIOR_PHOTO], list(range(28, 38)))
        self.assertEqual(cols[MediaType.VIDEO], list(range(38, 43)))
        self.assertEqual(cols[MediaType.INSTAGRAM], list(range(43, 48)))
        self.assertEqual(cols[MediaType.YOUTUBE], list(range(48, 53)))

    def test_slot_numbers_are_one_based_and_sequential(self):
        layout = M.detect_media_layout(_make_header_row())
        slots = [s for s, _c in layout.groups[MediaType.EXTERIOR_PHOTO]]
        self.assertEqual(slots, list(range(1, 11)))


class TestRowMapping(unittest.TestCase):
    def test_maps_urls_to_inventory_media(self):
        layout = M.detect_media_layout(_make_header_row())
        row = [None] * 52
        row[13] = "MH12AB1234"
        row[17] = "https://proj.supabase.co/storage/v1/object/public/car-media/a.jpg"
        row[37] = "https://proj.supabase.co/storage/v1/object/public/car-media/v.mp4"
        media = M.build_media_for_row("MH12AB1234", tuple(row), layout)
        by_type = {m.media_type: m for m in media}
        self.assertIn(MediaType.EXTERIOR_PHOTO, by_type)
        self.assertIn(MediaType.VIDEO, by_type)
        ext = by_type[MediaType.EXTERIOR_PHOTO]
        self.assertEqual(ext.slot, 1)
        self.assertTrue(ext.is_primary)            # first exterior = hero
        self.assertEqual(ext.registration_no, "MH12AB1234")

    def test_non_url_cells_are_ignored(self):
        layout = M.detect_media_layout(_make_header_row())
        row = [None] * 52
        row[13] = "MH12AB1234"
        row[17] = "TODO upload"                    # not an http(s) URL
        row[18] = "https://x.supabase.co/p/b.jpg"  # valid
        media = M.build_media_for_row("MH12AB1234", tuple(row), layout)
        self.assertEqual(len(media), 1)
        self.assertEqual(media[0].slot, 2)


class TestLoaderAttachesMedia(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.wb = os.path.join(self.dir, "dnj.xlsx")
        build_dnj_workbook(self.wb)

    def test_load_inventory_with_media_attaches(self):
        write_url_to_slot(self.wb, "MH12AB1234", MediaType.EXTERIOR_PHOTO,
                          "https://p.supabase.co/o/ext1.jpg")
        write_url_to_slot(self.wb, "MH12AB1234", MediaType.EXTERIOR_PHOTO,
                          "https://p.supabase.co/o/ext2.jpg")
        items = load_inventory(self.wb)
        creta = next(i for i in items if i.model == "Creta")
        self.assertEqual(len(creta.media), 2)
        self.assertTrue(all(m.media_type == MediaType.EXTERIOR_PHOTO for m in creta.media))

    def test_with_media_false_skips(self):
        write_url_to_slot(self.wb, "MH12AB1234", MediaType.EXTERIOR_PHOTO,
                          "https://p.supabase.co/o/ext1.jpg")
        items = load_inventory(self.wb, with_media=False)
        creta = next(i for i in items if i.model == "Creta")
        self.assertEqual(creta.media, [])

    def test_no_media_columns_is_noop(self):
        wb2 = os.path.join(self.dir, "nomedia.xlsx")
        build_dnj_workbook(wb2, with_media_columns=False)
        items = load_inventory(wb2)                # must not raise
        self.assertTrue(all(not i.media for i in items))


class TestFullChainToMediaService(unittest.TestCase):
    """Upload → URL → Excel → Loader → MediaService → chat response."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.wb = os.path.join(self.dir, "dnj.xlsx")
        build_dnj_workbook(self.wb)

    def _upload_and_record(self, reg, media_type, sub, fname):
        """Run the REAL production store (Upload → public URL), then Excel write."""
        from media_sync_service import InMemoryMediaStore, build_object_path
        from _util import sha1_file

        store = InMemoryMediaStore(project="pilotproj")
        local = os.path.join(self.dir, fname)
        with open(local, "wb") as f:
            f.write(b"\xff\xd8\xff binary-" + fname.encode())
        chash = sha1_file(local)
        ext = fname.rsplit(".", 1)[-1]
        obj = build_object_path(reg, sub, chash, ext)
        url = store.upload("car-media", obj, local, "image/jpeg")
        self.assertTrue(url.startswith("https://pilotproj.supabase.co/"))   # URL gen
        write_url_to_slot(self.wb, reg, media_type, url)                    # Excel update
        return url

    def test_photo_chain(self):
        url1 = self._upload_and_record("MH14CD5678", MediaType.EXTERIOR_PHOTO,
                                       "exterior", "front.jpg")
        url2 = self._upload_and_record("MH14CD5678", MediaType.EXTERIOR_PHOTO,
                                       "exterior", "side.jpg")
        items = load_inventory(self.wb)                       # Inventory Loader
        engine = RetrievalEngine(items)
        svc = MediaService(engine, InventoryMediaProvider())  # Media Service
        resp = svc.get_media("send me photos of the Swift")   # Chat Response
        self.assertEqual(resp["status"], "ok")
        self.assertIn("Swift", resp["vehicle"])
        self.assertEqual(set(resp["photos"]), {url1, url2})

    def test_video_chain(self):
        vurl = self._upload_and_record("MH01EF9012", MediaType.VIDEO,
                                       "video", "walk.mp4")
        items = load_inventory(self.wb)
        svc = MediaService(RetrievalEngine(items), InventoryMediaProvider())
        resp = svc.get_media("any video of the Fortuner?")
        self.assertEqual(resp["status"], "ok")
        self.assertEqual(resp["videos"], [vurl])

    def test_vehicle_without_media_reports_unavailable(self):
        # Creta has no uploads -> media_unavailable (not a crash, not other car's)
        items = load_inventory(self.wb)
        svc = MediaService(RetrievalEngine(items), InventoryMediaProvider())
        resp = svc.get_media("photos of the Creta")
        self.assertEqual(resp["status"], "media_unavailable")
        self.assertEqual(resp["photos"], [])


class TestProductionUploadWriterIsReal(unittest.TestCase):
    """Sanity-check the upstream media_sync writer (Upload → Excel) still passes
    on its native layout — the other half of the chain."""

    def test_poc_run_sync_uploads_and_writes(self):
        from media_sync_poc import (create_test_workbook, run_sync,
                                     InMemoryStorageUploader)
        d = tempfile.mkdtemp()
        wb = os.path.join(d, "wb.xlsx")
        create_test_workbook(wb)
        up = os.path.join(d, "uploads", "MH12AB1234", "exterior")
        os.makedirs(up)
        with open(os.path.join(up, "f.jpg"), "wb") as f:
            f.write(b"\xff\xd8\xff jpeg")
        rep = run_sync(os.path.join(d, "uploads"), wb,
                       uploader=InMemoryStorageUploader())
        self.assertTrue(rep.saved)
        self.assertEqual(rep.succeeded, 1)
        self.assertTrue(rep.results[0].url.startswith("https://"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
