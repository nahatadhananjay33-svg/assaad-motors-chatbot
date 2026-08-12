"""
media_sync_tests.py — production media sync test suite (50+ tests).

Covers: upload/excel success & failure, journal state machine + recovery,
reconciliation (reapply / clobber / missing file / orphan), cleanup (sold,
removed, grace, re-list, dry-run), duplicate handling, and multi-user safety
(locking, atomic writes, no lost updates / duplicates).
"""

import os
import time
import tempfile
import threading
import unittest
from unittest import mock

import openpyxl

import media_journal as J
import media_audit as A
import media_sync_service as S
from media_sync_service import (MediaSyncService, InMemoryMediaStore,
                                build_object_path, apply_pending_to_excel)
from media_reconciliation import ReconciliationEngine
from media_cleanup import CleanupEngine
import _util as U

T0 = "2026-06-10T10:00:00+00:00"
T1 = "2026-06-10T10:05:00+00:00"


def make_workbook(path, vehicles, *, status=True, ext=3, inte=2, vid=2):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Inventory"
    headers = ["CAR NUMB", "MODEL"] + (["STATUS"] if status else [])
    headers += [f"Exterior_Photo_{i}" for i in range(1, ext + 1)]
    headers += [f"Interior_Photo_{i}" for i in range(1, inte + 1)]
    headers += [f"Video_{i}" for i in range(1, vid + 1)]
    ws.append(headers)
    for v in vehicles:
        reg, model = v[0], v[1]
        st = v[2] if len(v) > 2 else "AVAILABLE"
        ws.append([reg, model] + ([st] if status else []) + [None] * (ext + inte + vid))
    wb.save(path)


def put(uploads, reg, mtype, fname, content=b"\xff\xd8\xff data"):
    os.makedirs(os.path.join(uploads, reg, mtype), exist_ok=True)
    with open(os.path.join(uploads, reg, mtype, fname), "wb") as f:
        f.write(content)


class Base(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.uploads = os.path.join(self.dir, "uploads")
        self.wb = os.path.join(self.dir, "inv.xlsx")
        make_workbook(self.wb, [["MH12AB1234", "Creta"], ["MH14CD5678", "Swift"]])
        self.jr = J.MediaJournal()
        self.au = A.MediaAudit()
        self.st = InMemoryMediaStore()
        self.svc = MediaSyncService(self.uploads, self.wb, store=self.st,
                                    journal=self.jr, audit=self.au,
                                    lock_path=self.wb + ".lock")

    def cell(self, reg, header_prefix, idx):
        wb = openpyxl.load_workbook(self.wb); ws = wb.active
        from _poc import read_layout, index_rows
        layout = read_layout(ws); rows = index_rows(ws, layout.car_col)
        r = rows[reg]
        key = {"Exterior": "exterior", "Interior": "interior", "Video": "video"}[header_prefix]
        return ws.cell(r, layout.slots[key][idx]).value

    def recon(self):
        return ReconciliationEngine(self.jr, self.wb, self.st, self.au,
                                    bucket="car-media", lock_path=self.wb + ".lock")

    def cleanup(self, grace=30, dbp=":memory:"):
        return CleanupEngine(self.jr, self.wb, self.st, self.au, bucket="car-media",
                             lock_path=self.wb + ".lock", db_path=dbp, grace_days=grace)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Journal
# ─────────────────────────────────────────────────────────────────────────────
class TestJournal(Base):
    def test_get_or_create_new_is_pending_upload(self):
        e = self.jr.get_or_create(registration_no="MH12AB1234", file_name="a.jpg",
                                  media_type="exterior", supabase_path="p1", now=T0)
        self.assertEqual(e["status"], J.PENDING_UPLOAD)

    def test_get_or_create_idempotent(self):
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="exterior",
                              supabase_path="p1", now=T0)
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="exterior",
                              supabase_path="p1", now=T0)
        self.assertEqual(len(self.jr.all()), 1)

    def test_state_transitions(self):
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="exterior",
                              supabase_path="p", now=T0)
        self.jr.mark_uploaded("p", "http://u", now=T0)
        self.assertEqual(self.jr.get("p")["status"], J.UPLOADED)
        self.jr.mark_pending_excel("p", now=T0)
        self.assertEqual(self.jr.get("p")["status"], J.PENDING_EXCEL)
        self.jr.mark_completed("p", slot=5, now=T0)
        self.assertEqual(self.jr.get("p")["status"], J.COMPLETED)
        self.assertEqual(self.jr.get("p")["slot"], 5)

    def test_mark_failed(self):
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="video",
                              supabase_path="p", now=T0)
        self.jr.mark_failed("p", "boom", now=T0)
        self.assertEqual(self.jr.get("p")["status"], J.FAILED)
        self.assertEqual(self.jr.get("p")["error"], "boom")

    def test_by_status_and_registration(self):
        self.jr.get_or_create(registration_no="R1", file_name="a", media_type="exterior",
                              supabase_path="p1", now=T0)
        self.jr.get_or_create(registration_no="R2", file_name="b", media_type="exterior",
                              supabase_path="p2", now=T0)
        self.jr.mark_uploaded("p1", "u", now=T0)
        self.assertEqual(len(self.jr.by_status(J.UPLOADED)), 1)
        self.assertEqual(len(self.jr.by_registration("R1")), 1)

    def test_delete_registration(self):
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="exterior",
                              supabase_path="p1", now=T0)
        self.jr.get_or_create(registration_no="R", file_name="b", media_type="video",
                              supabase_path="p2", now=T0)
        self.assertEqual(self.jr.delete_registration("R"), 2)
        self.assertEqual(self.jr.all(), [])

    def test_counts(self):
        self.jr.get_or_create(registration_no="R", file_name="a", media_type="exterior",
                              supabase_path="p", now=T0)
        self.assertEqual(self.jr.counts().get(J.PENDING_UPLOAD), 1)

    def test_registrations(self):
        self.jr.get_or_create(registration_no="R1", file_name="a", media_type="exterior",
                              supabase_path="p1", now=T0)
        self.assertEqual(self.jr.registrations(), ["R1"])


# ─────────────────────────────────────────────────────────────────────────────
# 2. Audit
# ─────────────────────────────────────────────────────────────────────────────
class TestAudit(Base):
    def test_log_and_query(self):
        self.au.log(A.UPLOAD, registration="R", user="alice", now=T0)
        self.assertEqual(self.au.count(A.UPLOAD), 1)
        self.assertEqual(self.au.by_registration("R")[0]["user"], "alice")

    def test_failure_helper(self):
        self.au.failure(registration="R", detail="x", now=T0)
        self.assertEqual(self.au.by_action(A.FAILURE)[0]["result"], "failure")

    def test_summary(self):
        self.au.log(A.UPLOAD, now=T0); self.au.log(A.CLEANUP, now=T0)
        s = self.au.summary()
        self.assertEqual(s[A.UPLOAD], 1); self.assertEqual(s[A.CLEANUP], 1)

    def test_records_user(self):
        self.au.log(A.EXCEL_UPDATE, user="bob", now=T0)
        self.assertEqual(self.au.all()[0]["user"], "bob")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Upload + Excel success
# ─────────────────────────────────────────────────────────────────────────────
class TestUploadExcelSuccess(Base):
    def test_photo_uploaded_and_written(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        r = self.svc.sync_once(user="alice", now=T0)
        self.assertEqual(r.uploaded, 1)
        self.assertTrue(r.excel["saved"])
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))

    def test_video_routing(self):
        put(self.uploads, "MH12AB1234", "video", "v.mp4")
        self.svc.sync_once(now=T0)
        self.assertIsNotNone(self.cell("MH12AB1234", "Video", 0))

    def test_journal_completed(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        self.assertEqual(self.jr.counts().get(J.COMPLETED), 1)

    def test_audit_records_upload_and_excel(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(user="alice", now=T0)
        self.assertEqual(self.au.count(A.UPLOAD), 1)
        self.assertEqual(self.au.count(A.EXCEL_UPDATE), 1)

    def test_unknown_vehicle_not_uploaded(self):
        put(self.uploads, "MH99ZZ0000", "exterior", "a.jpg")
        r = self.svc.sync_once(now=T0)
        self.assertEqual(r.count(S.ST_UNKNOWN_VEHICLE), 1)
        self.assertEqual(len(self.st.objects), 0)

    def test_invalid_extension(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.txt")
        r = self.svc.sync_once(now=T0)
        self.assertEqual(r.count(S.ST_INVALID_FILE), 1)

    def test_empty_file(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg", content=b"")
        r = self.svc.sync_once(now=T0)
        self.assertEqual(r.count(S.ST_INVALID_FILE), 1)

    def test_multiple_photos_fill_slots(self):
        for n in ("a.jpg", "b.jpg", "c.jpg"):
            put(self.uploads, "MH12AB1234", "exterior", n, content=b"\xff" + n.encode())
        self.svc.sync_once(now=T0)
        vals = [self.cell("MH12AB1234", "Exterior", i) for i in range(3)]
        self.assertTrue(all(vals))
        self.assertEqual(len(set(vals)), 3)

    def test_existing_data_preserved(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        wb = openpyxl.load_workbook(self.wb)
        self.assertEqual(wb.active.cell(2, 2).value, "Creta")   # MODEL intact

    def test_correct_row_only(self):
        put(self.uploads, "MH14CD5678", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        self.assertIsNotNone(self.cell("MH14CD5678", "Exterior", 0))
        self.assertIsNone(self.cell("MH12AB1234", "Exterior", 0))


# ─────────────────────────────────────────────────────────────────────────────
# 4. Upload failure / Excel failure
# ─────────────────────────────────────────────────────────────────────────────
class TestFailures(Base):
    def test_upload_failure_recorded(self):
        put(self.uploads, "MH12AB1234", "video", "v.mp4")
        self.st.fail_on = {"v.mp4"}
        r = self.svc.sync_once(now=T0)
        self.assertEqual(r.count(S.ST_UPLOAD_ERROR), 1)
        self.assertEqual(self.jr.counts().get(J.FAILED), 1)
        self.assertEqual(self.au.count(A.FAILURE), 1)

    def test_excel_write_failure_keeps_pending(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        with mock.patch.object(S, "atomic_save_workbook", side_effect=IOError("disk full")):
            r = self.svc.sync_once(now=T0)
        self.assertFalse(r.excel["saved"])
        self.assertEqual(self.jr.counts().get(J.PENDING_EXCEL), 1)   # upload kept
        self.assertEqual(len(self.st.objects), 1)                    # file is in storage
        # excel unchanged on disk
        self.assertIsNone(self.cell("MH12AB1234", "Exterior", 0))

    def test_excel_failure_then_recovered_next_sync(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        with mock.patch.object(S, "atomic_save_workbook", side_effect=IOError("x")):
            self.svc.sync_once(now=T0)
        self.svc.sync_once(now=T1)                                    # save works now
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))
        self.assertEqual(self.jr.counts().get(J.COMPLETED), 1)

    def test_defer_when_excel_open(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        # simulate Office lock file
        open(os.path.join(self.dir, "~$inv.xlsx"), "w").close()
        r = self.svc.sync_once(now=T0)
        self.assertEqual(r.excel.get("deferred"), 1)
        self.assertIsNone(self.cell("MH12AB1234", "Exterior", 0))    # not written yet


# ─────────────────────────────────────────────────────────────────────────────
# 5. Duplicate handling / idempotency
# ─────────────────────────────────────────────────────────────────────────────
class TestDuplicates(Base):
    def test_rerun_no_duplicate(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        r2 = self.svc.sync_once(now=T1)
        self.assertEqual(r2.count(S.ST_DUPLICATE), 1)
        self.assertEqual(r2.uploaded, 0)

    def test_no_duplicate_cells(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0); self.svc.sync_once(now=T1)
        vals = [self.cell("MH12AB1234", "Exterior", i) for i in range(3)]
        self.assertEqual(len([v for v in vals if v]), 1)

    def test_same_content_same_path(self):
        c = b"\xff\xd8 identical"
        h1 = U.sha1_file.__wrapped__ if hasattr(U.sha1_file, "__wrapped__") else None
        put(self.uploads, "MH12AB1234", "exterior", "x.jpg", content=c)
        put(self.uploads, "MH14CD5678", "exterior", "y.jpg", content=c)
        self.svc.sync_once(now=T0)
        # two different vehicles, same bytes -> two objects (paths include reg)
        self.assertEqual(len(self.st.objects), 2)

    def test_no_duplicate_uploads_on_rerun(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        before = len(self.st.objects)
        self.svc.sync_once(now=T1)
        self.assertEqual(len(self.st.objects), before)

    def test_slot_exhaustion(self):
        for i in range(4):                       # 4 photos, 3 slots
            put(self.uploads, "MH12AB1234", "exterior", f"{i}.jpg", content=b"\xff" + str(i).encode())
        self.svc.sync_once(now=T0)
        vals = [self.cell("MH12AB1234", "Exterior", i) for i in range(3)]
        self.assertEqual(len([v for v in vals if v]), 3)             # 3 written, 4th waits


# ─────────────────────────────────────────────────────────────────────────────
# 6. Reconciliation
# ─────────────────────────────────────────────────────────────────────────────
class TestReconciliation(Base):
    def test_reapply_after_excel_failure(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        with mock.patch.object(S, "atomic_save_workbook", side_effect=IOError("x")):
            self.svc.sync_once(now=T0)
        rep = self.recon().reconcile(now=T1)
        self.assertEqual(rep.excel_reapplied, 1)
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))

    def test_clobber_repaired(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        # human save blanks the media cell (first exterior slot column)
        from _poc import read_layout
        wb = openpyxl.load_workbook(self.wb); ws = wb.active
        col = read_layout(ws).slots["exterior"][0]
        ws.cell(2, col).value = None; wb.save(self.wb)
        rep = self.recon().reconcile(now=T1)
        self.assertEqual(rep.clobbers_repaired, 1)
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))   # restored

    def test_missing_storage_file_flagged(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        self.st.objects.clear()                          # file vanished from storage
        rep = self.recon().reconcile(now=T1)
        self.assertEqual(rep.missing_file_flagged, 1)
        self.assertEqual(self.jr.counts().get(J.FAILED), 1)

    def test_orphan_storage_detected(self):
        self.st.objects["car-media/MH00XX0000/exterior/deadbeef.jpg"] = b"x"  # no journal
        rep = self.recon().reconcile(now=T0)
        self.assertEqual(len(rep.orphan_storage), 1)

    def test_reconciliation_report_clean(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        rep = self.recon().reconciliation_report()
        self.assertEqual(rep["missing_storage_files"], [])
        self.assertEqual(rep["orphan_storage_files"], [])
        self.assertEqual(rep["clobbered_excel"], [])

    def test_reconciliation_report_shows_pending(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        with mock.patch.object(S, "atomic_save_workbook", side_effect=IOError("x")):
            self.svc.sync_once(now=T0)
        self.assertEqual(self.recon().reconciliation_report()["pending_excel"], 1)

    def test_reconcile_audits(self):
        self.recon().reconcile(now=T0)
        self.assertGreaterEqual(self.au.count(A.RECONCILIATION), 1)


# ─────────────────────────────────────────────────────────────────────────────
# 7. Cleanup / sold workflow / grace
# ─────────────────────────────────────────────────────────────────────────────
class TestCleanup(Base):
    def _sync_creta(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)

    def _mark_sold(self, reg="MH12AB1234"):
        wb = openpyxl.load_workbook(self.wb); ws = wb.active
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == reg:
                ws.cell(r, 3).value = "SOLD"
        wb.save(self.wb)

    def test_sold_marks_pending(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup()
        rep = cl.scan(now="2026-07-01T10:00:00+00:00")
        self.assertEqual(rep.marked, 1)
        self.assertEqual(cl.state_of("MH12AB1234"), "sold_pending_cleanup")

    def test_no_delete_within_grace(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup(grace=30)
        cl.scan(now="2026-07-01T10:00:00+00:00")
        rep = cl.run(now="2026-07-15T10:00:00+00:00")          # < 30 days
        self.assertEqual(rep.deleted, 0)
        self.assertEqual(len(self.st.objects), 1)

    def test_delete_after_grace(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup(grace=30)
        cl.scan(now="2026-07-01T10:00:00+00:00")
        rep = cl.run(now="2026-08-15T10:00:00+00:00")          # > 30 days
        self.assertEqual(rep.deleted, 1)
        self.assertEqual(rep.files_deleted, 1)
        self.assertEqual(len(self.st.objects), 0)
        self.assertEqual(self.jr.all(), [])                    # journal cleared

    def test_delete_clears_excel_cells(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup(grace=30)
        cl.scan(now="2026-07-01T10:00:00+00:00")
        cl.run(now="2026-08-15T10:00:00+00:00")
        self.assertIsNone(self.cell("MH12AB1234", "Exterior", 0))

    def test_relist_cancels_cleanup(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup()
        cl.scan(now="2026-07-01T10:00:00+00:00")
        # re-list: STATUS back to AVAILABLE
        wb = openpyxl.load_workbook(self.wb); wb.active.cell(2, 3).value = "AVAILABLE"; wb.save(self.wb)
        rep = cl.scan(now="2026-07-05T10:00:00+00:00")
        self.assertEqual(rep.cancelled, 1)
        self.assertIsNone(cl.state_of("MH12AB1234"))

    def test_removed_row_eligible(self):
        self._sync_creta()
        # remove the row entirely
        wb = openpyxl.load_workbook(self.wb); wb.active.delete_rows(2); wb.save(self.wb)
        cl = self.cleanup()
        rep = cl.scan(now="2026-07-01T10:00:00+00:00")
        self.assertEqual(rep.marked, 1)

    def test_dry_run_deletes_nothing(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup(grace=30)
        cl.scan(now="2026-07-01T10:00:00+00:00")
        rep = cl.run(now="2026-08-15T10:00:00+00:00", dry_run=True)
        self.assertEqual(rep.deleted, 0)
        self.assertEqual(len(self.st.objects), 1)              # nothing removed
        self.assertTrue(rep.details)

    def test_cleanup_report(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup()
        cl.scan(now="2026-07-01T10:00:00+00:00")
        rep = cl.cleanup_report()
        self.assertEqual(rep["grace_days"], 30)
        self.assertEqual(rep["by_state"].get("sold_pending_cleanup"), 1)

    def test_cleanup_audits(self):
        self._sync_creta(); self._mark_sold()
        cl = self.cleanup(grace=30)
        cl.scan(now="2026-07-01T10:00:00+00:00")
        cl.run(now="2026-08-15T10:00:00+00:00")
        self.assertGreaterEqual(self.au.count(A.CLEANUP), 2)   # marked + deleted


# ─────────────────────────────────────────────────────────────────────────────
# 8. Multi-user safety
# ─────────────────────────────────────────────────────────────────────────────
class TestMultiUser(Base):
    def test_filelock_mutual_exclusion(self):
        lock = self.wb + ".lk"
        a = U.FileLock(lock, timeout=0.3)
        a.acquire()
        with self.assertRaises(TimeoutError):
            U.FileLock(lock, timeout=0.2).acquire()
        a.release()
        U.FileLock(lock, timeout=0.5).acquire().release()      # free again

    def test_concurrent_sync_no_corruption(self):
        for reg, n in [("MH12AB1234", "a.jpg"), ("MH14CD5678", "b.jpg")]:
            put(self.uploads, reg, "exterior", n, content=b"\xff" + n.encode())
        errs = []

        def worker():
            try:
                self.svc.sync_once(user="u", now=T0)
            except Exception as e:               # noqa
                errs.append(e)

        ts = [threading.Thread(target=worker) for _ in range(3)]
        [t.start() for t in ts]; [t.join() for t in ts]
        self.assertEqual(errs, [])
        # both vehicles written exactly once; no duplicate cells
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))
        self.assertIsNotNone(self.cell("MH14CD5678", "Exterior", 0))
        self.assertIsNone(self.cell("MH12AB1234", "Exterior", 1))
        self.assertEqual(self.jr.counts().get(J.COMPLETED), 2)

    def test_no_lost_update_two_vehicles(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        put(self.uploads, "MH14CD5678", "video", "v.mp4")
        self.svc.sync_once(now=T1)
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))   # earlier write kept
        self.assertIsNotNone(self.cell("MH14CD5678", "Video", 0))

    def test_atomic_save_leaves_no_temp(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.svc.sync_once(now=T0)
        leftovers = [f for f in os.listdir(self.dir) if f.endswith(".xlsx") and f != "inv.xlsx"]
        self.assertEqual(leftovers, [])

    def test_backup_created_when_configured(self):
        bdir = os.path.join(self.dir, "backups")
        svc = MediaSyncService(self.uploads, self.wb, store=self.st, journal=self.jr,
                               audit=self.au, lock_path=self.wb + ".lock", backup_dir=bdir)
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        svc.sync_once(now=T0)
        self.assertTrue(os.path.isdir(bdir) and len(os.listdir(bdir)) >= 1)


# ─────────────────────────────────────────────────────────────────────────────
# 9. End-to-end journal recovery + utils
# ─────────────────────────────────────────────────────────────────────────────
class TestEndToEndRecovery(Base):
    def test_full_recovery_cycle(self):
        put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        with mock.patch.object(S, "atomic_save_workbook", side_effect=IOError("x")):
            self.svc.sync_once(now=T0)                 # storage ok, excel failed
        self.assertEqual(self.jr.counts().get(J.PENDING_EXCEL), 1)
        self.recon().reconcile(now=T1)                 # recover
        self.assertEqual(self.jr.counts().get(J.COMPLETED), 1)
        self.assertIsNotNone(self.cell("MH12AB1234", "Exterior", 0))


class TestUtils(unittest.TestCase):
    def test_add_days_and_is_past(self):
        d = U.add_days(T0, 30)
        self.assertTrue(U.is_past(d, "2026-08-01T00:00:00+00:00"))
        self.assertFalse(U.is_past(d, "2026-06-20T00:00:00+00:00"))

    def test_excel_open_detection(self):
        dd = tempfile.mkdtemp()
        p = os.path.join(dd, "x.xlsx")
        open(p, "w").close()
        self.assertFalse(U.excel_open_by_user(p))
        open(os.path.join(dd, "~$x.xlsx"), "w").close()
        self.assertTrue(U.excel_open_by_user(p))

    def test_sha1_stable(self):
        dd = tempfile.mkdtemp(); p = os.path.join(dd, "f")
        with open(p, "wb") as f: f.write(b"hello")
        self.assertEqual(U.sha1_file(p), U.sha1_file(p))

    def test_build_object_path(self):
        self.assertEqual(build_object_path("MH12AB1234", "exterior", "abc", "jpg"),
                         "MH12AB1234/exterior/abc.jpg")


if __name__ == "__main__":
    unittest.main(verbosity=2)
