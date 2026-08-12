"""
media_cleanup.py — grace-period media cleanup for sold/removed vehicles.

Inventory stays in Excel. When a row's STATUS = SOLD, or a row is removed, its
media becomes eligible for cleanup — but never deleted immediately:

    active → sold_pending_cleanup → (grace period, default 30 days) → delete

Delete = Supabase files + journal entries + Excel media cells (orphan metadata).
Re-listing (STATUS back to AVAILABLE) within the grace window cancels cleanup.
`cleanup_report()` summarizes state.
"""

from __future__ import annotations

import sqlite3
import threading
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import openpyxl

from _poc import read_layout, index_rows, normalize_registration
from _util import (utcnow_iso, add_days, is_past, FileLock, atomic_save_workbook,
                   excel_open_by_user)
import media_journal as J
import media_audit as A

DEFAULT_GRACE_DAYS = 30

# cleanup states
ACTIVE = "active"
PENDING = "sold_pending_cleanup"
DELETED = "deleted"

_SCHEMA = """
CREATE TABLE IF NOT EXISTS cleanup_state (
    registration_no TEXT PRIMARY KEY,
    state           TEXT NOT NULL,
    reason          TEXT,
    marked_at       TEXT,
    delete_after    TEXT,
    deleted_at      TEXT
);
"""

_SOLD_STATUS = "SOLD"
_STATUS_HEADERS = {"STATUS", "LISTING_STATUS", "STATE"}


def _find_status_col(ws):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().upper() in _STATUS_HEADERS:
            return c
    return None


@dataclass
class CleanupReport:
    marked: int = 0
    cancelled: int = 0
    deleted: int = 0
    files_deleted: int = 0
    dry_run: bool = False
    details: List[Dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> Dict[str, Any]:
        return {"marked": self.marked, "cancelled": self.cancelled,
                "deleted": self.deleted, "files_deleted": self.files_deleted,
                "dry_run": self.dry_run, "details": self.details}


class CleanupEngine:
    def __init__(self, journal: J.MediaJournal, workbook_path: str, store,
                 audit: A.MediaAudit, *, bucket: str, lock_path: str,
                 db_path: str = ":memory:", grace_days: int = DEFAULT_GRACE_DAYS,
                 backup_dir: Optional[str] = None, clock=utcnow_iso):
        self.journal = journal
        self.workbook_path = workbook_path
        self.store = store
        self.audit = audit
        self.bucket = bucket
        self.lock_path = lock_path
        self.grace_days = grace_days
        self.backup_dir = backup_dir
        self.clock = clock
        self._lock = threading.Lock()
        self._conn = sqlite3.connect(db_path, check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._conn.execute(_SCHEMA)
        self._conn.commit()

    # ── workbook scan: who is SOLD, who is still present ──
    def _scan_workbook(self):
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb.active
        layout = read_layout(ws)
        status_col = _find_status_col(ws)
        present, sold = set(), set()
        if layout.car_col:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, layout.car_col).value
                if not v:
                    continue
                reg = normalize_registration(str(v))
                present.add(reg)
                if status_col and str(ws.cell(r, status_col).value or "").strip().upper() == _SOLD_STATUS:
                    sold.add(reg)
        wb.close()
        return present, sold

    # ── mark eligible (sold OR removed); cancel re-listed ──
    def scan(self, *, user: str = "system", now: Optional[str] = None) -> CleanupReport:
        now = now or self.clock()
        report = CleanupReport()
        present, sold = self._scan_workbook()
        journal_regs = set(self.journal.registrations())
        removed = journal_regs - present                  # had media, row gone
        eligible = sold | removed

        with self._lock:
            existing = {r["registration_no"]: dict(r) for r in
                        self._conn.execute("SELECT * FROM cleanup_state").fetchall()}

            for reg in eligible:
                if reg in existing and existing[reg]["state"] != ACTIVE:
                    continue
                reason = "sold" if reg in sold else "row_removed"
                self._conn.execute(
                    "INSERT OR REPLACE INTO cleanup_state "
                    "(registration_no,state,reason,marked_at,delete_after,deleted_at) "
                    "VALUES (?,?,?,?,?,NULL)",
                    (reg, PENDING, reason, now, add_days(now, self.grace_days)))
                report.marked += 1
                self.audit.log(A.CLEANUP, registration=reg, result="marked",
                               user=user, detail=f"{reason}; grace {self.grace_days}d", now=now)

            # re-listed (now AVAILABLE again and present) → cancel pending cleanup
            for reg, row in existing.items():
                if row["state"] == PENDING and reg in present and reg not in sold:
                    self._conn.execute("DELETE FROM cleanup_state WHERE registration_no=?", (reg,))
                    report.cancelled += 1
                    self.audit.log(A.CLEANUP, registration=reg, result="cancelled",
                                   user=user, detail="re-listed within grace", now=now)
            self._conn.commit()
        return report

    # ── delete media for entries whose grace has elapsed ──
    def run(self, *, user: str = "system", now: Optional[str] = None,
            dry_run: bool = False) -> CleanupReport:
        now = now or self.clock()
        report = CleanupReport(dry_run=dry_run)
        with self._lock:
            due = [dict(r) for r in self._conn.execute(
                "SELECT * FROM cleanup_state WHERE state=?", (PENDING,)).fetchall()]
        due = [r for r in due if is_past(r["delete_after"], now)]

        for row in due:
            reg = row["registration_no"]
            entries = self.journal.by_registration(reg)
            paths = [e["supabase_path"] for e in entries]
            if dry_run:
                report.details.append({"registration": reg, "would_delete": len(paths)})
                continue

            deleted_files = 0
            for p in paths:
                try:
                    if self.store.delete(self.bucket, p):
                        deleted_files += 1
                except Exception:
                    pass
            self.journal.delete_registration(reg)         # remove orphan metadata
            self._clear_excel_cells(reg, now)             # blank media cells if row present
            with self._lock:
                self._conn.execute(
                    "UPDATE cleanup_state SET state=?, deleted_at=? WHERE registration_no=?",
                    (DELETED, now, reg))
                self._conn.commit()
            report.deleted += 1
            report.files_deleted += deleted_files
            self.audit.log(A.CLEANUP, registration=reg, result="deleted", user=user,
                           detail=f"{deleted_files} files + journal + cells", now=now)
        return report

    def _clear_excel_cells(self, reg: str, now: str) -> None:
        if excel_open_by_user(self.workbook_path):
            return
        with FileLock(self.lock_path):
            wb = openpyxl.load_workbook(self.workbook_path)
            ws = wb.active
            layout = read_layout(ws)
            rows = index_rows(ws, layout.car_col) if layout.car_col else {}
            if reg not in rows:
                wb.close(); return
            ri = rows[reg]
            changed = False
            for cols in layout.slots.values():
                for c in cols:
                    if ws.cell(ri, c).value:
                        ws.cell(ri, c).value = None
                        changed = True
            if changed:
                atomic_save_workbook(wb, self.workbook_path, self.backup_dir, now)
            else:
                wb.close()

    # ── report ──
    def cleanup_report(self) -> Dict[str, Any]:
        with self._lock:
            rows = self._conn.execute(
                "SELECT state, COUNT(*) FROM cleanup_state GROUP BY state").fetchall()
            by_state = {s: n for s, n in rows}
            pending = [dict(r) for r in self._conn.execute(
                "SELECT registration_no, reason, delete_after FROM cleanup_state "
                "WHERE state=?", (PENDING,)).fetchall()]
        return {"by_state": by_state, "pending": pending,
                "grace_days": self.grace_days}

    def state_of(self, reg: str) -> Optional[str]:
        with self._lock:
            r = self._conn.execute(
                "SELECT state FROM cleanup_state WHERE registration_no=?", (reg,)).fetchone()
            return r[0] if r else None

    def close(self) -> None:
        self._conn.close()
