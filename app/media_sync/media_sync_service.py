"""
media_sync_service.py — production media synchronization layer.

    Folder Watcher → Upload → URL Generation → Excel Update

Reuses the validated POC pipeline (scan, registration matching, slot writes) and
adds production concerns: a recovery JOURNAL, an AUDIT trail, cross-process
LOCKING, ATOMIC workbook writes, and safe RETRIES — so 10–20 staff can upload
photos/videos and edit Excel concurrently without corruption, duplicates, or
lost updates.

Excel (DNJ) stays the ONLY source of truth; Supabase Storage holds files only.
This layer writes ONLY media URL columns into Excel.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import openpyxl

from _poc import (scan_uploads, normalize_registration, read_layout, index_rows,
                  content_type, _ext, _valid_ext_for_type,
                  InMemoryStorageUploader, SupabaseStorageUploader)
from _util import (FileLock, atomic_save_workbook, excel_open_by_user, sha1_file,
                   utcnow_iso)
import media_journal as J
import media_audit as A

DEFAULT_BUCKET = "car-media"

# upload-phase result statuses
ST_UPLOADED = "uploaded"
ST_ALREADY_UPLOADED = "already_uploaded"
ST_DUPLICATE = "skipped_duplicate"
ST_UNKNOWN_VEHICLE = "unknown_vehicle"
ST_INVALID_FILE = "invalid_file"
ST_UPLOAD_ERROR = "upload_error"


# ─────────────────────────────────────────────────────────────────────────────
# Media store (production): reuse POC uploaders + add exists/delete/list
# ─────────────────────────────────────────────────────────────────────────────
class InMemoryMediaStore(InMemoryStorageUploader):
    def exists(self, bucket: str, object_path: str) -> bool:
        return f"{bucket}/{object_path}" in self.objects

    def delete(self, bucket: str, object_path: str) -> bool:
        return self.objects.pop(f"{bucket}/{object_path}", None) is not None

    def list_paths(self, bucket: str) -> List[str]:
        pre = bucket + "/"
        return sorted(k[len(pre):] for k in self.objects if k.startswith(pre))


class SupabaseMediaStore(SupabaseStorageUploader):
    """Storage-only adapter (design parity with InMemoryMediaStore)."""
    def exists(self, bucket: str, object_path: str) -> bool:  # pragma: no cover
        folder = os.path.dirname(object_path)
        name = os.path.basename(object_path)
        listed = self._client.storage.from_(bucket).list(folder)
        return any(x.get("name") == name for x in (listed or []))

    def delete(self, bucket: str, object_path: str) -> bool:  # pragma: no cover
        self._client.storage.from_(bucket).remove([object_path])
        return True

    def list_paths(self, bucket: str) -> List[str]:           # pragma: no cover
        return [x.get("name") for x in (self._client.storage.from_(bucket).list() or [])]


def build_object_path(reg: str, media_type: str, content_hash: str, ext: str) -> str:
    return f"{reg}/{media_type}/{content_hash}.{ext}"


# ─────────────────────────────────────────────────────────────────────────────
# Result types
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class FileResult:
    filename: str
    registration: str
    media_type: str
    status: str
    url: Optional[str] = None
    detail: Optional[str] = None


@dataclass
class SyncRunReport:
    uploads: List[FileResult] = field(default_factory=list)
    excel: Dict[str, Any] = field(default_factory=dict)

    def count(self, status: str) -> int:
        return sum(1 for r in self.uploads if r.status == status)

    @property
    def uploaded(self) -> int:
        return self.count(ST_UPLOADED)

    def summary(self) -> Dict[str, Any]:
        s: Dict[str, Any] = {}
        for r in self.uploads:
            s[r.status] = s.get(r.status, 0) + 1
        s["excel"] = self.excel
        return s


# ─────────────────────────────────────────────────────────────────────────────
# Excel write phase (shared by the service AND reconciliation)
# ─────────────────────────────────────────────────────────────────────────────
def apply_pending_to_excel(journal: J.MediaJournal, workbook_path: str,
                           audit: A.MediaAudit, *, bucket: str, lock_path: str,
                           backup_dir: Optional[str], user: str,
                           now: Optional[str] = None,
                           defer_if_open: bool = True) -> Dict[str, Any]:
    """Write every uploaded/pending-excel URL into the right DNJ slot — under a
    cross-process lock, with an atomic save. Idempotent and lost-update-safe."""
    now = now or utcnow_iso()
    pending = [e for e in (journal.by_status(J.UPLOADED) + journal.by_status(J.PENDING_EXCEL))
               if e.get("public_url")]
    if not pending:
        return {"written": 0, "completed": 0, "deferred": 0, "saved": True}

    if defer_if_open and excel_open_by_user(workbook_path):
        for e in pending:
            journal.mark_pending_excel(e["supabase_path"], now=now)
        audit.log(A.EXCEL_UPDATE, result="deferred", user=user,
                  detail="excel open by a user; deferred", now=now)
        return {"written": 0, "completed": 0, "deferred": len(pending),
                "saved": True, "reason": "excel_open"}

    with FileLock(lock_path):
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active
        layout = read_layout(ws)
        rows = index_rows(ws, layout.car_col)

        written = 0
        to_complete: List[tuple] = []          # (entry, slot_col)
        for e in pending:
            reg, mtype, url = e["registration_no"], e["media_type"], e["public_url"]
            if reg not in rows:                # row not in sheet (yet/removed) -> leave
                continue
            ri = rows[reg]
            slot_cols = layout.slots.get(mtype, [])
            if not slot_cols:
                continue
            existing = [ws.cell(ri, c).value for c in slot_cols]
            if url in [v for v in existing if v]:          # already present -> done
                col = slot_cols[next(i for i, v in enumerate(existing) if v == url)]
                to_complete.append((e, col))
                continue
            empty = next((c for c in slot_cols if not ws.cell(ri, c).value), None)
            if empty is None:                  # no free slot -> leave pending
                continue
            ws.cell(ri, empty).value = url
            written += 1
            to_complete.append((e, empty))

        if written > 0:
            try:
                atomic_save_workbook(wb, workbook_path, backup_dir, now)
            except Exception as ex:
                for e in pending:
                    journal.mark_pending_excel(e["supabase_path"], now=now)
                audit.failure(detail=f"excel save failed: {ex}", user=user, now=now)
                return {"written": 0, "completed": 0, "deferred": 0,
                        "saved": False, "error": str(ex)}

    # mark completed + audit (outside the lock)
    for e, col in to_complete:
        journal.mark_completed(e["supabase_path"], slot=col, now=now)
        audit.log(A.EXCEL_UPDATE, registration=e["registration_no"], result="success",
                  user=user, file_name=e["file_name"], media_type=e["media_type"], now=now)
    return {"written": written, "completed": len(to_complete), "deferred": 0, "saved": True}


# ─────────────────────────────────────────────────────────────────────────────
# The service
# ─────────────────────────────────────────────────────────────────────────────
class MediaSyncService:
    def __init__(self, uploads_dir: str, workbook_path: str, *, store,
                 journal: J.MediaJournal, audit: A.MediaAudit,
                 bucket: str = DEFAULT_BUCKET, lock_path: Optional[str] = None,
                 backup_dir: Optional[str] = None, clock=utcnow_iso):
        self.uploads_dir = uploads_dir
        self.workbook_path = workbook_path
        self.store = store
        self.journal = journal
        self.audit = audit
        self.bucket = bucket
        self.lock_path = lock_path or (workbook_path + ".lock")
        self.backup_dir = backup_dir
        self.clock = clock

    def _workbook_registrations(self) -> set:
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb.active
        layout = read_layout(ws)
        regs = set(index_rows(ws, layout.car_col).keys()) if layout.car_col else set()
        wb.close()
        return regs

    def _process_uploads(self, user: str, now: str) -> List[FileResult]:
        results: List[FileResult] = []
        valid_regs = self._workbook_registrations()
        for item in scan_uploads(self.uploads_dir):
            nreg = normalize_registration(item.registration)
            ext = _ext(item.filename)
            fr = FileResult(item.filename, nreg, item.media_type, ST_UPLOADED)

            if not _valid_ext_for_type(item.media_type, ext):
                fr.status, fr.detail = ST_INVALID_FILE, f"bad extension '.{ext}'"
                self.audit.failure(registration=nreg, user=user, file_name=item.filename,
                                   media_type=item.media_type, detail=fr.detail, now=now)
                results.append(fr); continue
            if not os.path.isfile(item.path) or os.path.getsize(item.path) == 0:
                fr.status, fr.detail = ST_INVALID_FILE, "missing or empty file"
                self.audit.failure(registration=nreg, user=user, file_name=item.filename,
                                   media_type=item.media_type, detail=fr.detail, now=now)
                results.append(fr); continue
            if nreg not in valid_regs:           # quarantine: never upload orphan media
                fr.status, fr.detail = ST_UNKNOWN_VEHICLE, "no matching CAR NUMB row"
                self.audit.failure(registration=nreg, user=user, file_name=item.filename,
                                   media_type=item.media_type, detail=fr.detail, now=now)
                results.append(fr); continue

            chash = sha1_file(item.path)
            object_path = build_object_path(nreg, item.media_type, chash, ext)
            entry = self.journal.get_or_create(
                registration_no=nreg, file_name=item.filename, media_type=item.media_type,
                supabase_path=object_path, content_hash=chash, now=now)

            if entry["status"] == J.COMPLETED:
                fr.status, fr.url = ST_DUPLICATE, entry["public_url"]
                results.append(fr); continue
            if entry["status"] in (J.UPLOADED, J.PENDING_EXCEL) and entry["public_url"]:
                fr.status, fr.url = ST_ALREADY_UPLOADED, entry["public_url"]
                results.append(fr); continue

            # upload (pending_upload or a prior failure) — idempotent store path
            try:
                url = self.store.upload(self.bucket, object_path, item.path, content_type(ext))
                self.journal.mark_uploaded(object_path, url, now=now)
                self.audit.log(A.UPLOAD, registration=nreg, result="success", user=user,
                               file_name=item.filename, media_type=item.media_type, now=now)
                fr.url, fr.status = url, ST_UPLOADED
            except Exception as ex:
                self.journal.mark_failed(object_path, str(ex), now=now)
                self.audit.failure(registration=nreg, user=user, file_name=item.filename,
                                   media_type=item.media_type,
                                   detail=f"upload failed: {ex}", now=now)
                fr.status, fr.detail = ST_UPLOAD_ERROR, str(ex)
            results.append(fr)
        return results

    def sync_once(self, *, user: str = "system", now: Optional[str] = None,
                  process_uploads: bool = True) -> SyncRunReport:
        now = now or self.clock()
        uploads = self._process_uploads(user, now) if process_uploads else []
        excel = apply_pending_to_excel(
            self.journal, self.workbook_path, self.audit, bucket=self.bucket,
            lock_path=self.lock_path, backup_dir=self.backup_dir, user=user, now=now)
        return SyncRunReport(uploads, excel)
