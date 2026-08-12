"""
media_sync_poc.py
================

Proof of concept for the Excel-master + Supabase-media sync (Phase 4A.3-PoC).

Proves the pipeline:
    Upload Folder → Supabase Storage → Public URL → Excel Update

POC scope: PHOTOS + VIDEOS only. NO Instagram/YouTube, NO cleanup, NO audit log,
NO concurrency, NO production architecture. Self-contained: does NOT touch the
chatbot or the inventory engine.

Vehicle identity = the upload **folder name** = the DNJ **CAR NUMB** column
(registration number). Never row numbers, never inventory ids.

Storage is pluggable so the POC runs offline:
  * `InMemoryStorageUploader`  — used by tests / dry runs (no network).
  * `SupabaseStorageUploader`  — reuses the legacy Storage upload + public-URL
                                 pattern (lazy `supabase` import). Storage ONLY.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Callable, Dict, List, Optional, Protocol, Tuple

import openpyxl

# ── allowed media (POC: photos + videos only) ──
PHOTO_EXT = {"jpg", "jpeg", "png", "webp"}
VIDEO_EXT = {"mp4", "mov"}
MEDIA_TYPES = ("exterior", "interior", "video")     # = upload subfolders

_CONTENT_TYPE = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "webp": "image/webp", "mp4": "video/mp4", "mov": "video/quicktime",
}

DEFAULT_BUCKET = "car-media"


def normalize_registration(value: str) -> str:
    """Trim, uppercase, drop spaces — same spirit as the chatbot's reg key."""
    return (value or "").strip().upper().replace(" ", "")


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _valid_ext_for_type(media_type: str, ext: str) -> bool:
    if media_type in ("exterior", "interior"):
        return ext in PHOTO_EXT
    if media_type == "video":
        return ext in VIDEO_EXT
    return False


def content_type(ext: str) -> str:
    return _CONTENT_TYPE.get(ext, "application/octet-stream")


# ─────────────────────────────────────────────────────────────────────────────
# Storage uploader (pluggable)
# ─────────────────────────────────────────────────────────────────────────────
class StorageUploader(Protocol):
    def upload(self, bucket: str, object_path: str, local_path: str,
               ctype: str) -> str:
        """Upload a local file; return its public URL. Idempotent."""
        ...


class InMemoryStorageUploader:
    """Offline fake — stores bytes in memory, returns a Supabase-style public URL.
    Idempotent: re-uploading the same object_path is a no-op (skips re-upload)."""

    def __init__(self, project: str = "demo", fail_on: Optional[set] = None):
        self.project = project
        self.objects: Dict[str, bytes] = {}
        self.fail_on = fail_on or set()             # object_paths that should error

    def _url(self, bucket: str, object_path: str) -> str:
        return (f"https://{self.project}.supabase.co/storage/v1/object/public/"
                f"{bucket}/{object_path}")

    def upload(self, bucket: str, object_path: str, local_path: str,
               ctype: str) -> str:
        key = f"{bucket}/{object_path}"
        if object_path in self.fail_on or os.path.basename(local_path) in self.fail_on:
            raise RuntimeError("simulated storage upload failure")
        if key in self.objects:                     # idempotent: already uploaded
            return self._url(bucket, object_path)
        with open(local_path, "rb") as f:
            data = f.read()
        if not data:                                # guard: empty/corrupt file
            raise ValueError("empty file")
        self.objects[key] = data
        return self._url(bucket, object_path)


class SupabaseStorageUploader:
    """
    Reuses the legacy Storage upload + public-URL pattern (storageService.ts).
    Storage ONLY — never touches the cars table / sheets-sync / operational
    tables. Lazy `supabase` import so the POC runs without the dependency.
    """

    def __init__(self, url: str, key: str):
        from supabase import create_client                      # optional dep
        self._client = create_client(url, key)

    def upload(self, bucket: str, object_path: str, local_path: str,
               ctype: str) -> str:
        with open(local_path, "rb") as f:
            data = f.read()
        store = self._client.storage.from_(bucket)
        try:
            store.upload(object_path, data, {"content-type": ctype, "upsert": "true"})
        except Exception:
            # if it already exists (upsert disabled on the project), ignore and
            # fall through to the public URL — idempotent behaviour.
            pass
        return store.get_public_url(object_path)


# ─────────────────────────────────────────────────────────────────────────────
# Upload-folder scanning
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class UploadItem:
    registration: str       # folder name (raw)
    media_type: str         # exterior | interior | video
    filename: str
    path: str


def scan_uploads(uploads_dir: str) -> List[UploadItem]:
    """uploads/<REG>/<type>/<file>  →  list of UploadItem (sorted, deterministic)."""
    items: List[UploadItem] = []
    if not os.path.isdir(uploads_dir):
        return items
    for reg in sorted(os.listdir(uploads_dir)):
        reg_path = os.path.join(uploads_dir, reg)
        if not os.path.isdir(reg_path) or reg.startswith("_"):
            continue
        for mtype in sorted(os.listdir(reg_path)):
            type_path = os.path.join(reg_path, mtype)
            if not os.path.isdir(type_path) or mtype.lower() not in MEDIA_TYPES:
                continue
            for fname in sorted(os.listdir(type_path)):
                fpath = os.path.join(type_path, fname)
                if os.path.isfile(fpath):
                    items.append(UploadItem(reg, mtype.lower(), fname, fpath))
    return items


# ─────────────────────────────────────────────────────────────────────────────
# Workbook helpers
# ─────────────────────────────────────────────────────────────────────────────
_CAR_NUMB_HEADERS = {"CAR NUMB", "CAR NUMBER", "CAR_NUMB", "REGISTRATION", "REG NO"}


@dataclass
class SheetLayout:
    car_col: Optional[int]
    slots: Dict[str, List[int]]      # media_type -> ordered list of column indices


def read_layout(ws) -> SheetLayout:
    car_col = None
    slots = {"exterior": [], "interior": [], "video": []}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip()
        hu = h.upper()
        hl = h.lower()
        if hu in _CAR_NUMB_HEADERS:
            car_col = c
        elif hl.startswith("exterior"):
            slots["exterior"].append(c)
        elif hl.startswith("interior"):
            slots["interior"].append(c)
        elif hl.startswith("video"):
            slots["video"].append(c)
    return SheetLayout(car_col, slots)


def index_rows(ws, car_col: int) -> Dict[str, int]:
    rows: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, car_col).value
        if v:
            rows[normalize_registration(str(v))] = r
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Sync result types
# ─────────────────────────────────────────────────────────────────────────────
# statuses
ST_UPLOADED = "uploaded"
ST_DUPLICATE = "skipped_duplicate"
ST_UNKNOWN_VEHICLE = "unknown_vehicle"
ST_NO_SLOT = "no_slot"
ST_INVALID_FILE = "invalid_file"
ST_UPLOAD_ERROR = "upload_error"


@dataclass
class FileResult:
    filename: str
    registration: str
    media_type: str
    status: str
    url: Optional[str] = None
    detail: Optional[str] = None


@dataclass
class SyncReport:
    results: List[FileResult] = field(default_factory=list)
    saved: bool = False
    save_error: Optional[str] = None
    changed: bool = False

    def count(self, status: str) -> int:
        return sum(1 for r in self.results if r.status == status)

    @property
    def total(self) -> int:
        return len(self.results)

    @property
    def succeeded(self) -> int:
        return self.count(ST_UPLOADED)

    def summary(self) -> Dict[str, int]:
        out: Dict[str, int] = {}
        for r in self.results:
            out[r.status] = out.get(r.status, 0) + 1
        return out


# ─────────────────────────────────────────────────────────────────────────────
# The sync (the 7-step workflow)
# ─────────────────────────────────────────────────────────────────────────────
def run_sync(uploads_dir: str, workbook_path: str, *,
             uploader: StorageUploader, bucket: str = DEFAULT_BUCKET,
             save_fn: Optional[Callable] = None) -> SyncReport:
    """
    Step 1 scan → 2 identify reg → 3 upload → 4 public URL → 5 locate row →
    6 populate media columns → 7 save workbook.
    """
    save_fn = save_fn or (lambda wb, path: wb.save(path))
    report = SyncReport()

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    layout = read_layout(ws)
    if layout.car_col is None:
        report.save_error = "no CAR NUMB column in workbook"
        return report
    rows = index_rows(ws, layout.car_col)

    for item in scan_uploads(uploads_dir):                 # (1) scan
        nreg = normalize_registration(item.registration)   # (2) identify
        ext = _ext(item.filename)
        fr = FileResult(item.filename, nreg, item.media_type, ST_UPLOADED)

        # validate file type + content
        if not _valid_ext_for_type(item.media_type, ext):
            fr.status, fr.detail = ST_INVALID_FILE, f"bad extension '.{ext}' for {item.media_type}"
            report.results.append(fr); continue
        if not os.path.isfile(item.path) or os.path.getsize(item.path) == 0:
            fr.status, fr.detail = ST_INVALID_FILE, "missing or empty file"
            report.results.append(fr); continue

        # (5) locate matching row by registration
        if nreg not in rows:
            fr.status, fr.detail = ST_UNKNOWN_VEHICLE, "no matching CAR NUMB row"
            report.results.append(fr); continue
        row_idx = rows[nreg]
        slot_cols = layout.slots.get(item.media_type, [])
        if not slot_cols:
            fr.status, fr.detail = ST_NO_SLOT, f"no {item.media_type} columns in sheet"
            report.results.append(fr); continue

        # (3) upload  (deterministic path → idempotent)
        object_path = f"{nreg}/{item.media_type}/{item.filename}"
        try:
            url = uploader.upload(bucket, object_path, item.path, content_type(ext))
        except Exception as e:                              # upload failed
            fr.status, fr.detail = ST_UPLOAD_ERROR, str(e)
            report.results.append(fr); continue
        fr.url = url                                        # (4) public URL

        # (6) populate — duplicate-safe, next empty slot
        existing = [ws.cell(row_idx, c).value for c in slot_cols]
        if url in [v for v in existing if v]:
            fr.status, fr.detail = ST_DUPLICATE, "url already present"
            report.results.append(fr); continue
        empty = next((c for c in slot_cols if not ws.cell(row_idx, c).value), None)
        if empty is None:
            fr.status, fr.detail = ST_NO_SLOT, f"all {item.media_type} slots full"
            report.results.append(fr); continue
        ws.cell(row_idx, empty).value = url
        report.changed = True
        report.results.append(fr)

    # (7) save
    if report.changed:
        try:
            save_fn(wb, workbook_path)
            report.saved = True
        except Exception as e:
            report.saved = False
            report.save_error = str(e)
    else:
        report.saved = True                                 # nothing to persist
    return report


# ─────────────────────────────────────────────────────────────────────────────
# Test workbook generator (5 realistic vehicles)
# ─────────────────────────────────────────────────────────────────────────────
def create_test_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = (["CAR NUMB", "MODEL", "YEAR", "RATE"]
               + [f"Exterior_Photo_{i}" for i in range(1, 6)]
               + [f"Interior_Photo_{i}" for i in range(1, 6)]
               + [f"Video_{i}" for i in range(1, 4)])
    ws.append(headers)
    rows = [
        ["MH12AB1234", "Creta", 2019, 750000],
        ["MH14CD5678", "Swift", 2018, 550000],
        ["MH01EF9012", "Fortuner", 2017, 2200000],
        ["MH02GH3456", "Nexon", 2021, 900000],
        ["MH04IJ7890", "City", 2016, 600000],
    ]
    pad = len(headers) - 4
    for r in rows:
        ws.append(r + [None] * pad)
    wb.save(path)


if __name__ == "__main__":
    import tempfile, json
    d = tempfile.mkdtemp()
    wb_path = os.path.join(d, "test_inventory.xlsx")
    create_test_workbook(wb_path)
    up = os.path.join(d, "uploads", "MH12AB1234", "exterior")
    os.makedirs(up)
    for n in ("front.jpg", "side.jpg"):
        with open(os.path.join(up, n), "wb") as f:
            f.write(b"\xff\xd8\xff demo-jpeg")
    rep = run_sync(os.path.join(d, "uploads"), wb_path,
                   uploader=InMemoryStorageUploader())
    print("summary:", json.dumps(rep.summary()), "saved:", rep.saved)
    for r in rep.results:
        print(f"  [{r.status}] {r.registration}/{r.media_type}/{r.filename} -> {r.url}")
