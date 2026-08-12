"""
media_sync_poc_tests.py
======================

Validates the media-sync POC: folder parsing, registration matching, URL
generation, Excel updates, and failure cases. Runs fully offline (InMemory
storage). 20+ tests.

Run:  python media_sync_poc_tests.py
"""

import os
import tempfile
import unittest

import openpyxl
import media_sync_poc as m


def _mkfile(path: str, content: bytes = b"\xff\xd8\xff data"):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(content)


def _put(uploads: str, reg: str, mtype: str, fname: str, content: bytes = b"\xff\xd8\xff data"):
    _mkfile(os.path.join(uploads, reg, mtype, fname), content)


class Base(unittest.TestCase):
    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.uploads = os.path.join(self.dir, "uploads")
        self.wb = os.path.join(self.dir, "test_inventory.xlsx")
        m.create_test_workbook(self.wb)
        self.up = m.InMemoryStorageUploader()

    def sync(self, **kw):
        return m.run_sync(self.uploads, self.wb, uploader=self.up, **kw)

    def cells(self, reg):
        wb = openpyxl.load_workbook(self.wb)
        ws = wb.active
        layout = m.read_layout(ws)
        rows = m.index_rows(ws, layout.car_col)
        r = rows[m.normalize_registration(reg)]
        return {t: [ws.cell(r, c).value for c in cols] for t, cols in layout.slots.items()}, ws, r, layout


# ─────────────────────────────────────────────────────────────────────────────
# 1. Folder parsing
# ─────────────────────────────────────────────────────────────────────────────
class TestFolderParsing(Base):
    def test_scans_all_types(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        _put(self.uploads, "MH12AB1234", "interior", "b.png")
        _put(self.uploads, "MH12AB1234", "video", "c.mp4")
        items = m.scan_uploads(self.uploads)
        self.assertEqual(len(items), 3)
        self.assertEqual({i.media_type for i in items}, {"exterior", "interior", "video"})

    def test_ignores_unknown_subfolders(self):
        _put(self.uploads, "MH12AB1234", "documents", "rc.jpg")   # not a media type
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        items = m.scan_uploads(self.uploads)
        self.assertEqual([i.media_type for i in items], ["exterior"])

    def test_ignores_underscore_dirs(self):
        _put(self.uploads, "_processed", "exterior", "a.jpg")
        self.assertEqual(m.scan_uploads(self.uploads), [])

    def test_empty_uploads_dir(self):
        self.assertEqual(m.scan_uploads(os.path.join(self.dir, "nope")), [])


# ─────────────────────────────────────────────────────────────────────────────
# 2. Registration matching
# ─────────────────────────────────────────────────────────────────────────────
class TestRegistration(Base):
    def test_normalize(self):
        self.assertEqual(m.normalize_registration(" mh12ab1234 "), "MH12AB1234")
        self.assertEqual(m.normalize_registration("MH 12 AB 1234"), "MH12AB1234")

    def test_matched_vehicle_updates_row(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        rep = self.sync()
        self.assertEqual(rep.succeeded, 1)

    def test_case_and_space_insensitive_folder_matches(self):
        _put(self.uploads, "mh12ab1234", "exterior", "a.jpg")    # lowercase folder
        rep = self.sync()
        self.assertEqual(rep.succeeded, 1)                       # matches CAR NUMB

    def test_unknown_vehicle_safe(self):
        _put(self.uploads, "MH99ZZ0000", "exterior", "a.jpg")    # not in sheet
        rep = self.sync()
        self.assertEqual(rep.count(m.ST_UNKNOWN_VEHICLE), 1)
        self.assertEqual(rep.succeeded, 0)
        self.assertEqual(len(self.up.objects), 0)                # not uploaded


# ─────────────────────────────────────────────────────────────────────────────
# 3. URL generation + Storage
# ─────────────────────────────────────────────────────────────────────────────
class TestUrlAndStorage(Base):
    def test_public_url_format(self):
        _put(self.uploads, "MH12AB1234", "exterior", "front.jpg")
        rep = self.sync()
        url = rep.results[0].url
        self.assertTrue(url.startswith("https://"))
        self.assertIn("/storage/v1/object/public/car-media/MH12AB1234/exterior/front.jpg", url)

    def test_object_stored(self):
        _put(self.uploads, "MH12AB1234", "video", "v.mp4")
        self.sync()
        self.assertEqual(len(self.up.objects), 1)
        self.assertIn("car-media/MH12AB1234/video/v.mp4", self.up.objects)

    def test_content_type(self):
        self.assertEqual(m.content_type("jpg"), "image/jpeg")
        self.assertEqual(m.content_type("mp4"), "video/mp4")


# ─────────────────────────────────────────────────────────────────────────────
# 4. Excel updates
# ─────────────────────────────────────────────────────────────────────────────
class TestExcelUpdates(Base):
    def test_exterior_goes_to_first_exterior_slot(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.sync()
        cells, *_ = self.cells("MH12AB1234")
        self.assertTrue(cells["exterior"][0])                    # slot 1 filled
        self.assertIsNone(cells["exterior"][1])

    def test_interior_and_video_routing(self):
        _put(self.uploads, "MH12AB1234", "interior", "i.png")
        _put(self.uploads, "MH12AB1234", "video", "v.mp4")
        self.sync()
        cells, *_ = self.cells("MH12AB1234")
        self.assertTrue(cells["interior"][0])
        self.assertTrue(cells["video"][0])
        self.assertFalse(any(cells["exterior"]))

    def test_multiple_photos_fill_successive_slots(self):
        for n in ("a.jpg", "b.jpg", "c.jpg"):
            _put(self.uploads, "MH12AB1234", "exterior", n)
        self.sync()
        cells, *_ = self.cells("MH12AB1234")
        filled = [v for v in cells["exterior"] if v]
        self.assertEqual(len(filled), 3)
        self.assertEqual(len(set(filled)), 3)                    # distinct URLs

    def test_correct_row_only(self):
        _put(self.uploads, "MH01EF9012", "exterior", "a.jpg")    # Fortuner row
        self.sync()
        fort, *_ = self.cells("MH01EF9012")
        creta, *_ = self.cells("MH12AB1234")
        self.assertTrue(fort["exterior"][0])
        self.assertFalse(any(creta["exterior"]))                 # untouched

    def test_existing_data_preserved(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.sync()
        wb = openpyxl.load_workbook(self.wb)
        ws = wb.active
        # row 2 = MH12AB1234: MODEL/YEAR/RATE intact
        self.assertEqual(ws.cell(2, 2).value, "Creta")
        self.assertEqual(ws.cell(2, 3).value, 2019)
        self.assertEqual(ws.cell(2, 4).value, 750000)

    def test_preexisting_media_url_preserved(self):
        # pre-fill exterior slot 1 manually, then sync a new photo -> goes to slot 2
        wb = openpyxl.load_workbook(self.wb); ws = wb.active
        layout = m.read_layout(ws)
        ws.cell(2, layout.slots["exterior"][0]).value = "https://existing/url1.jpg"
        wb.save(self.wb)
        _put(self.uploads, "MH12AB1234", "exterior", "new.jpg")
        self.sync()
        cells, *_ = self.cells("MH12AB1234")
        self.assertEqual(cells["exterior"][0], "https://existing/url1.jpg")
        self.assertTrue(cells["exterior"][1])

    def test_saved_and_reloadable(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        rep = self.sync()
        self.assertTrue(rep.saved)
        cells, *_ = self.cells("MH12AB1234")                     # reload from disk
        self.assertTrue(cells["exterior"][0])


# ─────────────────────────────────────────────────────────────────────────────
# 5. Idempotency / duplicates
# ─────────────────────────────────────────────────────────────────────────────
class TestIdempotency(Base):
    def test_duplicate_upload_skipped(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.sync()
        rep2 = self.sync()                                       # run again
        self.assertEqual(rep2.count(m.ST_DUPLICATE), 1)
        self.assertEqual(rep2.succeeded, 0)

    def test_no_duplicate_cells_after_rerun(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.sync(); self.sync()
        cells, *_ = self.cells("MH12AB1234")
        filled = [v for v in cells["exterior"] if v]
        self.assertEqual(len(filled), 1)                         # exactly one

    def test_rerun_makes_no_changes(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")
        self.sync()
        rep2 = self.sync()
        self.assertFalse(rep2.changed)

    def test_no_slot_when_full(self):
        for n in ("a", "b", "c", "d", "e", "f"):                 # 6 photos, 5 slots
            _put(self.uploads, "MH12AB1234", "exterior", f"{n}.jpg")
        rep = self.sync()
        self.assertEqual(rep.succeeded, 5)
        self.assertEqual(rep.count(m.ST_NO_SLOT), 1)


# ─────────────────────────────────────────────────────────────────────────────
# 6. Failure handling
# ─────────────────────────────────────────────────────────────────────────────
class TestFailures(Base):
    def test_excel_write_fails_safely(self):
        _put(self.uploads, "MH12AB1234", "exterior", "a.jpg")

        def boom(wb, path):
            raise IOError("disk full (simulated)")

        rep = self.sync(save_fn=boom)
        self.assertFalse(rep.saved)
        self.assertIn("disk full", rep.save_error)
        self.assertEqual(len(self.up.objects), 1)                # upload DID happen
        # on-disk workbook unchanged (no partial/corrupt write)
        cells, *_ = self.cells("MH12AB1234")
        self.assertFalse(any(cells["exterior"]))

    def test_empty_file_rejected(self):
        _put(self.uploads, "MH12AB1234", "exterior", "empty.jpg", content=b"")
        rep = self.sync()
        self.assertEqual(rep.count(m.ST_INVALID_FILE), 1)
        self.assertEqual(rep.succeeded, 0)

    def test_invalid_extension_rejected(self):
        _put(self.uploads, "MH12AB1234", "exterior", "notes.txt")
        rep = self.sync()
        self.assertEqual(rep.count(m.ST_INVALID_FILE), 1)

    def test_video_file_in_photo_folder_rejected(self):
        _put(self.uploads, "MH12AB1234", "exterior", "clip.mp4")  # video in photo dir
        rep = self.sync()
        self.assertEqual(rep.count(m.ST_INVALID_FILE), 1)

    def test_upload_error_recorded(self):
        _put(self.uploads, "MH12AB1234", "video", "v.mp4")
        up = m.InMemoryStorageUploader(fail_on={"v.mp4"})
        rep = m.run_sync(self.uploads, self.wb, uploader=up)
        self.assertEqual(rep.count(m.ST_UPLOAD_ERROR), 1)
        self.assertEqual(rep.succeeded, 0)

    def test_mixed_batch_partial_success(self):
        _put(self.uploads, "MH12AB1234", "exterior", "good.jpg")   # ok
        _put(self.uploads, "MH99ZZ0000", "exterior", "x.jpg")      # unknown vehicle
        _put(self.uploads, "MH14CD5678", "exterior", "bad.txt")    # invalid
        rep = self.sync()
        self.assertEqual(rep.succeeded, 1)
        self.assertEqual(rep.count(m.ST_UNKNOWN_VEHICLE), 1)
        self.assertEqual(rep.count(m.ST_INVALID_FILE), 1)
        self.assertTrue(rep.saved)                                 # the good one saved


# ─────────────────────────────────────────────────────────────────────────────
# 7. End-to-end across the 5-vehicle workbook
# ─────────────────────────────────────────────────────────────────────────────
class TestEndToEnd(Base):
    def test_full_pipeline_five_vehicles(self):
        regs = ["MH12AB1234", "MH14CD5678", "MH01EF9012", "MH02GH3456", "MH04IJ7890"]
        for reg in regs:
            _put(self.uploads, reg, "exterior", "front.jpg")
            _put(self.uploads, reg, "video", "walk.mp4")
        rep = self.sync()
        self.assertEqual(rep.succeeded, 10)                        # 5x2
        self.assertTrue(rep.saved)
        for reg in regs:
            cells, *_ = self.cells(reg)
            self.assertTrue(cells["exterior"][0])
            self.assertTrue(cells["video"][0])


if __name__ == "__main__":
    unittest.main(verbosity=2)
